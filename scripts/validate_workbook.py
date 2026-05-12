import re, zipfile
from pathlib import Path
from openpyxl import load_workbook
ERR_PAT=["#REF!","#VALUE!","#N/A","#NAME?","#DIV/0!","#NULL!","#ССЫЛКА!","#ЗНАЧ!","#Н/Д","#ИМЯ?"]
CRITICAL=[('ScenarioText_UP','B3'),('ScenarioText_UP','B4'),('ScenarioText_UP','B5'),('ScenarioText_UP','B6'),('ScenarioText_UP','B31'),('ScenarioText_DOWN','B3'),('ScenarioText_DOWN','B4'),('ScenarioText_DOWN','B5'),('ScenarioText_DOWN','B6'),('ScenarioText_DOWN','B31'),('BasketSummary','B13'),('BasketSummary','C13')]

def check(path):
 wb=load_workbook(path,data_only=False)
 issues=[]
 for sh in wb.worksheets:
  for row in sh.iter_rows(min_row=1,max_row=sh.max_row,min_col=1,max_col=sh.max_column):
   for c in row:
    v=c.value
    if isinstance(v,str) and any(e in v for e in ERR_PAT): issues.append((sh.title,c.coordinate,'FORMULA_ERROR',v))
 for sh,cell in CRITICAL:
  if sh in wb.sheetnames:
   v=wb[sh][cell].value
   if v in (None,''): issues.append((sh,cell,'EMPTY_CRITICAL','empty'))
 if 'Settings' in wb.sheetnames:
  s=wb['Settings']
  if s['B31'].value in (None,''): issues.append(('Settings','B31','EMPTY_BID',''))
  if s['B32'].value in (None,''): issues.append(('Settings','B32','EMPTY_ASK',''))
 if 'Scenario_UP' in wb.sheetnames:
  su=wb['Scenario_UP']
  if su['B6'].value in (None,''): issues.append(('Scenario_UP','B6','EMPTY_SCENARIO_ASK',''))
  if su['B7'].value in (None,''): issues.append(('Scenario_UP','B7','EMPTY_SCENARIO_MID',''))
  if su['B8'].value in (None,''): issues.append(('Scenario_UP','B8','EMPTY_SCENARIO_SPREAD',''))
 if 'Scenario_DOWN' in wb.sheetnames:
  sd=wb['Scenario_DOWN']
  if sd['B6'].value in (None,''): issues.append(('Scenario_DOWN','B6','EMPTY_SCENARIO_ASK',''))
  if sd['B7'].value in (None,''): issues.append(('Scenario_DOWN','B7','EMPTY_SCENARIO_MID',''))
  if sd['B8'].value in (None,''): issues.append(('Scenario_DOWN','B8','EMPTY_SCENARIO_SPREAD',''))
 # xml grep
 with zipfile.ZipFile(path) as z:
  xml='\n'.join(z.read(n).decode('utf-8','ignore') for n in z.namelist() if n.startswith('xl/'))
  for e in ERR_PAT:
   if e in xml: issues.append(('XML','-', 'ERROR_TOKEN', e))
 print(path.name,'issues',len(issues))
 for i in issues[:50]: print(' -',i)
 return len(issues)

if __name__=='__main__':
 root=Path('recovery_lock_cascade_next_step')
 total=0
 for f in ['recovery_lock_cascade_next_step.xlsx','recovery_lock_cascade_next_step_ru.xlsx']:
  total+=check(root/f)
 print('TOTAL_ISSUES',total)
