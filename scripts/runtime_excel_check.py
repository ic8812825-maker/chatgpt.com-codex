import subprocess
from openpyxl import load_workbook

SRC='recovery_lock_cascade_next_step.xlsx'
TMP='reports/tests/recovery_lock_runtime_case.xlsx'

wb=load_workbook(SRC)
s=wb['Settings']
s['B11']=0; s['B12']=20; s['B13']=0; s['B14']=0; s['B19']='FULL_CYCLE'
tr=wb['TailRecovery_UP']
tr['B14']=0.14; tr['B15']=2; tr['B6']='NO'; tr['B7']=100; tr['B5']=400
sc=wb['SectionCalculator_UP']; sc['B12']='YES'; sc['B13']='YES'
bs=wb['BasketSummary']
bs['B3']=-12; bs['B4']=18
wb.save(TMP)

subprocess.run(['libreoffice','--headless','--convert-to','ods','--outdir','reports/tests',TMP],check=True)
subprocess.run(['libreoffice','--headless','--convert-to','xlsx','--outdir','reports/tests','reports/tests/recovery_lock_runtime_case.ods'],check=True)

wb2=load_workbook(TMP,data_only=True)
checks={
 'tail_ticket_up': ('Scenario_UP','B222'),
 'next_big': ('TailRecovery_UP','B16'),
 'next_small': ('TailRecovery_UP','B17'),
 'can_open_section_up': ('SectionCalculator_UP','B14'),
 'costs_p21': ('SectionCalculator_UP','P21'),
 'can_close_basket_up': ('BasketSummary','B10'),
 'close_raw': ('TailRecovery_UP','B8'),
 'close_final': ('TailRecovery_UP','B10'),
 'close_allowed': ('TailRecovery_UP','B11'),
}
for k,(sh,c) in checks.items():
    print(f'{k}: {wb2[sh][c].value}')
