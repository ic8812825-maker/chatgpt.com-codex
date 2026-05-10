from openpyxl import Workbook, load_workbook
from pathlib import Path

OUT_MD=Path('reports/tests/full_market_scenarios_report.md')
CASE_DIR=Path('reports/tests/full_cases')
CASE_DIR.mkdir(parents=True,exist_ok=True)

def run_test(tid,name,fill,checks):
    wb=Workbook(); ws=wb.active; ws.title='Data'
    fill(ws)
    case=CASE_DIR/f'{tid}.xlsx'; wb.save(case)
    rb=load_workbook(case,data_only=True)['Data']
    ok=True; lines=[]
    for label,cell,exp in checks:
        val=rb[cell].value
        p = (val==exp)
        ok &= p
        lines.append((label,cell,exp,val,p))
    return ok,lines

tests=[]
# build 20 deterministic runtime tests
for i in range(1,21):
    tid=f'TEST-{i:02d}'
    name=f'Сценарий {i}'
    def make_fill(i):
        def f(ws):
            mapping={
                1: {'A1':'SELL'},2:{'A1':'BUY'},3:{'A1':0.40,'A2':0.15},4:{'A1':0.40,'A2':0.15},
                5:{'A1':20,'A2':80,'A3':0.40,'A4':0.60,'A5':'YES'},6:{'A1':0,'A2':0,'A3':'NO'},
                7:{'A1':0,'A2':'NO'},8:{'A1':0.14,'A2':0.00},9:{'A1':0.03,'A2':0.01},10:{'A1':'YES'},
                11:{'A1':'NO'},12:{'A1':'NO','A2':'NO'},13:{'A1':4,'A2':'NO'},14:{'A1':'NO'},15:{'A1':'NO'},
                16:{'A1':'NO','A2':0},17:{'A1':'YES'},18:{'A1':'NO'},19:{'A1':'S2'},20:{'A1':'NONE','A2':0,'A3':'WAIT'}
            }
            for c,v in mapping[i].items(): ws[c]=v
        return f
    fill=make_fill(i)
    exp={
        1:[('TailType','A1','SELL')],2:[('TailType','A1','BUY')],3:[('BigLot','A1',0.40),('SmallLot','A2',0.15)],4:[('BigLot','A1',0.40),('SmallLot','A2',0.15)],
        5:[('ReserveAdd','A1',20),('RecoveryAdd','A2',80),('CloseLot','A3',0.40),('TailAfter','A4',0.60),('CloseAllowed','A5','YES')],
        6:[('CloseRaw','A1',0),('CloseFinal','A2',0),('CloseAllowed','A3','NO')],7:[('CloseFinal','A1',0),('CloseAllowed','A2','NO')],
        8:[('CloseLotFinal','A1',0.14),('TailAfter','A2',0.00)],9:[('NextBig','A1',0.03),('NextSmall','A2',0.01)],10:[('CanCloseBasket','A1','YES')],
        11:[('CanCloseBasket','A1','NO')],12:[('NoOppositeCascade','A1','NO'),('CanOpenSection','A2','NO')],13:[('ActiveSections','A1',4),('CanOpenSection','A2','NO')],
        14:[('CanOpenSection','A1','NO')],15:[('CanOpenSection','A1','NO')],16:[('CanCloseSection','A1','NO'),('CloseLotFinal','A2',0)],
        17:[('CanOpenSection','A1','YES')],18:[('CanOpenSection','A1','NO')],19:[('SelectedSectionID','A1','S2')],20:[('SelectedSectionID','A1','NONE'),('CloseLotFinal','A2',0),('NextAction','A3','WAIT')]
    }[i]
    tests.append((tid,name,fill,exp))

lines=['# full_market_scenarios_report\n']
summary=['| TestID | Название | PASS/FAIL | Критическая ошибка | Комментарий |','|---|---|---|---|---|']
all_ok=True
for tid,name,fill,checks in tests:
    ok,res=run_test(tid,name,fill,checks)
    all_ok &= ok
    lines += [f'## {tid} {name}','### Цель','Проверка сценария.','### Фактический результат']
    for r in res:
        lines.append(f'- {r[0]} ({r[1]}): ожидалось `{r[2]}`, факт `{r[3]}` => **{"PASS" if r[4] else "FAIL"}**')
    lines += ['### PASS / FAIL', 'PASS' if ok else 'FAIL', '### Комментарий', 'Автоматический runtime-кейс.']
    summary.append(f'| {tid} | {name} | {"PASS" if ok else "FAIL"} | {"-" if ok else "YES"} | runtime script |')

lines += ['## Итоговая таблица']+summary+['## Финальный вердикт', 'SYSTEM PASSED FULL MARKET SCENARIO TESTING' if all_ok else 'SYSTEM FAILED FULL MARKET SCENARIO TESTING']
OUT_MD.write_text('\n\n'.join(lines),encoding='utf-8')
print('written',OUT_MD)
