from __future__ import annotations
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

REF_RE = re.compile(r"(?:('(?:[^']|'')+'|[A-Za-zА-Яа-я0-9_ ]+)!)?(\$?[A-Z]{1,3}\$?\d+)")


def normalize_sheet(raw: str | None, current: str) -> str:
    if not raw:
        return current
    s = raw[:-1]
    if s.startswith("'") and s.endswith("'"):
        s = s[1:-1].replace("''", "'")
    return s


def build_graph(xlsx_path: Path):
    wb = load_workbook(xlsx_path, data_only=False)
    graph: dict[str, list[str]] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    node = f"{ws.title}!{cell.coordinate}"
                    deps = []
                    for m in REF_RE.finditer(cell.value):
                        dep_sheet = normalize_sheet(m.group(1), ws.title)
                        dep_cell = m.group(2).replace("$", "")
                        deps.append(f"{dep_sheet}!{dep_cell}")
                    graph[node] = deps
    return graph


def find_cycle(graph: dict[str, list[str]]):
    visited: set[str] = set()
    active: set[str] = set()
    parent: dict[str, str] = {}

    def dfs(node: str):
        visited.add(node)
        active.add(node)
        for dep in graph.get(node, []):
            if dep not in graph:
                continue
            if dep not in visited:
                parent[dep] = node
                cyc = dfs(dep)
                if cyc:
                    return cyc
            elif dep in active:
                cycle = [dep]
                cur = node
                while cur != dep:
                    cycle.append(cur)
                    cur = parent[cur]
                cycle.append(dep)
                cycle.reverse()
                return cycle
        active.remove(node)
        return None

    for n in graph:
        if n not in visited:
            parent[n] = n
            cyc = dfs(n)
            if cyc:
                return cyc
    return None


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("adaptive_ev_calculator.xlsx")
    if not path.exists():
        print(f"ERROR: workbook not found: {path}")
        return 2
    graph = build_graph(path)
    cyc = find_cycle(graph)
    if cyc:
        print("ERROR: Circular formula dependency found:")
        print(" -> ".join(cyc))
        return 1
    print(f"OK: no formula cycles found in {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
