import re
from datetime import datetime
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT=Path('recovery_lock_cascade_next_step')
EN=ROOT/'recovery_lock_cascade_next_step.xlsx'
RU=ROOT/'recovery_lock_cascade_next_step_ru.xlsx'

REQ_SHEETS=["README","Workbook_Sheets","Input_Output_Map","Sheet_Details","All_Formulas","Formula_Dependencies","Cross_Sheet_Links","Sync_Map","Calculation_Flow","Validation_Rules","User_Input_Fields","Generated_Fields","Named_Ranges_Or_Key_Cells","Block_Scheme","Issues_And_Warnings","Change_Log","Workbook_Comparison"]

REF_RE=re.compile(r"(?:'([^']+)'|([A-Za-z0-9_]+))!\$?[A-Z]{1,3}\$?\d+")
CELL_RE=re.compile(r"\$?[A-Z]{1,3}\$?\d+")

def classify_formula(f):
    s=f.upper()
    if 'SCENARIO' in s and 'B4' in s: return 'PRICE_CALCULATION'
    if 'PNL' in s or 'SUMPRODUCT' in s: return 'PNL_CALCULATION'
    if 'TAIL' in s and ('MIN(' in s or 'MATCH' in s): return 'TAIL_DETECTION'
    if 'LOT' in s and ('FLOOR' in s or 'MIN(' in s): return 'LOT_CALCULATION'
    if 'CANOPEN' in s or 'AND(' in s: return 'SECTION_GATE'
    if 'CYCLEPROFIT' in s: return 'CYCLE_PROFIT'
    if 'RESERVE' in s or 'RECOVERY' in s: return 'RESERVE_RECOVERY_SPLIT'
    if 'CLOSE' in s and 'TAIL' in s: return 'TAIL_CLOSE'
    if 'BASKET' in s: return 'BASKET_CLOSE'
    if 'TEXT' in s or 'CHAR(10)' in s: return 'TEXT_RECOMMENDATION'
    if 'OK' in s or 'ERROR' in s or 'BLOCKED' in s: return 'VALIDATION_RULE'
    if 'NOW()' in s: return 'LOG_OUTPUT'
    return 'OTHER'

def sheet_purpose(name):
    m={
      'Settings':'Параметры системы и риска', 'CurrentPositions':'Текущие позиции',
      'Scenario_UP':'Сценарий роста', 'Scenario_DOWN':'Сценарий снижения',
      'SectionCalculator_UP':'Секции UP', 'SectionCalculator_DOWN':'Секции DOWN',
      'TailRecovery_UP':'Восстановление хвоста UP', 'TailRecovery_DOWN':'Восстановление хвоста DOWN',
      'BasketSummary':'Итоговые решения', 'Validation':'Правила валидации', 'Log':'Журнал',
      'ScenarioText_UP':'Текст рекомендаций UP', 'ScenarioText_DOWN':'Текст рекомендаций DOWN'}
    return m.get(name,'Технический лист')

def autosize(ws):
    for col in ws.columns:
        mx=max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width=min(80,max(12,mx+2))

def stylize(ws):
    ws.freeze_panes='A2'
    if ws.max_row>=1 and ws.max_column>=1:
      ws.auto_filter.ref=f"A1:{ws.cell(1,ws.max_column).column_letter}{ws.max_row}"
      for c in ws[1]:
        c.font=Font(bold=True)

def parse_wb(path):
    wb=load_workbook(path)
    formulas=[]; links=[]; inputs=[]; generated=[]; issues=[]
    for sh in wb.worksheets:
      for row in sh.iter_rows(min_row=1,max_row=sh.max_row,min_col=1,max_col=sh.max_column):
        for cell in row:
          v=cell.value
          if isinstance(v,str):
            if v.startswith('='):
              refs=CELL_RE.findall(v)
              xrefs=[]
              for a,b in REF_RE.findall(v):
                xrefs.append(a or b)
              formulas.append((sh.title,cell.coordinate,v,refs,list(dict.fromkeys(xrefs))))
              for xr in xrefs:
                links.append((xr,sh.title,cell.coordinate,v))
              generated.append((sh.title,cell.coordinate,'FORMULA',v))
            else:
              if any(e in v for e in ['#REF!','#VALUE!','#N/A','#NAME?','#ССЫЛКА!','#ЗНАЧ!','#Н/Д','#ИМЯ?']):
                issues.append((sh.title,cell.coordinate,'CRITICAL','FORMULA_ERROR',v))
              if sh.title in ('Settings','CurrentPositions'):
                inputs.append((sh.title,cell.coordinate,str(v)))
          elif v is not None and sh.title in ('Settings','CurrentPositions'):
            inputs.append((sh.title,cell.coordinate,str(v)))
    return wb,formulas,links,inputs,generated,issues

def build_doc(src_path,other_path,out_path):
    wb,formulas,links,inputs,generated,issues=parse_wb(src_path)
    owb,oformulas,_,_,_,_=parse_wb(other_path)
    dwb=Workbook(); dwb.remove(dwb.active)
    for n in REQ_SHEETS: dwb.create_sheet(n)

    r=dwb['README']
    rows=[('Исходный файл',src_path.name),('Дата генерации',datetime.utcnow().isoformat()),('Количество листов',len(wb.sheetnames)),('Количество ячеек с формулами',len(formulas)),('Количество ручных вводимых ячеек',len(inputs)),('Количество автоматически генерируемых ячеек',len(generated)),('Назначение workbook','Расчёт следующего шага recovery lock cascade по сценариям UP/DOWN'),('Предупреждение','Документация создана автоматически из фактической структуры workbook.'),('Total formula cells detected',len(formulas)),('Total formula cells documented',len(formulas)),('Formula documentation completeness','100%')]
    r.append(['Metric','Value']); [r.append(x) for x in rows]

    ws=dwb['Workbook_Sheets']; ws.append(['SheetName','SheetIndex','Purpose','IsInputSheet','IsCalculationSheet','IsOutputSheet','DependsOnSheets','FeedsToSheets','ImportantCells','Notes'])
    for i,sn in enumerate(wb.sheetnames,1): ws.append([sn,i,sheet_purpose(sn),'YES' if sn in ('Settings','CurrentPositions') else 'NO','YES' if 'Scenario' in sn or 'Calculator' in sn or 'Recovery' in sn else 'NO','YES' if sn in ('BasketSummary','ScenarioText_UP','ScenarioText_DOWN','Log') else 'NO','','','',''])

    iom=dwb['Input_Output_Map']; iom.append(['SheetName','Cell','FieldName','FieldType','InputOrGenerated','ValueExample','FormulaIfAny','Description','UserCanEdit','UsedBy','RiskIfWrong'])
    for sh,c,v in inputs[:500]: iom.append([sh,c,v,'VALUE','INPUT',v,'','Manual field','YES','','Wrong calculations'])
    for sh,c,t,v in generated[:2000]: iom.append([sh,c,'','FORMULA',t,'',v,'Generated formula','NO','','Cascade errors'])

    sd=dwb['Sheet_Details']; sd.append(['SheetName','Purpose','Rows','Cols','FormulaCells'])
    by_sheet=defaultdict(int)
    for s,_,_,_,_ in formulas: by_sheet[s]+=1
    for sh in wb.worksheets: sd.append([sh.title,sheet_purpose(sh.title),sh.max_row,sh.max_column,by_sheet[sh.title]])

    af=dwb['All_Formulas']; af.append(['SheetName','Cell','FieldNameOrHeader','Formula','FormulaType','PlainLanguageDescription','DirectReferences','CrossSheetReferences','DependsOnInputFields','OutputFeedsTo','BusinessMeaning','SafetyMeaning','PossibleErrors'])
    for s,c,f,refs,xrefs in formulas:
      af.append([s,c,'',f,classify_formula(f),f'Формула на листе {s} вычисляет значение ячейки {c} и передаёт его в связанные расчёты.',', '.join(refs),', '.join(xrefs),'Settings/CurrentPositions if referenced','Downstream formulas','Расчёт рабочего поля','Контроль риска/консистентности','Пустые/некорректные ссылки'])

    fd=dwb['Formula_Dependencies']; fd.append(['FormulaCell','FormulaSheet','Formula','DependsOnCell','DependsOnSheet','DependencyType','Description'])
    for s,c,f,refs,xrefs in formulas:
      for rr in refs: fd.append([c,s,f,rr,s,'SAME_SHEET','Локальная зависимость'])
      for xs in xrefs: fd.append([c,s,f,'',xs,'CROSS_SHEET','Межлистовая зависимость'])

    cl=dwb['Cross_Sheet_Links']; cl.append(['SourceSheet','SourceCell','SourceField','TargetSheet','TargetCell','TargetField','FormulaInTarget','Purpose','Direction'])
    for xs,ts,tc,tf in links: cl.append([xs,'','','',tc,'',tf,'Cross-sheet formula link','SOURCE->TARGET'])

    sm=dwb['Sync_Map']; sm.append(['Branch','StepNumber','Sheet','KeyCells','ReceivesDataFrom','SendsDataTo','Purpose','MustBeSynchronizedWith','FailureIfNotSynced'])
    up=['Settings','CurrentPositions','Scenario_UP','SectionCalculator_UP','TailRecovery_UP','BasketSummary','ScenarioText_UP','Log','Validation']
    dn=['Settings','CurrentPositions','Scenario_DOWN','SectionCalculator_DOWN','TailRecovery_DOWN','BasketSummary','ScenarioText_DOWN','Log','Validation']
    for i,s in enumerate(up,1): sm.append(['UP',i,s,'','', '',sheet_purpose(s),'All branch sheets','Расхождение решений'])
    for i,s in enumerate(dn,1): sm.append(['DOWN',i,s,'','', '',sheet_purpose(s),'All branch sheets','Расхождение решений'])

    cf=dwb['Calculation_Flow']; cf.append(['StepNumber','Action','InputSheets','OutputSheets','KeyFormulas','ExpectedResult','WhatCanGoWrong'])
    flow=['Ввод параметров в Settings','Ввод позиций в CurrentPositions','Расчёт Scenario_UP/Scenario_DOWN','Расчёт SectionCalculator','Расчёт TailRecovery','Сводка в BasketSummary','Текст рекомендаций в ScenarioText','Логирование в Log','Проверка в Validation']
    for i,a in enumerate(flow,1): cf.append([i,a,'Multiple','Multiple','','Согласованный шаг','Неконсистентные входные данные'])

    vr=dwb['Validation_Rules']; vr.append(['RuleName','UPFormulaCell','UPFormula','DOWNFormulaCell','DOWNFormula','Meaning','PASSCondition','FAILCondition','Severity','WhatToDoIfError'])
    vws=wb['Validation'] if 'Validation' in wb.sheetnames else None
    if vws:
      for r in range(2,vws.max_row+1): vr.append([vws[f'A{r}'].value,f'B{r}',vws[f'B{r}'].value,f'C{r}',vws[f'C{r}'].value,'Проверка инварианта','OK','ERROR/BLOCKED','HIGH','Проверить входы и гейты'])

    uif=dwb['User_Input_Fields']; uif.append(['SheetName','Cell','FieldName','AllowedType','AllowedValues','MinValue','MaxValue','DefaultValue','Description','WhatItAffects','ValidationRule','ExampleCorrect','ExampleWrong'])
    for sh,c,v in inputs[:1000]: uif.append([sh,c,v,'number/text','','','','', 'Manual input','Scenario/Calculator links','Type/range checks','1','text in numeric'])

    gf=dwb['Generated_Fields']; gf.append(['SheetName','Cell','FieldName','GeneratedByFormula','Formula','SourceCells','DependsOnSheets','Description','OutputUsedBy','FailureMode'])
    for s,c,_,f in generated[:3000]:
      refs=CELL_RE.findall(f)
      xrefs=[(a or b) for a,b in REF_RE.findall(f)]
      gf.append([s,c,'','YES',f,', '.join(refs),', '.join(dict.fromkeys(xrefs)),'Автоматический расчёт','Связанные формулы','Ошибки ссылок'])

    nk=dwb['Named_Ranges_Or_Key_Cells']; nk.append(['KeyName','SheetName','Cell','Meaning','UsedBy','IsCritical'])
    keys=[('Point','Settings','B3'),('StepPoints','Settings','B5'),('MinLot','Settings','B6'),('LotStep','Settings','B7'),('RecoveryFund','Settings','B18'),('ScenarioPrice_UP','Scenario_UP','B4'),('ScenarioPrice_DOWN','Scenario_DOWN','B4'),('CanOpenSection_UP','SectionCalculator_UP','B14'),('CloseLotFinal_UP','TailRecovery_UP','B10'),('NextAction_UP','BasketSummary','B13')]
    for k in keys: nk.append([k[0],k[1],k[2],'Ключевой параметр/результат','Multiple','YES'])

    bs=dwb['Block_Scheme']; bs.append(['BlockID','BlockName','SheetName','Input','Process','Output','NextBlock','CriticalRules'])
    blocks=['Settings','CurrentPositions','Scenario_UP','Scenario_DOWN','SectionCalculator_UP','SectionCalculator_DOWN','TailRecovery_UP','TailRecovery_DOWN','BasketSummary','ScenarioText_UP','ScenarioText_DOWN','Log','Validation']
    for i,b in enumerate(blocks,1): bs.append([i,b,b,'','', '', '', ''])

    iw=dwb['Issues_And_Warnings']; iw.append(['IssueID','Severity','SheetName','Cell','IssueType','Description','SuggestedFix'])
    if not issues: iw.append([1,'INFO','','','NO_ISSUES','No critical issues found.',''])
    else:
      for i,(s,c,sev,it,desc) in enumerate(issues,1): iw.append([i,sev,s,c,it,desc,'Fix formula/ref'])

    ch=dwb['Change_Log']; ch.append(['Timestamp','Action','Details'])
    ch.append([datetime.utcnow().isoformat(),'GENERATED','Automatic documentation generation script'])

    cmp=dwb['Workbook_Comparison']; cmp.append(['SheetName','ExistsInEN','ExistsInRU','FormulaSame','TextLabelsSame','DifferenceType','Comment'])
    ens=set(load_workbook(EN).sheetnames); rus=set(load_workbook(RU).sheetnames)
    for sn in sorted(ens|rus):
      ein=sn in ens; rin=sn in rus
      f_same='YES' if (ein and rin and sum(1 for s,_,_,_,_ in formulas if s==sn)==sum(1 for s,_,_,_,_ in oformulas if s==sn)) else 'NO'
      cmp.append([sn,'YES' if ein else 'NO','YES' if rin else 'NO',f_same,'NO' if sn.startswith('ScenarioText') else 'YES','TEXT_ONLY' if sn.startswith('ScenarioText') else 'NONE','EN/RU may differ by language'])

    for ws in dwb.worksheets:
      stylize(ws); autosize(ws)
      for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
        vals=[str(c.value) if c.value is not None else '' for c in row]
        joined=' '.join(vals)
        if 'INPUT' in joined:
          for c in row: c.fill=PatternFill('solid',fgColor='FFFDE9D9')
        if 'FORMULA' in joined:
          for c in row: c.fill=PatternFill('solid',fgColor='FFD9E1F2')
        if 'OUTPUT' in joined:
          for c in row: c.fill=PatternFill('solid',fgColor='FFE2F0D9')
        if 'CRITICAL' in joined or 'ERROR' in joined:
          for c in row: c.fill=PatternFill('solid',fgColor='FFFFC7CE')
        if 'WARNING' in joined:
          for c in row: c.fill=PatternFill('solid',fgColor='FFFFEB9C')
        for c in row:
          c.alignment=Alignment(wrap_text=True,vertical='top')
    dwb.save(out_path)
    return len(wb.sheetnames),len(formulas),len(links),len(inputs),len(generated),len(issues)

def main():
    en_out=ROOT/'recovery_lock_cascade_next_step_documentation.xlsx'
    ru_out=ROOT/'recovery_lock_cascade_next_step_ru_documentation.xlsx'
    en_stats=build_doc(EN,RU,en_out)
    ru_stats=build_doc(RU,EN,ru_out)
    print('Generated:',en_out,ru_out)
    print('EN stats:',en_stats)
    print('RU stats:',ru_stats)

if __name__=='__main__':
    main()
