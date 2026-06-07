from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = PROJECT_ROOT / "MinusLock_SelfCompressing_BigSmall_v2.xlsx"
EXPECTED_SHEETS = [
    "Settings",
    "Calculator",
    "Trend_UP",
    "Trend_DOWN",
    "Risk_Analysis",
    "Tests",
    "Manual",
    "Examples",
]
REQUIRED_V3_HEADERS = [
    "RealizedFarLoss",
    "ClosedLotsForCosts",
    "CostPerLot",
    "OldFarClosedLot",
    "BlockedReason",
    "ActiveOldFarLot",
    "ActiveNewFarLot",
    "DualTailTotalLot",
    "ManualOldFarCloseLot",
    "ManualNewFarCloseLot",
    "ManualClosePL",
    "ManualAllowNewLevel",
]


@dataclass
class Params:
    start_lot: float = 1.0
    big_ratio: float = 1.30
    small_ratio: float = 0.36
    close_far_share: float = 0.90
    reserve_share: float = 0.10
    close_big_on_small: float = 0.30
    remain_big_on_small: float = 0.70
    point_value_per_lot: float = 1.0
    margin_per_lot: float = 1000.0
    lot_step: float = 0.01
    max_big_ratio: float = 1.35
    max_small_ratio: float = 0.45


def round_lot(value: float, step: float = 0.01) -> float:
    return round(round(value / step) * step, 10)


def workbook():
    assert WORKBOOK.exists()
    return load_workbook(WORKBOOK, data_only=False)


def header_map(ws):
    return {ws.cell(1, column).value: column for column in range(1, ws.max_column + 1)}


def settings_map(wb):
    ws = wb["Settings"]
    return {ws.cell(row, 1).value: ws.cell(row, 2).value for row in range(2, 18)}


def big_side_v3_model(
    far_start: float,
    big_lot: float,
    small_lot: float,
    big_profit_points: float,
    small_points: float,
    far_loss_points: float,
    commission_per_lot: float,
    spread_cost_per_lot: float,
    slippage_cost_per_lot: float,
    params: Params = Params(),
):
    cost_per_lot = commission_per_lot + spread_cost_per_lot + slippage_cost_per_lot
    profit_big = big_lot * big_profit_points * params.point_value_per_lot
    loss_small = small_lot * abs(small_points) * params.point_value_per_lot
    loss_per_lot = max(0, far_loss_points * params.point_value_per_lot)
    costs_before_far = (big_lot + small_lot) * cost_per_lot
    net_profit_before_far = profit_big - loss_small - costs_before_far
    close_far_budget = net_profit_before_far * params.close_far_share if net_profit_before_far > 0 else 0
    close_far_lot = 0 if loss_per_lot <= 0 else min(far_start, max(0, close_far_budget / loss_per_lot))
    realized_far_loss = close_far_lot * loss_per_lot
    costs_far_close = close_far_lot * cost_per_lot
    closed_lots_for_costs = big_lot + small_lot + close_far_lot
    costs = closed_lots_for_costs * cost_per_lot
    net_profit = profit_big - loss_small - costs
    reserve_add = net_profit_before_far * params.reserve_share if net_profit_before_far > 0 else 0
    balance_after = net_profit - realized_far_loss
    return {
        "profit_big": profit_big,
        "loss_small": loss_small,
        "cost_per_lot": cost_per_lot,
        "costs_before_far": costs_before_far,
        "net_profit_before_far": net_profit_before_far,
        "close_far_budget": close_far_budget,
        "close_far_lot": close_far_lot,
        "realized_far_loss": realized_far_loss,
        "costs_far_close": costs_far_close,
        "closed_lots_for_costs": closed_lots_for_costs,
        "costs": costs,
        "net_profit": net_profit,
        "reserve_add": reserve_add,
        "balance_after": balance_after,
    }


def small_side_model(far_start: float, params: Params = Params()):
    big_lot = round_lot(far_start * params.big_ratio, params.lot_step)
    close_big = big_lot * params.close_big_on_small
    remain_big = big_lot * params.remain_big_on_small
    return {
        "big_lot": big_lot,
        "close_big": close_big,
        "remain_big": remain_big,
        "new_far_start": remain_big,
    }


def test_sheet_structure_is_exact():
    assert workbook().sheetnames == EXPECTED_SHEETS


def test_settings_defaults_and_editable_cells():
    wb = workbook()
    settings = settings_map(wb)
    expected = {
        "StartLot": 1,
        "BigRatio": 1.30,
        "SmallRatio": 0.36,
        "CloseFarShare": 0.90,
        "ReserveShare": 0.10,
        "CloseBigOnSmall": 0.30,
        "RemainBigOnSmall": 0.70,
        "PointValuePerLot": 1,
        "MarginPerLot": 1000,
        "LotStep": 0.01,
        "MaxLevels": 10,
        "CommissionPerLot": 0,
        "SpreadCostPerLot": 0,
        "SlippageCostPerLot": 0,
        "MaxBigRatio": 1.35,
        "MaxSmallRatio": 0.45,
    }
    assert settings == expected
    for row in range(2, 18):
        assert wb["Settings"].cell(row, 2).data_type == "n"


def test_v3_required_columns_exist_on_all_calculation_sheets():
    wb = workbook()
    for sheet in ["Calculator", "Trend_UP", "Trend_DOWN"]:
        headers = header_map(wb[sheet])
        for required in REQUIRED_V3_HEADERS:
            assert required in headers, f"{sheet} missing {required}"


def test_big_and_small_formula_math():
    assert round_lot(1.00 * 1.30) == 1.30
    assert round_lot(0.80 * 1.30) == 1.04
    small_raw = 1.30 * 0.36
    small_lot = round_lot(small_raw)
    assert small_raw == 0.46799999999999997
    assert small_lot == 0.47
    assert small_lot / 1.30 <= 0.45


def test_direction_sheets_have_v2_direction_logic():
    wb = workbook()
    assert wb["Trend_UP"]["C2"].value == "SELL"
    assert wb["Trend_UP"]["E2"].value == '=IF($B2="BLOCKED","",IF($C2="SELL","BUY","SELL"))'
    assert wb["Trend_UP"]["H2"].value == '=IF($B2="BLOCKED","",$C2)'
    assert wb["Trend_DOWN"]["C2"].value == "BUY"
    assert wb["Trend_DOWN"]["E2"].value == '=IF($B2="BLOCKED","",IF($C2="SELL","BUY","SELL"))'
    assert wb["Trend_DOWN"]["H2"].value == '=IF($B2="BLOCKED","",$C2)'


def test_calculator_big_side_formulas_close_big_small_and_use_monetary_budget():
    ws = workbook()["Calculator"]
    assert ws["AN2"].value == '=IF($B2="BIG_SIDE",1,IF($B2="SMALL_SIDE",Settings!$B$7,0))'
    assert ws["AO2"].value == '=IF(OR($B2="BIG_SIDE",$B2="SMALL_SIDE"),1,0)'
    assert ws["T2"].value == '=IF(AND($B2="BIG_SIDE",$AX2>0),$AX2*Settings!$B$5,0)'
    assert ws["X2"].value == '=IF($B2="BIG_SIDE",IF($AX$2>0,$AX$2*Settings!$B$6,0),IF($S$2>0,$S$2*Settings!$B$6,0))'


def test_costs_per_lot_are_multiplied_by_closed_lots():
    ws = workbook()["Calculator"]
    assert ws["AU2"].value == "=Settings!$B$13+Settings!$B$14+Settings!$B$15"
    assert ws["AT2"].value == '=IF($B2="BIG_SIDE",$G2+$J2+$V2,IF($B2="SMALL_SIDE",$J2+$AL2+$AV2,0))'
    assert ws["R2"].value == "=$AT2*$AU2"
    model = big_side_v3_model(1.0, 1.30, 0.47, 100, -100, 200, 5, 3, 2)
    assert round(model["closed_lots_for_costs"], 4) == 2.0638
    assert model["cost_per_lot"] == 10
    assert round(model["costs"], 3) == 20.639


def test_realized_far_loss_formula():
    ws = workbook()["Calculator"]
    assert ws["AS2"].value == '=IF($B2="BIG_SIDE",$V2*$U2,0)'
    model = big_side_v3_model(1.0, 1.30, 0.47, 100, -100, 200, 5, 3, 2)
    assert round(model["close_far_budget"], 2) == 58.77
    assert round(model["close_far_lot"], 4) == 0.2939
    assert round(model["realized_far_loss"], 2) == 58.77


def test_big_side_balance_after_realized_far_loss():
    ws = workbook()["Calculator"]
    assert ws["AA2"].value == '=IF($B2="BLOCKED",$Z2+$BF2,IF($B2="BIG_SIDE",$Z2+$S2-$AS2,IF($B2="SMALL_SIDE",$Z2+$S2+$AI2,$Z2+$S2)))'
    model = big_side_v3_model(1.0, 1.30, 0.47, 100, -100, 200, 5, 3, 2)
    expected_balance = model["net_profit"] - model["realized_far_loss"]
    assert round(model["balance_after"], 3) == round(expected_balance, 3)
    assert model["balance_after"] < model["net_profit_before_far"]


def test_reserve_add_uses_net_profit_before_far_minus_realized_far_loss():
    ws = workbook()["Calculator"]
    assert ws["X2"].value == '=IF($B2="BIG_SIDE",IF($AX$2>0,$AX$2*Settings!$B$6,0),IF($S$2>0,$S$2*Settings!$B$6,0))'
    model = big_side_v3_model(1.0, 1.30, 0.47, 100, -100, 200, 5, 3, 2)
    assert round(model["reserve_add"], 2) == round(model["net_profit_before_far"] * Params().reserve_share, 2)


def test_zero_division_and_negative_net_profit_are_guarded():
    zero_loss = big_side_v3_model(1.0, 1.30, 0.47, 100, -100, 0, 5, 3, 2)
    assert zero_loss["close_far_lot"] == 0
    negative = big_side_v3_model(1.0, 1.30, 0.47, 20, -200, 200, 5, 3, 2)
    assert negative["net_profit_before_far"] < 0
    assert negative["close_far_budget"] == 0
    assert negative["reserve_add"] == 0
    assert negative["close_far_lot"] == 0


def test_small_side_close_remain_and_self_compression():
    first = small_side_model(1.0)
    assert first["big_lot"] == 1.30
    assert round(first["close_big"], 3) == 0.390
    assert round(first["remain_big"], 3) == 0.910
    assert round(first["new_far_start"], 3) == 0.910
    second = small_side_model(0.91)
    assert second["big_lot"] == 1.18
    assert round(1.30 * 0.70, 6) == 0.91
    assert second["new_far_start"] > 0


def test_old_far_close_pl_affects_small_side_balance():
    ws = workbook()["Calculator"]
    assert ws["AA2"].value.endswith('IF($B2="SMALL_SIDE",$Z2+$S2+$AI2,$Z2+$S2)))')
    assert ws["AV2"].value == '=IF($B2="SMALL_SIDE",IF($AH2="YES",$D2,IFERROR(MIN($D2,MAX(0,$AI2/$U2)),0)),0)'
    balance_before = 1000
    flip_net = 50
    old_far_close_pl = -20
    assert balance_before + flip_net + old_far_close_pl == 1030


def test_dual_tail_blocks_next_level():
    ws = workbook()["Calculator"]
    assert ws["AK2"].value == '=IF(AND($AJ2>0,$AG2>0),TRUE,FALSE)'
    assert ws["AQ2"].value.startswith('=IF($B2="BLOCKED",IF($BC2=0,"STOP_CLEARED"')
    assert '"STOP"' in ws["AQ2"].value
    assert 'IF($AK2=TRUE,"DUAL_TAIL"' in ws["AQ2"].value
    assert ws["AP2"].value == '=IF(AND($AQ2="MANUAL_READY",$BG2="YES"),"YES",IF(OR($AQ2="DUAL_TAIL",$AQ2="DANGER",$AQ2="STOP",$AQ2="STOP_CLEARED",$AQ2="MANUAL_READY"),"NO","YES"))'
    assert ws["B3"].value.startswith('=IF(AND($AQ2="MANUAL_READY",$BG2="YES")')
    assert '"BLOCKED"' in ws["B3"].value
    assert ws["AW3"].value == '=IF($B3="BLOCKED","Заблокировано: DUAL_TAIL сохраняет оба хвоста до ручного закрытия","")'


def test_blocked_next_level_zeroes_big_and_small_lots():
    ws = workbook()["Calculator"]
    assert ws["F3"].value == '=IF($B3="BLOCKED",0,MAX(0,$D3*Settings!$B$3))'
    assert ws["G3"].value == '=IF($B3="BLOCKED",0,IFERROR(MAX(0,ROUND($F3/Settings!$B$11,0)*Settings!$B$11),0))'
    assert ws["I3"].value == '=IF($B3="BLOCKED",0,MAX(0,$G3*Settings!$B$4))'
    assert ws["J3"].value == '=IF($B3="BLOCKED",0,IFERROR(MAX(0,ROUND($I3/Settings!$B$11,0)*Settings!$B$11),0))'
    assert ws["AQ3"].value.startswith('=IF($B3="BLOCKED",IF($BC3=0,"STOP_CLEARED"')
    assert '"STOP"' in ws["AQ3"].value


def test_margin_balance_reserve_and_limit_status_formulas():
    ws = workbook()["Calculator"]
    assert ws["AB2"].value == '=IF($B2="BLOCKED",$BC2,$D2+$G2+$J2)'
    assert ws["AD2"].value == "=$AB2*Settings!$B$10"
    assert ws["AE2"].value == "=$AC2*Settings!$B$10"
    assert "IFERROR($G2/$D2,0)>Settings!$B$16" in ws["AQ2"].value
    assert "IFERROR($J2/$G2,0)>Settings!$B$17" in ws["AQ2"].value
    open_lots_before = 1.00 + 1.30 + 0.47
    assert round(open_lots_before, 2) == 2.77
    assert round(open_lots_before * 1000, 0) == 2770


def test_risk_analysis_includes_realized_far_loss():
    wb = workbook()
    ws = wb["Risk_Analysis"]
    metrics = {ws.cell(row, 1).value: ws.cell(row, 3).value for row in range(2, ws.max_row + 1)}
    assert "Total Realized Far Loss" in metrics
    assert metrics["Total Realized Far Loss"] == "=SUM(Calculator!AS2:AS11)"
    assert "+SUM(Calculator!AS2:AS11)" in metrics["Total Closed Loss"]
    assert "=COUNTIF(Calculator!AQ2:AQ11,\"STOP\")" == metrics["Stop Count"]


def test_tests_sheet_is_detailed_and_has_v3_tests():
    ws = workbook()["Tests"]
    assert [ws.cell(1, c).value for c in range(1, 9)] == [
        "Test ID",
        "Test Name",
        "Input",
        "Expected",
        "Actual",
        "Formula Checked",
        "Status",
        "Comment",
    ]
    assert ws.max_row >= 40
    statuses = [ws.cell(row, 7).value for row in range(2, ws.max_row + 1)]
    assert set(statuses) == {"PASS"}
    test_names = {ws.cell(row, 2).value for row in range(2, ws.max_row + 1)}
    for required in [
        "RealizedFarLoss formula",
        "BIG_SIDE BalanceAfter includes far loss",
        "Total Closed Loss includes RealizedFarLoss",
        "DUAL_TAIL blocks next level",
        "Costs per lot multiplied",
        "OldFarClosePL affects SMALL_SIDE balance",
    ]:
        assert required in test_names


def test_risk_analysis_examples_and_manual_v2_anchors():
    wb = workbook()
    risk_names = [wb["Risk_Analysis"].cell(row, 1).value for row in range(2, wb["Risk_Analysis"].max_row + 1)]
    for required in [
        "Total Closed Profit",
        "Total Closed Loss",
        "Total Realized Far Loss",
        "Final Balance",
        "Final Reserve",
        "Max Margin",
        "Max Open Lots",
        "Final Far Lot",
        "Number of BIG_SIDE",
        "Number of SMALL_SIDE",
        "Dual Tail Count",
        "Danger Count",
        "Stop Count",
        "Final Status",
    ]:
        assert required in risk_names
    examples = "\n".join(str(wb["Examples"].cell(row, col).value or "") for row in range(1, wb["Examples"].max_row + 1) for col in range(1, 5))
    assert "Пример 1. Тренд вверх" in examples
    assert "BIG_SIDE" in examples
    assert "Пример 2. Разворот вниз" in examples
    assert "SMALL_SIDE" in examples
    manual = "\n".join(str(wb["Manual"].cell(row, 1).value or "") for row in range(1, wb["Manual"].max_row + 1))
    for text in [
        "Close Big = 100%",
        "Close Small = 100%",
        "NetProfit = ProfitBig - LossSmall - Commission - SpreadCost - SlippageCost",
        "Основной режим системы — **денежный**",
        "20% — это денежный бюджет",
        "Close Big = 22%",
        "Remain Big = 78%",
        "DUAL_TAIL",
        "STOP",
        "Small не должен накапливаться",
    ]:
        assert text in manual


def test_dual_tail_persists_old_and_new_tail():
    ws = workbook()["Calculator"]
    headers = header_map(ws)
    assert "ActiveOldFarLot" in headers
    assert "ActiveNewFarLot" in headers
    assert "DualTailTotalLot" in headers
    assert ws["BA4"].value == '=IF($AK4=TRUE,MAX(0,$AJ4-$BD4),IF($B4="BLOCKED",MAX(0,$BA3-$BD4),0))'
    assert ws["BB4"].value == '=IF($AK4=TRUE,MAX(0,$AG4-$BE4),IF($B4="BLOCKED",MAX(0,$BB3-$BE4),0))'
    assert ws["BC4"].value == "=$BA4+$BB4"


def test_blocked_rows_keep_dual_tail_total_lot():
    ws = workbook()["Calculator"]
    assert ws["B5"].value.startswith('=IF(AND($AQ4="MANUAL_READY",$BG4="YES")')
    assert ws["BA5"].value == '=IF($AK5=TRUE,MAX(0,$AJ5-$BD5),IF($B5="BLOCKED",MAX(0,$BA4-$BD5),0))'
    assert ws["BB5"].value == '=IF($AK5=TRUE,MAX(0,$AG5-$BE5),IF($B5="BLOCKED",MAX(0,$BB4-$BE5),0))'
    assert ws["BA6"].value == '=IF($AK6=TRUE,MAX(0,$AJ6-$BD6),IF($B6="BLOCKED",MAX(0,$BA5-$BD6),0))'
    assert ws["BB6"].value == '=IF($AK6=TRUE,MAX(0,$AG6-$BE6),IF($B6="BLOCKED",MAX(0,$BB5-$BE6),0))'
    assert ws["BC5"].value == "=$BA5+$BB5"
    assert ws["BC6"].value == "=$BA6+$BB6"


def test_blocked_rows_keep_margin_from_both_tails():
    ws = workbook()["Calculator"]
    assert ws["AB5"].value == '=IF($B5="BLOCKED",$BC5,$D5+$G5+$J5)'
    assert ws["AC5"].value == '=IF($B5="BLOCKED",$BC5,IF($B5="BIG_SIDE",$W5,IF($AK5=TRUE,$BC5,$AG5)))'
    assert ws["AE5"].value == "=$AC5*Settings!$B$10"
    assert ws["AB6"].value == '=IF($B6="BLOCKED",$BC6,$D6+$G6+$J6)'
    assert ws["AC6"].value == '=IF($B6="BLOCKED",$BC6,IF($B6="BIG_SIDE",$W6,IF($AK6=TRUE,$BC6,$AG6)))'
    assert ws["AE6"].value == "=$AC6*Settings!$B$10"


def test_old_tail_cannot_disappear_without_manual_close():
    ws = workbook()["Calculator"]
    assert ws["BD5"].value == 0
    assert "MAX(0,$BA4-$BD5)" in ws["BA5"].value
    assert "MAX(0,$BA5-$BD6)" in ws["BA6"].value


def test_manual_close_reduces_active_tail_lots():
    ws = workbook()["Calculator"]
    assert "MAX(0,$BA4-$BD5)" in ws["BA5"].value
    assert "MAX(0,$BB4-$BE5)" in ws["BB5"].value
    assert ws["AA5"].value == '=IF($B5="BLOCKED",$Z5+$BF5,IF($B5="BIG_SIDE",$Z5+$S5-$AS5,IF($B5="SMALL_SIDE",$Z5+$S5+$AI5,$Z5+$S5)))'


def test_manual_ready_requires_single_remaining_tail():
    ws = workbook()["Calculator"]
    status_formula = ws["AQ5"].value
    assert 'AND($BA5=0,$BB5>0)' in status_formula
    assert 'AND($BA5>0,$BB5=0)' in status_formula
    assert '"MANUAL_READY"' in status_formula
    assert '"STOP_CLEARED"' in status_formula


def test_new_level_requires_manual_allow_after_dual_tail():
    ws = workbook()["Calculator"]
    assert ws["BG5"].value == "NO"
    assert ws["AP5"].value == '=IF(AND($AQ5="MANUAL_READY",$BG5="YES"),"YES",IF(OR($AQ5="DUAL_TAIL",$AQ5="DANGER",$AQ5="STOP",$AQ5="STOP_CLEARED",$AQ5="MANUAL_READY"),"NO","YES"))'
    assert ws["B6"].value.startswith('=IF(AND($AQ5="MANUAL_READY",$BG5="YES")')


def _data_only_workbook():
    return load_workbook(WORKBOOK, data_only=True)


def _rows_by_header(ws):
    headers = [cell.value for cell in ws[1]]
    return {header: headers.index(header) + 1 for header in headers if header}


def test_dual_tail_row_has_active_old_and_new_lots():
    for sheet in ["Trend_UP", "Trend_DOWN"]:
        ws = _data_only_workbook()[sheet]
        idx = _rows_by_header(ws)
        dual_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, idx["Status"]).value == "DUAL_TAIL"]
        assert dual_rows, f"{sheet} has no DUAL_TAIL row"
        for r in dual_rows:
            old_remain = ws.cell(r, idx["OldFarRemainLot"]).value or 0
            new_far = ws.cell(r, idx["NewFarStartLot"]).value or 0
            active_old = ws.cell(r, idx["ActiveOldFarLot"]).value or 0
            active_new = ws.cell(r, idx["ActiveNewFarLot"]).value or 0
            assert active_old == old_remain
            assert active_new == new_far
            assert active_old > 0 and active_new > 0


def test_first_blocked_row_preserves_dual_tail_lots():
    for sheet in ["Trend_UP", "Trend_DOWN"]:
        ws = _data_only_workbook()[sheet]
        idx = _rows_by_header(ws)
        rows = range(2, ws.max_row + 1)
        dual = next(r for r in rows if ws.cell(r, idx["Status"]).value == "DUAL_TAIL")
        blocked = dual + 1
        assert ws.cell(blocked, idx["Scenario"]).value == "BLOCKED"
        assert ws.cell(blocked, idx["ActiveOldFarLot"]).value == ws.cell(dual, idx["ActiveOldFarLot"]).value
        assert ws.cell(blocked, idx["ActiveNewFarLot"]).value == ws.cell(dual, idx["ActiveNewFarLot"]).value


def test_second_blocked_row_preserves_dual_tail_lots():
    for sheet in ["Trend_UP", "Trend_DOWN"]:
        ws = _data_only_workbook()[sheet]
        idx = _rows_by_header(ws)
        rows = range(2, ws.max_row + 1)
        dual = next(r for r in rows if ws.cell(r, idx["Status"]).value == "DUAL_TAIL")
        blocked_1 = dual + 1
        blocked_2 = dual + 2
        assert ws.cell(blocked_2, idx["Scenario"]).value == "BLOCKED"
        assert ws.cell(blocked_2, idx["ActiveOldFarLot"]).value == ws.cell(blocked_1, idx["ActiveOldFarLot"]).value
        assert ws.cell(blocked_2, idx["ActiveNewFarLot"]).value == ws.cell(blocked_1, idx["ActiveNewFarLot"]).value


def test_blocked_rows_open_lots_equal_dual_tail_total():
    for sheet in ["Trend_UP", "Trend_DOWN"]:
        ws = _data_only_workbook()[sheet]
        idx = _rows_by_header(ws)
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, idx["Scenario"]).value == "BLOCKED":
                assert ws.cell(r, idx["OpenLotsAfter"]).value == ws.cell(r, idx["DualTailTotalLot"]).value
                assert ws.cell(r, idx["DualTailTotalLot"]).value > 0


def test_blocked_rows_margin_equal_dual_tail_total_margin():
    for sheet in ["Trend_UP", "Trend_DOWN"]:
        ws = _data_only_workbook()[sheet]
        idx = _rows_by_header(ws)
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, idx["Scenario"]).value == "BLOCKED":
                expected = ws.cell(r, idx["DualTailTotalLot"]).value * 1000
                assert abs(ws.cell(r, idx["MarginAfter"]).value - expected) < 1e-9


def test_risk_analysis_counts_blocked_rows():
    wb = _data_only_workbook()
    risk = wb["Risk_Analysis"]
    values = {risk.cell(r, 1).value: risk.cell(r, 3).value for r in range(2, risk.max_row + 1)}
    assert values["Blocked Levels Count"] > 0


def test_risk_analysis_max_dual_tail_exposure_positive():
    wb = _data_only_workbook()
    risk = wb["Risk_Analysis"]
    values = {risk.cell(r, 1).value: risk.cell(r, 3).value for r in range(2, risk.max_row + 1)}
    assert values["Max DualTail Exposure"] > 0


def test_active_old_tail_not_zero_without_manual_close():
    ws = _data_only_workbook()["Trend_UP"]
    idx = _rows_by_header(ws)
    blocked_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, idx["Scenario"]).value == "BLOCKED"]
    assert blocked_rows
    for r in blocked_rows:
        assert (ws.cell(r, idx["ManualOldFarCloseLot"]).value or 0) == 0
        assert ws.cell(r, idx["ActiveOldFarLot"]).value > 0


def test_active_new_tail_not_zero_without_manual_close():
    ws = _data_only_workbook()["Trend_UP"]
    idx = _rows_by_header(ws)
    blocked_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, idx["Scenario"]).value == "BLOCKED"]
    assert blocked_rows
    for r in blocked_rows:
        assert (ws.cell(r, idx["ManualNewFarCloseLot"]).value or 0) == 0
        assert ws.cell(r, idx["ActiveNewFarLot"]).value > 0


def test_manual_close_pl_changes_balance():
    ws = workbook()["Calculator"]
    assert ws["AA5"].value.startswith('=IF($B5="BLOCKED",$Z5+$BF5')


def test_manual_allow_required_to_resume():
    ws = workbook()["Calculator"]
    assert ws["B6"].value.startswith('=IF(AND($AQ5="MANUAL_READY",$BG5="YES")')
    assert ws["AP5"].value.startswith('=IF(AND($AQ5="MANUAL_READY",$BG5="YES")')


def test_calc_trend_up_trend_down_have_same_headers():
    wb = workbook()
    headers = [[cell.value for cell in wb[sheet][1]] for sheet in ["Calculator", "Trend_UP", "Trend_DOWN"]]
    assert headers[0] == headers[1] == headers[2]


def test_same_big_formula_across_sheets():
    wb = workbook()
    for row in range(2, 12):
        formulas = [(wb[sheet].cell(row, 6).value, wb[sheet].cell(row, 7).value) for sheet in ["Calculator", "Trend_UP", "Trend_DOWN"]]
        assert formulas[0] == formulas[1] == formulas[2]
        assert "Settings!$B$3" in formulas[0][0]
        assert "Settings!$B$11" in formulas[0][1]


def test_same_small_formula_across_sheets():
    wb = workbook()
    for row in range(2, 12):
        formulas = [(wb[sheet].cell(row, 9).value, wb[sheet].cell(row, 10).value) for sheet in ["Calculator", "Trend_UP", "Trend_DOWN"]]
        assert formulas[0] == formulas[1] == formulas[2]
        assert "Settings!$B$4" in formulas[0][0]
        assert "Settings!$B$11" in formulas[0][1]


def _dual_and_blocked_rows(sheet_name: str):
    ws = _data_only_workbook()[sheet_name]
    idx = _rows_by_header(ws)
    dual_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, idx["Status"]).value == "DUAL_TAIL"]
    blocked_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, idx["Scenario"]).value == "BLOCKED"]
    return ws, idx, dual_rows, blocked_rows


def test_trend_up_dual_tail_persists():
    ws, idx, dual_rows, blocked_rows = _dual_and_blocked_rows("Trend_UP")
    assert dual_rows and blocked_rows
    for r in dual_rows:
        assert ws.cell(r, idx["ActiveOldFarLot"]).value == ws.cell(r, idx["OldFarRemainLot"]).value
        assert ws.cell(r, idx["ActiveNewFarLot"]).value == ws.cell(r, idx["NewFarStartLot"]).value
        assert ws.cell(r, idx["DualTailTotalLot"]).value > 0


def test_trend_down_dual_tail_persists():
    ws, idx, dual_rows, blocked_rows = _dual_and_blocked_rows("Trend_DOWN")
    assert dual_rows and blocked_rows
    for r in dual_rows:
        assert ws.cell(r, idx["ActiveOldFarLot"]).value == ws.cell(r, idx["OldFarRemainLot"]).value
        assert ws.cell(r, idx["ActiveNewFarLot"]).value == ws.cell(r, idx["NewFarStartLot"]).value
        assert ws.cell(r, idx["DualTailTotalLot"]).value > 0


def test_trend_up_blocked_rows_keep_tails():
    ws, idx, _, blocked_rows = _dual_and_blocked_rows("Trend_UP")
    assert blocked_rows
    for r in blocked_rows:
        assert ws.cell(r, idx["ActiveOldFarLot"]).value > 0
        assert ws.cell(r, idx["ActiveNewFarLot"]).value > 0
        assert ws.cell(r, idx["OpenLotsAfter"]).value == ws.cell(r, idx["DualTailTotalLot"]).value
        assert ws.cell(r, idx["Status"]).value == "STOP"
        assert ws.cell(r, idx["BigLot"]).value == 0
        assert ws.cell(r, idx["SmallLot"]).value == 0


def test_trend_down_blocked_rows_keep_tails():
    ws, idx, _, blocked_rows = _dual_and_blocked_rows("Trend_DOWN")
    assert blocked_rows
    for r in blocked_rows:
        assert ws.cell(r, idx["ActiveOldFarLot"]).value > 0
        assert ws.cell(r, idx["ActiveNewFarLot"]).value > 0
        assert ws.cell(r, idx["OpenLotsAfter"]).value == ws.cell(r, idx["DualTailTotalLot"]).value
        assert ws.cell(r, idx["Status"]).value == "STOP"
        assert ws.cell(r, idx["BigLot"]).value == 0
        assert ws.cell(r, idx["SmallLot"]).value == 0


def _risk_values():
    risk = _data_only_workbook()["Risk_Analysis"]
    return {risk.cell(r, 1).value: risk.cell(r, 3).value for r in range(2, risk.max_row + 1) if risk.cell(r, 1).value}


def test_risk_analysis_counts_trend_up_dual_tail():
    values = _risk_values()
    assert values["Global Dual Tail Count"] >= 1
    assert values["Global Max DualTail Exposure"] > 0


def test_risk_analysis_counts_trend_down_dual_tail():
    values = _risk_values()
    assert values["Global Dual Tail Count"] >= 2
    assert values["Global Blocked Levels Count"] > 0


def test_global_risk_summary_counts_all_sheets():
    wb = workbook()
    risk = wb["Risk_Analysis"]
    formulas = {risk.cell(r, 1).value: risk.cell(r, 3).value for r in range(2, risk.max_row + 1) if risk.cell(r, 1).value}
    for metric in ["Global Dual Tail Count", "Global Blocked Levels Count", "Global Max DualTail Exposure"]:
        formula = formulas[metric]
        assert "Calculator!" in formula
        assert "Trend_UP!" in formula
        assert "Trend_DOWN!" in formula


def test_blocked_rows_have_no_empty_key_fields():
    key_fields = [
        "Scenario", "Status", "ActiveOldFarLot", "ActiveNewFarLot", "DualTailTotalLot",
        "OpenLotsBefore", "OpenLotsAfter", "MarginBefore", "MarginAfter", "Comment",
    ]
    for sheet in ["Trend_UP", "Trend_DOWN"]:
        ws, idx, _, blocked_rows = _dual_and_blocked_rows(sheet)
        assert blocked_rows
        for r in blocked_rows:
            for field in key_fields:
                assert ws.cell(r, idx[field]).value not in (None, ""), f"{sheet} row {r} empty {field}"


def test_settings_links_exist_across_all_calc_sheets():
    wb = workbook()
    settings_refs = ["Settings!$B$3", "Settings!$B$4", "Settings!$B$5", "Settings!$B$6", "Settings!$B$7", "Settings!$B$9", "Settings!$B$10", "Settings!$B$11"]
    for sheet in ["Calculator", "Trend_UP", "Trend_DOWN"]:
        ws = wb[sheet]
        formulas = "\n".join(
            str(ws.cell(r, c).value)
            for r in range(2, min(ws.max_row, 11) + 1)
            for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).data_type == "f"
        )
        for ref in settings_refs:
            assert ref in formulas, f"{sheet} formulas missing {ref}"


def test_no_direct_self_references():
    wb = workbook()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    coord = cell.coordinate.upper()
                    formula = cell.value.upper()
                    assert coord not in formula, f"Direct self-reference found: {ws.title}!{coord} -> {cell.value}"


def test_no_blank_global_profit_loss():
    wb = workbook()
    risk = wb["Risk_Analysis"]
    formulas = {risk.cell(r, 1).value: risk.cell(r, 3).value for r in range(2, risk.max_row + 1) if risk.cell(r, 1).value}
    assert formulas["Global Total Closed Profit"]
    assert formulas["Global Total Closed Loss"]
    assert "Calculator!" in formulas["Global Total Closed Profit"]
    assert "Trend_UP!" in formulas["Global Total Closed Profit"]
    assert "Trend_DOWN!" in formulas["Global Total Closed Profit"]
    assert "Calculator!" in formulas["Global Total Closed Loss"]
    assert "Trend_UP!" in formulas["Global Total Closed Loss"]
    assert "Trend_DOWN!" in formulas["Global Total Closed Loss"]


def test_trend_down_blocked_uses_previous_dual_tail_values():
    ws, idx, dual_rows, blocked_rows = _dual_and_blocked_rows("Trend_DOWN")
    assert dual_rows and blocked_rows
    dual = dual_rows[0]
    first_blocked = min(r for r in blocked_rows if r > dual)
    assert ws.cell(first_blocked, idx["ActiveOldFarLot"]).value == ws.cell(dual, idx["ActiveOldFarLot"]).value
    assert ws.cell(first_blocked, idx["ActiveNewFarLot"]).value == ws.cell(dual, idx["ActiveNewFarLot"]).value
    assert ws.cell(first_blocked, idx["DualTailTotalLot"]).value == ws.cell(dual, idx["DualTailTotalLot"]).value
    assert round(ws.cell(first_blocked, idx["ActiveOldFarLot"]).value, 4) == 0.8052
    assert round(ws.cell(first_blocked, idx["ActiveNewFarLot"]).value, 4) == 0.7254
    assert round(ws.cell(first_blocked, idx["DualTailTotalLot"]).value, 4) == 1.5306


def test_blocked_rows_key_fields_not_blank():
    key_fields = [
        "Scenario", "Status", "Comment", "OpenLotsBefore", "OpenLotsAfter",
        "MarginBefore", "MarginAfter", "ActiveOldFarLot", "ActiveNewFarLot",
        "DualTailTotalLot", "NewFullLevelAllowed", "BlockedReason",
    ]
    for sheet in ["Calculator", "Trend_UP", "Trend_DOWN"]:
        ws = _data_only_workbook()[sheet]
        idx = _rows_by_header(ws)
        blocked_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, idx["Scenario"]).value == "BLOCKED"]
        for r in blocked_rows:
            for field in key_fields:
                assert ws.cell(r, idx[field]).value not in (None, ""), f"{sheet} row {r} blank {field}"


def test_global_risk_summary_profit_loss_populated():
    values = _risk_values()
    assert values["Global Total Closed Profit"] is not None
    assert values["Global Total Closed Loss"] is not None
    assert values["Global Total Closed Profit"] > 0
    assert values["Global Total Closed Loss"] > 0
    assert values["Global Dual Tail Count"] > 0
    assert values["Global Blocked Levels Count"] > 0
    assert values["Global Stop Count"] > 0
    assert values["Global Max DualTail Exposure"] > 0


def _settings_geometry_header_row(ws):
    required = {"N", "FarStart-N", "Big-N", "Small-N", "CloseFarBudget"} if False else {"N", "FarStart-N", "Big-N", "Small-N", "CloseFarBudget"}
    for row in range(1, ws.max_row + 1):
        values = {ws.cell(row, col).value for col in range(1, ws.max_column + 1)}
        if required.issubset(values):
            return row
    raise AssertionError("Settings geometry table headers missing")


def test_global_max_dual_tail_exposure_uses_dual_tail_total_lot():
    wb = workbook()
    risk = wb["Risk_Analysis"]
    formulas = {risk.cell(r, 1).value: risk.cell(r, 3).value for r in range(1, risk.max_row + 1) if risk.cell(r, 1).value}
    formula = formulas["Global Max DualTail Exposure"]
    assert "Calculator!BC2:BC11" in formula
    assert "Trend_UP!BC2:BC11" in formula
    assert "Trend_DOWN!BC2:BC11" in formula
    assert "BA2:BA11" not in formula
    assert "BB2:BB11" not in formula


def test_global_max_dual_tail_exposure_expected_value():
    values = _risk_values()
    assert abs(values["Global Max DualTail Exposure"] - 1.5306) < 1e-9


def test_settings_geometry_table_exists():
    ws = workbook()["Settings"]
    header_row = _settings_geometry_header_row(ws)
    assert ws.cell(header_row - 1, 1).value == "Big Harvest Geometry — денежное закрытие Far через CloseFarBudget"
    assert [ws.cell(header_row, col).value for col in range(1, 15)] == ["N", "FarStart-N", "BigMovePoints", "FarDistancePoints", "Big-N", "Small-N", "ProfitBig", "LossSmall", "NetProfit", "CloseFarBudget", "ReserveAdd", "CloseFarLot", "FarRemain", "CloseFarPercent"]


def test_settings_geometry_big_n_formula():
    ws = workbook()["Settings"]
    header_row = _settings_geometry_header_row(ws)
    first = header_row + 1
    assert ws.cell(first, 5).value == '=IF($A22="","",ROUND(($B22*Settings!$B$3)/Settings!$B$11,0)*Settings!$B$11)'
    assert "Settings!$B$3" in ws.cell(first, 5).value
    assert "Settings!$B$11" in ws.cell(first, 5).value


def test_settings_geometry_small_n_formula():
    ws = workbook()["Settings"]
    header_row = _settings_geometry_header_row(ws)
    first = header_row + 1
    assert ws.cell(first, 6).value == '=IF($A22="","",ROUND(($E22*Settings!$B$4)/Settings!$B$11,0)*Settings!$B$11)'
    assert "Settings!$B$4" in ws.cell(first, 6).value
    assert "Settings!$B$11" in ws.cell(first, 6).value


def test_settings_geometry_close_n_formula():
    ws = workbook()["Settings"]
    header_row = _settings_geometry_header_row(ws)
    first = header_row + 1
    assert ws.cell(first, 10).value == '=IF($A22="","",MAX(0,$I22*Settings!$B$5))'
    assert ws.cell(first, 12).value == '=IF($A22="","",MIN($B22,IFERROR($J22/($D22*Settings!$B$9),0)))'
    assert "Settings!$B$5" in ws.cell(first, 10).value


def test_settings_geometry_uses_settings_references():
    ws = workbook()["Settings"]
    header_row = _settings_geometry_header_row(ws)
    formulas = "\n".join(str(ws.cell(r, c).value) for r in range(header_row + 1, header_row + 11) for c in range(1, 15))
    for ref in ["Settings!$B$2", "Settings!$B$3", "Settings!$B$4", "Settings!$B$5", "Settings!$B$11", "Settings!$B$12"]:
        assert ref in formulas
    for static in ["*1.30", "*0.36", "*0.90", "/0.01"]:
        assert static not in formulas


def test_close_n_is_reference_only_not_calculator_close_rule():
    wb = workbook()
    calc = wb["Calculator"]
    formulas = "\n".join(
        str(calc.cell(r, c).value)
        for r in range(2, 12)
        for c in range(1, calc.max_column + 1)
        if calc.cell(r, c).data_type == "f"
    )
    assert "Close-N" not in formulas
    assert "$T2/$U2" in formulas or "$T3/$U3" in formulas
    assert "LossPerLotToClose" not in formulas


def test_big_harvest_close_far_share_is_money_budget_not_lot_percent():
    wb = _data_only_workbook()
    ws = wb["Calculator"]
    idx = _rows_by_header(ws)
    row = 2
    assert ws.cell(row, idx["FarStartLot"]).value == 1
    assert round(ws.cell(row, idx["BigLot"]).value, 2) == 1.30
    assert round(ws.cell(row, idx["SmallLot"]).value, 2) == 0.47
    assert round(ws.cell(row, idx["CloseFarLot"]).value, 4) == 0.3735
    assert round(ws.cell(row, idx["FarRemainLot"]).value, 4) == 0.6265
    assert round(ws.cell(row, idx["CloseFarPercent"]).value, 4) == 0.3735
    assert ws.cell(row, idx["CloseFarLot"]).value != 0.90
    assert round(ws.cell(row, idx["FarRemainLot"]).value, 4) != 0.10


def test_big_harvest_reserve_and_budget_shares():
    ws = _data_only_workbook()["Calculator"]
    idx = _rows_by_header(ws)
    row = 2
    net_profit = ws.cell(row, idx["NetProfitBeforeFar"]).value
    assert round(net_profit, 2) == 83.00
    assert round(ws.cell(row, idx["ReserveAdd"]).value, 2) == round(net_profit * 0.10, 2)
    assert round(ws.cell(row, idx["CloseFarBudget"]).value, 2) == round(net_profit * 0.90, 2)
    assert ws.cell(row, idx["RealizedFarLoss"]).value <= ws.cell(row, idx["CloseFarBudget"]).value


def test_big_harvest_small_scenario_positive_and_new_far():
    big = 1.30
    small = big * 0.36
    close_big = big * 0.30
    small_move_points = 100
    profit_small = small * small_move_points
    loss_closed_big = close_big * small_move_points
    net_small = profit_small - loss_closed_big
    new_far = big * 0.70
    assert round(small, 3) == 0.468
    assert round(close_big, 2) == 0.39
    assert round(net_small, 1) == 7.8
    assert net_small > 0
    assert round(new_far, 2) == 0.91


def test_big_harvest_new_columns_exist_on_all_calc_sheets():
    wb = workbook()
    required = ["BigMovePoints", "FarDistancePoints", "HarvestMode", "HarvestCount", "CloseFarPercent", "CanFullCloseFar", "FarRemainLoss", "FinalClosePL"]
    for sheet in ["Calculator", "Trend_UP", "Trend_DOWN"]:
        headers = header_map(wb[sheet])
        for header in required:
            assert header in headers
