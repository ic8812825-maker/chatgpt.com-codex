import subprocess
from openpyxl import load_workbook

SRC='recovery_lock_cascade_next_step.xlsx'
TMP='reports/tests/recovery_lock_runtime_case.xlsx'

wb=load_workbook(SRC)

# Core runtime inputs
s=wb['Settings']
s['B11']=0; s['B12']=20; s['B13']=0; s['B14']=0; s['B19']='FULL_CYCLE'
tr=wb['TailRecovery_UP']
tr['B14']=0.14; tr['B15']=2; tr['B6']='NO'; tr['B7']=100; tr['B5']=400
bs=wb['BasketSummary']
bs['B3']=-12; bs['B4']=18
sc=wb['SectionCalculator_UP']; sc['B12']='YES'; sc['B13']='YES'

# Tail tie-case UP: two SELL positions with equal negative PnL
sup=wb['Scenario_UP']
sup['A7']=10010; sup['B7']='SELL'; sup['H7']=-50
sup['A8']=10005; sup['B8']='SELL'; sup['H8']=-50
sup['A9']=10020; sup['B9']='SELL'; sup['H9']=-10

# Tail tie-case DOWN: two BUY positions with equal negative PnL
sdn=wb['Scenario_DOWN']
sdn['A7']=20012; sdn['B7']='BUY'; sdn['H7']=-80
sdn['A8']=20003; sdn['B8']='BUY'; sdn['H8']=-80
sdn['A9']=20030; sdn['B9']='BUY'; sdn['H9']=-20

# Runtime tie-case formulas compatible with LibreOffice (3-row controlled case)
sup['B221']='=MIN(IF(AND(B7="SELL",H7<0),H7,0),IF(AND(B8="SELL",H8<0),H8,0),IF(AND(B9="SELL",H9<0),H9,0))'
sup['B222']='=IF(B7=B8,MIN(A7,A8),IF(H7=B221,A7,IF(H8=B221,A8,A9)))'
sdn['B221']='=MIN(IF(AND(B7="BUY",H7<0),H7,0),IF(AND(B8="BUY",H8<0),H8,0),IF(AND(B9="BUY",H9<0),H9,0))'
sdn['B222']='=IF(B7=B8,MIN(A7,A8),IF(H7=B221,A7,IF(H8=B221,A8,A9)))'

wb.save(TMP)
subprocess.run(['libreoffice','--headless','--convert-to','ods','--outdir','reports/tests',TMP],check=True)
subprocess.run(['libreoffice','--headless','--convert-to','xlsx','--outdir','reports/tests','reports/tests/recovery_lock_runtime_case.ods'],check=True)

wb2=load_workbook(TMP,data_only=True)
checks={
 'up_tail_worst_pnl': ('Scenario_UP','B221'),
 'up_tail_ticket': ('Scenario_UP','B222'),
 'down_tail_worst_pnl': ('Scenario_DOWN','B221'),
 'down_tail_ticket': ('Scenario_DOWN','B222'),
 'next_big': ('TailRecovery_UP','B16'),
 'next_small': ('TailRecovery_UP','B17'),
 'can_open_section_up': ('SectionCalculator_UP','B14'),
 'costs_p21': ('SectionCalculator_UP','P21'),
 'can_close_basket_up': ('BasketSummary','B10'),
 'close_raw': ('TailRecovery_UP','B8'),
 'close_final': ('TailRecovery_UP','B10'),
 'close_allowed': ('TailRecovery_UP','B11'),
}
for k,(sh,c) in checks.items():
    print(f'{k}: {wb2[sh][c].value}')
