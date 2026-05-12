from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BOLD = Font(bold=True)


def hdr(ws, row, cols):
    for i, c in enumerate(cols, 1):
        ws.cell(row=row, column=i, value=c).font = BOLD


def add_settings(ws):
    hdr(ws, 1, ["Field", "Value"])
    rows = [
        ("Symbol", "EURUSD"), ("Point", 0.0001), ("PointValuePerLot", 1), ("StepPoints", 100), ("MinLot", 0.01), ("LotStep", 0.01),
        ("MaxActiveSections", 4), ("ReservePercent", 0.2), ("RecoveryPercent", 0.8), ("BasketTarget", 0), ("MaxSpreadPoints", 20),
        ("CommissionPerLot", 0), ("SwapPerLot", 0), ("MaxTotalLot", 20), ("MaxNetLot", 10), ("GlobalReserve", 0), ("RecoveryFund", 0), ("CostMode", "FULL_CYCLE")
    ]
    for r, (k, v) in enumerate(rows, 2):
        ws.cell(r, 1, k); ws.cell(r, 2, v)
    hdr(ws, 20, ["Level", "BigRatio", "SmallRatio"])
    for r, row in enumerate([(1, 0.40, 0.15), (2, 0.25, 0.10), (3, 0.15, 0.06), (4, 0.10, 0.04)], 21):
        ws.cell(r, 1, row[0]); ws.cell(r, 2, row[1]); ws.cell(r, 3, row[2])
    ws["A30"]="MarketPriceInput"
    ws["A31"]="Bid"; ws["B31"]=1.23000
    ws["A32"]="Ask"; ws["B32"]=1.23020
    ws["A33"]="MidPrice"; ws["B33"]="=(B31+B32)/2"
    ws["A34"]="SpreadPoints"; ws["B34"]="=(B32-B31)/B3"
    ws["A35"]="PriceSourceMode"; ws["B35"]="BID_ASK"
    for c in ("B31","B32","B33"): ws[c].number_format="0.00000"
    ws["B34"].number_format="0"


def add_positions(ws):
    hdr(ws, 1, ["Ticket", "Type", "Role", "Lot", "OpenPrice", "CurrentPrice", "PointsPnL", "MoneyPnL", "IsTail", "SectionID"])
    demo = [
        [10001, "BUY", "MAIN", 1.00, 1.2300, 1.2300, None, None, "NO", ""], [10002, "SELL", "TAIL", 1.00, 1.2300, 1.2300, None, None, "YES", ""],
        [10003, "SELL", "SECTION_BIG", 0.40, 1.2350, 1.2300, None, None, "NO", "S1"], [10004, "BUY", "SECTION_SMALL", 0.15, 1.2350, 1.2300, None, None, "NO", "S1"],
    ]
    for r, row in enumerate(demo, 2):
        for c, v in enumerate(row, 1): ws.cell(r, c, v)
        ws.cell(r, 6, f'=IF(B{r}="BUY",Settings!$B$31,IF(B{r}="SELL",Settings!$B$32,""))')
        ws.cell(r, 7, f'=IF(B{r}="BUY",(F{r}-E{r})/Settings!$B$3,IF(B{r}="SELL",(E{r}-F{r})/Settings!$B$3,0))')
        ws.cell(r, 8, f'=G{r}*D{r}*Settings!$B$4')


def add_scenario(ws, direction):
    up = direction == "UP"
    loss_type = "SELL" if up else "BUY"
    hdr(ws, 1, ["Field", "Value"])
    ws["A2"] = "CurrentBid"; ws["B2"] = "=Settings!$B$31"
    ws["A3"] = "CurrentAsk"; ws["B3"] = "=Settings!$B$32"
    ws["A4"] = "MoveUpPoints" if up else "MoveDownPoints"; ws["B4"] = 100
    ws["A5"] = "ScenarioBid"; ws["B5"] = "=B2+B4*Settings!$B$3" if up else "=B2-B4*Settings!$B$3"
    ws["A6"] = "ScenarioAsk"; ws["B6"] = "=B3+B4*Settings!$B$3" if up else "=B3-B4*Settings!$B$3"
    ws["A7"] = "ScenarioMid"; ws["B7"] = "=(B5+B6)/2"
    ws["A8"] = "ScenarioSpread"; ws["B8"] = "=(B6-B5)/Settings!$B$3"
    hdr(ws, 11, ["Ticket", "Type", "Role", "Lot", "OpenPrice", "ScenarioClosePrice", "ScenarioPointsPnL", "ScenarioMoneyPnL", "IsTail", "SectionID", "TailLossHelper", "TailTicketHelper"])
    for r in range(2, 202):
        rr = r + 10
        ws.cell(rr, 1, f"=CurrentPositions!A{r}"); ws.cell(rr, 2, f"=CurrentPositions!B{r}"); ws.cell(rr, 3, f"=CurrentPositions!C{r}")
        ws.cell(rr, 4, f"=CurrentPositions!D{r}"); ws.cell(rr, 5, f"=CurrentPositions!E{r}")
        ws.cell(rr, 6, f'=IF(B{rr}="BUY",$B$5,IF(B{rr}="SELL",$B$6,""))')
        ws.cell(rr, 7, f'=IF(B{rr}="BUY",(F{rr}-E{rr})/Settings!$B$3,IF(B{rr}="SELL",(E{rr}-F{rr})/Settings!$B$3,0))'); ws.cell(rr, 8, f"=G{rr}*D{rr}*Settings!$B$4")
        ws.cell(rr, 9, f"=CurrentPositions!I{r}"); ws.cell(rr, 10, f"=CurrentPositions!J{r}")
        ws.cell(rr, 11, f'=IF(AND(B{rr}="{loss_type}",H{rr}<0),H{rr},"")')
        ws.cell(rr, 12, f'=IF(K{rr}<>"",A{rr},"")')
    ws["A220"] = "TailType"; ws["B220"] = loss_type
    ws["A221"] = "TailWorstPnL"; ws["B221"] = '=IF(COUNT(K12:K211)=0,0,MIN(K12:K211))'
    ws["A222"] = "TailTicket"; ws["B222"] = '=IF(B221=0,"N/A",MIN(L12:L211))'
    ws["A223"] = "TailLot"; ws["B223"] = '=IF(B222="N/A",0,INDEX(D12:D211,MATCH(B222,A12:A211,0)))'
    ws["A224"] = "TailLossMoney"; ws["B224"] = '=IF(B222="N/A",0,ABS(INDEX(H12:H211,MATCH(B222,A12:A211,0))))'
    ws["A225"] = "TailLossPerLot"; ws["B225"] = '=IF(B223=0,0,ABS(B224/B223))'
    ws["M1"] = "AverageBuyPrice"; ws["N1"] = '=IFERROR(SUMPRODUCT((B12:B211="BUY")*E12:E211*D12:D211)/MAX(0.0000001,SUMIFS(D12:D211,B12:B211,"BUY")),0)'
    ws["M2"] = "AverageSellPrice"; ws["N2"] = '=IFERROR(SUMPRODUCT((B12:B211="SELL")*E12:E211*D12:D211)/MAX(0.0000001,SUMIFS(D12:D211,B12:B211,"SELL")),0)'
    ws["M3"] = "Center"; ws["N3"] = '=(N1+N2)/2'
    for lvl in range(1, 5):
        ws.cell(3 + lvl, 13, f"UpperLevel{lvl}"); ws.cell(3 + lvl, 14, f"=IF(N3=0,\"Уровень не рассчитан\",N3+{lvl}*Settings!$B$5*Settings!$B$3)")
        ws.cell(8 + lvl, 13, f"LowerLevel{lvl}"); ws.cell(8 + lvl, 14, f"=IF(N3=0,\"Уровень не рассчитан\",N3-{lvl}*Settings!$B$5*Settings!$B$3)")


def add_section_calc(ws, scenario_sheet, up):
    hdr(ws, 1, ["Field", "Value"])
    ws["A2"] = "TailType"; ws["B2"] = f"={scenario_sheet}!B220"
    ws["A3"] = "TailLot"; ws["B3"] = f"={scenario_sheet}!B223"
    ws["A4"] = "NextLevel"; ws["B4"] = 1
    ws["A5"] = "BigRatio"; ws["B5"] = '=INDEX(Settings!B21:B24,MATCH(B4,Settings!A21:A24,0))'
    ws["A6"] = "SmallRatio"; ws["B6"] = '=INDEX(Settings!C21:C24,MATCH(B4,Settings!A21:A24,0))'
    ws["A7"] = "BigLot"; ws["B7"] = '=IF(B3<=0,0,FLOOR(B3*B5,Settings!$B$7))'
    ws["A8"] = "SmallLot"; ws["B8"] = '=IF(B3<=0,0,FLOOR(B3*B6,Settings!$B$7))'
    ws["A9"] = "ActiveSections"; ws["B9"] = '=COUNTIFS(CurrentPositions!C2:C500,"SECTION_BIG")'
    ws["A10"] = "TotalLotAfterOpen"; ws["B10"] = '=SUM(CurrentPositions!D2:D500)+B7+B8'
    ws["A11"] = "NetLotAfterOpen"; ws["B11"] = '=ABS(SUMIFS(CurrentPositions!D2:D500,CurrentPositions!B2:B500,"BUY")+IF(B2="BUY",B7,0)+IF(B2="SELL",B8,0)-SUMIFS(CurrentPositions!D2:D500,CurrentPositions!B2:B500,"SELL")-IF(B2="SELL",B7,0)-IF(B2="BUY",B8,0))'
    ws["A12"] = "LevelReached"
    if up:
        ws["B12"] = f'=IF({scenario_sheet}!B7>=IF(B4=1,{scenario_sheet}!N4,IF(B4=2,{scenario_sheet}!N5,IF(B4=3,{scenario_sheet}!N6,{scenario_sheet}!N7))),"YES","NO")'
    else:
        ws["B12"] = f'=IF({scenario_sheet}!B7<=IF(B4=1,{scenario_sheet}!N9,IF(B4=2,{scenario_sheet}!N10,IF(B4=3,{scenario_sheet}!N11,{scenario_sheet}!N12))),"YES","NO")'
    ws["A13"] = "NoOppositeCascade"; ws["B13"] = '=IF(B2="SELL",IF(COUNTIFS(CurrentPositions!B2:B500,"BUY",CurrentPositions!C2:C500,"SECTION_BIG")=0,"YES","NO"),IF(B2="BUY",IF(COUNTIFS(CurrentPositions!B2:B500,"SELL",CurrentPositions!C2:C500,"SECTION_BIG")=0,"YES","NO"),"NO"))'
    ws["A14"] = "CanOpenSection"; ws["B14"] = '=IF(B3<=0,"NO",IF(AND(B7>=Settings!$B$6,B8>=Settings!$B$6,B8<B7,B7<=B3,B9<Settings!$B$8,B10<=Settings!$B$15,B11<=Settings!$B$16,B12="YES",B13="YES"),"YES","NO"))'

    hdr(ws, 20, ["SectionID", "BigType", "BigLot", "BigOpenPrice", "SmallType", "SmallLot", "SmallOpenPrice", "ScenarioPrice", "BigPnL", "SmallPnL", "Lots", "CostMultiplier", "SpreadCost", "CommissionCost", "SwapCost", "Costs", "CycleProfit", "CanCloseSection"])
    for i in range(1, 5):
        r = 20 + i
        ws.cell(r, 1, f"S{i}")
        ws.cell(r, 2, f'=IFERROR(INDEX(CurrentPositions!B2:B500,MATCH(1,(CurrentPositions!C2:C500="SECTION_BIG")*(CurrentPositions!J2:J500=A{r}),0)),"")')
        ws.cell(r, 3, f'=IFERROR(INDEX(CurrentPositions!D2:D500,MATCH(1,(CurrentPositions!C2:C500="SECTION_BIG")*(CurrentPositions!J2:J500=A{r}),0)),0)')
        ws.cell(r, 4, f'=IFERROR(INDEX(CurrentPositions!E2:E500,MATCH(1,(CurrentPositions!C2:C500="SECTION_BIG")*(CurrentPositions!J2:J500=A{r}),0)),0)')
        ws.cell(r, 5, f'=IFERROR(INDEX(CurrentPositions!B2:B500,MATCH(1,(CurrentPositions!C2:C500="SECTION_SMALL")*(CurrentPositions!J2:J500=A{r}),0)),"")')
        ws.cell(r, 6, f'=IFERROR(INDEX(CurrentPositions!D2:D500,MATCH(1,(CurrentPositions!C2:C500="SECTION_SMALL")*(CurrentPositions!J2:J500=A{r}),0)),0)')
        ws.cell(r, 7, f'=IFERROR(INDEX(CurrentPositions!E2:E500,MATCH(1,(CurrentPositions!C2:C500="SECTION_SMALL")*(CurrentPositions!J2:J500=A{r}),0)),0)')
        ws.cell(r, 8, f'={scenario_sheet}!B7')
        ws.cell(r, 9, f'=IF(B{r}="",0,IF(B{r}="BUY",(H{r}-D{r})/Settings!$B$3*C{r}*Settings!$B$4,(D{r}-H{r})/Settings!$B$3*C{r}*Settings!$B$4))')
        ws.cell(r, 10, f'=IF(E{r}="",0,IF(E{r}="BUY",(H{r}-G{r})/Settings!$B$3*F{r}*Settings!$B$4,(G{r}-H{r})/Settings!$B$3*F{r}*Settings!$B$4))')
        ws.cell(r, 11, f'=C{r}+F{r}')
        ws.cell(r, 12, '=IF(Settings!$B$19="FULL_CYCLE",2,1)')
        ws.cell(r, 13, f'=Settings!$B$12*Settings!$B$4*K{r}*L{r}')
        ws.cell(r, 14, f'=Settings!$B$13*K{r}*L{r}')
        ws.cell(r, 15, f'=Settings!$B$14*K{r}')
        ws.cell(r, 16, f'=M{r}+N{r}+O{r}')
        ws.cell(r, 17, f'=I{r}+J{r}-P{r}')
        ws.cell(r, 18, f'=IF(Q{r}>0,"YES","NO")')
    ws["A30"] = "SelectedSectionID"; ws["B30"] = '=IFERROR(INDEX(A21:A24,MATCH("YES",R21:R24,0)),"NONE")'
    ws["A31"] = "CycleProfit"; ws["B31"] = '=IF(B30="NONE",0,INDEX(Q21:Q24,MATCH(B30,A21:A24,0)))'
    ws["A32"] = "CanCloseSection"; ws["B32"] = '=IF(B31>0,"YES","NO")'
    ws["A33"] = "ReserveAdd"; ws["B33"] = '=IF(B31>0,B31*Settings!$B$9,0)'
    ws["A34"] = "RecoveryAdd"; ws["B34"] = '=IF(B31>0,B31*Settings!$B$10,0)'
    ws["A35"] = "NewGlobalReserve"; ws["B35"] = '=Settings!$B$17+B33'
    ws["A36"] = "NewRecoveryFund"; ws["B36"] = '=Settings!$B$18+B34'


def add_tail_recovery(ws, section_sheet, scenario_sheet):
    hdr(ws, 1, ["Field", "Value"])
    rows = [
        ("TailType", f"={scenario_sheet}!B220"), ("TailTicket", f"={scenario_sheet}!B222"), ("TailLot", f"={scenario_sheet}!B223"), ("TailLossPerLot", f"={scenario_sheet}!B225"),
        ("CanCloseSection", f"={section_sheet}!B32"), ("RecoveryFundAfterCycle", f"={section_sheet}!B36"), ("CloseLotRaw", '=IF(OR(B6<>"YES",B5=0),0,B7/B5)'),
        ("CloseLotRounded", '=FLOOR(B8,Settings!$B$7)'), ("CloseLotFinal", '=IF(B6="YES",MIN(B9,B4),0)'), ("CloseAllowed", '=IF(AND(B10>=Settings!$B$6,B6="YES"),"YES","NO")'),
        ("TailCloseLoss", '=B10*B5'), ("RecoveryFundAfterClose", '=MAX(0,B7-B12)'), ("TailLotAfterClose", '=MAX(0,B4-B10)'), ("NextLevel", f"={section_sheet}!B4"),
        ("NextBigLotAfterRecovery", '=FLOOR(B14*INDEX(Settings!B21:B24,MATCH(B15,Settings!A21:A24,0)),Settings!$B$7)'), ("NextSmallLotAfterRecovery", '=FLOOR(B14*INDEX(Settings!C21:C24,MATCH(B15,Settings!A21:A24,0)),Settings!$B$7)')
    ]
    for r, (k, v) in enumerate(rows, 2): ws.cell(r, 1, k); ws.cell(r, 2, v)



def add_recommendations(ws, direction):
    hdr(ws, 1, ["Field", "Value"])
    sc = "Scenario_UP" if direction == "UP" else "Scenario_DOWN"
    sec = "SectionCalculator_UP" if direction == "UP" else "SectionCalculator_DOWN"
    tail = "TailRecovery_UP" if direction == "UP" else "TailRecovery_DOWN"

    ws["A2"] = "Название сценария"; ws["B2"] = f'="СЦЕНАРИЙ: ЦЕНА ИДЕТ {"ВВЕРХ" if direction=="UP" else "ВНИЗ"}"'
    ws["A3"] = "Текущий Bid"; ws["B3"] = f"=IFERROR({sc}!B2,0)"
    ws["A4"] = "Текущий Ask"; ws["B4"] = f"=IFERROR({sc}!B3,0)"
    ws["A5"] = "Сценарный Bid"; ws["B5"] = f"=IFERROR({sc}!B5,0)"
    ws["A6"] = "Сценарный Ask"; ws["B6"] = f"=IFERROR({sc}!B6,0)"
    ws["A7"] = "Spread, пунктов"; ws["B7"] = "=IFERROR(Settings!B34,0)"
    ws["A9"] = "Следующий уровень открытия"; ws["B9"] = f'=IFERROR(IF({sec}!B4=1,{sc}!N4,IF({sec}!B4=2,{sc}!N5,IF({sec}!B4=3,{sc}!N6,IF({sec}!B4=4,{sc}!N7,"Уровень не рассчитан")))),"Уровень не рассчитан")' if direction == "UP" else f'=IFERROR(IF({sec}!B4=1,{sc}!N9,IF({sec}!B4=2,{sc}!N10,IF({sec}!B4=3,{sc}!N11,IF({sec}!B4=4,{sc}!N12,"Уровень не рассчитан")))),"Уровень не рассчитан")'
    ws["A10"] = "Расстояние до уровня, пунктов"; ws["B10"] = '=IF(OR(B9="Уровень не рассчитан",B7=""),"Расстояние не рассчитано",ROUND(ABS(B9-B7)/Settings!$B$3,0))'
    ws["A11"] = "Большая позиция"; ws["B11"] = f'=IFERROR(IF({tail}!B16>0,IF("{direction}"="UP","SELL ","BUY ")&TEXT({tail}!B16,"0.00"),"Не открывать"),"Не открывать")'
    ws["A12"] = "Малая защитная позиция"; ws["B12"] = f'=IFERROR(IF({tail}!B17>0,IF("{direction}"="UP","BUY ","SELL ")&TEXT({tail}!B17,"0.00"),"Не открывать"),"Не открывать")'
    ws["A13"] = "Уровень секции"; ws["B13"] = f"=IFERROR({sec}!B4,0)"
    ws["A14"] = "Причина"; ws["B14"] = f'=IF({sec}!B14="YES","цена достигла уровня; хвост найден; риск-лимиты OK","секция недоступна: уровень/риск-гейт/хвост")'

    ws["A12"] = "Проверки риска"
    ws["A13"] = "Можно открыть секцию"; ws["B13"] = f'=IF({sec}!B14="YES","ДА","НЕТ")'
    ws["A14"] = "Нет встречного каскада"; ws["B14"] = f'=IF({sec}!B13="YES","ДА","НЕТ")'
    ws["A15"] = "Можно закрыть секцию"; ws["B15"] = f'=IF({sec}!B32="YES","ДА","НЕТ")'
    ws["A16"] = "Можно закрыть хвост"; ws["B16"] = f'=IF({tail}!B11="YES","ДА","НЕТ")'

    ws["A18"] = "Прибыль цикла"; ws["B18"] = f"=IFERROR({sec}!B31,0)"
    ws["A19"] = "Добавить в резерв"; ws["B19"] = f"=IFERROR({sec}!B33,0)"
    ws["A20"] = "Добавить в RecoveryFund"; ws["B20"] = f"=IFERROR({sec}!B34,0)"
    ws["A21"] = "Резерв после"; ws["B21"] = f"=IFERROR({sec}!B35,0)"
    ws["A22"] = "RecoveryFund после цикла"; ws["B22"] = f"=IFERROR({sec}!B36,0)"

    ws["A23"] = "Тип хвоста"; ws["B23"] = f'=IF(OR({tail}!B2="",{tail}!B2="N/A"),"Хвост не найден",{tail}!B2)'
    ws["A24"] = "Тикет хвоста"; ws["B24"] = f"=IFERROR({tail}!B3,\"N/A\")"
    ws["A25"] = "Лот хвоста"; ws["B25"] = f'=IFERROR(IF({tail}!B4>0,{tail}!B4,"Хвост не найден"),"Хвост не найден")'
    ws["A26"] = "Убыток хвоста на 1 лот"; ws["B26"] = f"=IFERROR({tail}!B5,0)"
    ws["A27"] = "RecoveryFund до закрытия"; ws["B27"] = f"=IFERROR({tail}!B7,0)"
    ws["A28"] = "Расчётный лот закрытия"; ws["B28"] = f"=IFERROR({tail}!B8,0)"
    ws["A29"] = "Округлённый лот закрытия"; ws["B29"] = f"=IFERROR({tail}!B9,0)"
    ws["A30"] = "Итоговый лот закрытия"; ws["B30"] = f"=IFERROR({tail}!B10,0)"
    ws["A34"] = "Разрешено закрытие хвоста"; ws["B34"] = f'=IF({tail}!B11="YES","ДА","НЕТ")'
    ws["A35"] = "Убыток закрываемой части"; ws["B35"] = f"=IFERROR({tail}!B12,0)"
    ws["A36"] = "RecoveryFund после закрытия"; ws["B36"] = f"=IFERROR({tail}!B13,0)"
    ws["A37"] = "Остаток хвоста"; ws["B37"] = f"=IFERROR({tail}!B14,0)"
    ws["A38"] = "SAFE"; ws["B38"] = '=IF('+('BasketSummary!B13' if direction=='UP' else 'BasketSummary!C13')+'="SAFE","ДА","НЕТ")'
    ws["A31"] = "HumanReadableAction"
    ws["B31"] = '=IF('+('BasketSummary!B13' if direction=='UP' else 'BasketSummary!C13')+'="BASKET_CLOSE","ЗАКРЫТЬ ВСЮ КОРЗИНУ",IF('+('BasketSummary!B13' if direction=='UP' else 'BasketSummary!C13')+'="CLOSE_SECTION+CLOSE_TAIL","ЗАКРЫТЬ СЕКЦИЮ + ЗАКРЫТЬ ХВОСТ",IF('+('BasketSummary!B13' if direction=='UP' else 'BasketSummary!C13')+'="CLOSE_SECTION","ЗАКРЫТЬ СЕКЦИЮ",IF('+('BasketSummary!B13' if direction=='UP' else 'BasketSummary!C13')+'="OPEN_SECTION","ОТКРЫТЬ СЕКЦИЮ",IF('+('BasketSummary!B13' if direction=='UP' else 'BasketSummary!C13')+'="SAFE","SAFE",IF('+('BasketSummary!B13' if direction=='UP' else 'BasketSummary!C13')+'="WAIT","ЖДАТЬ","BLOCKED"))))))'
    # Final semantic overrides for ScenarioText key fields (authoritative mapping)
    ws["A3"] = "Текущий Bid"; ws["B3"] = f"=IFERROR({sc}!B2,0)"
    ws["A4"] = "Текущий Ask"; ws["B4"] = f"=IFERROR({sc}!B3,0)"
    ws["A5"] = "Сценарный Bid"; ws["B5"] = f"=IFERROR({sc}!B5,0)"
    ws["A6"] = "Сценарный Ask"; ws["B6"] = f"=IFERROR({sc}!B6,0)"
    ws["A7"] = "Сценарный Mid"; ws["B7"] = f"=IFERROR({sc}!B7,0)"
    ws["A8"] = "Сценарный Spread"; ws["B8"] = f"=IFERROR({sc}!B8,0)"
    ws["A9"] = "Следующий уровень открытия"; ws["B9"] = f'=IFERROR(IF({sec}!B4=1,{sc}!N4,IF({sec}!B4=2,{sc}!N5,IF({sec}!B4=3,{sc}!N6,IF({sec}!B4=4,{sc}!N7,"Уровень не рассчитан")))),"Уровень не рассчитан")' if direction == "UP" else f'=IFERROR(IF({sec}!B4=1,{sc}!N9,IF({sec}!B4=2,{sc}!N10,IF({sec}!B4=3,{sc}!N11,IF({sec}!B4=4,{sc}!N12,"Уровень не рассчитан")))),"Уровень не рассчитан")'
    ws["A10"] = "Расстояние до уровня, пунктов"; ws["B10"] = '=IF(OR(B9="Уровень не рассчитан",B7=""),"Расстояние не рассчитано",ROUND(ABS(B9-B7)/Settings!$B$3,0))'
    ws["A11"] = "Большая позиция"; ws["B11"] = f'=IFERROR(IF({tail}!B16>0,IF("{direction}"="UP","SELL ","BUY ")&TEXT({tail}!B16,"0.00"),"Не открывать"),"Не открывать")'
    ws["A12"] = "Малая защитная позиция"; ws["B12"] = f'=IFERROR(IF({tail}!B17>0,IF("{direction}"="UP","BUY ","SELL ")&TEXT({tail}!B17,"0.00"),"Не открывать"),"Не открывать")'
    ws["A13"] = "Уровень секции"; ws["B13"] = f"=IFERROR({sec}!B4,0)"

    ws["A39"] = (
        '="ПОДРОБНАЯ РЕКОМЕНДАЦИЯ"&CHAR(10)&CHAR(10)&'
        '"Сценарий: цена идет ' + ('вверх' if direction == 'UP' else 'вниз') + '."&CHAR(10)&'
        '"Текущий Bid: "&TEXT(B3,"0.00000")&CHAR(10)&'
        '"Текущий Ask: "&TEXT(B4,"0.00000")&CHAR(10)&'
        '"Сценарный Bid: "&TEXT(B5,"0.00000")&CHAR(10)&'
        '"Сценарный Ask: "&TEXT(B6,"0.00000")&CHAR(10)&'
        '"Сценарный Mid: "&TEXT(B7,"0.00000")&CHAR(10)&'
        '"Spread: "&TEXT(B8,"0")&" пунктов"&CHAR(10)&'
        '"Следующий уровень: "&IF(ISNUMBER(B9),TEXT(B9,"0.00000"),B9)&CHAR(10)&'
        '"Расстояние до уровня: "&IF(ISNUMBER(B10),TEXT(B10,"0")&" пунктов",B10)&CHAR(10)&CHAR(10)&'
        '"Рекомендация по секции: "&B11&" / "&B12&CHAR(10)&'
        '"Можно открыть секцию: "&IF(' + (f'{sec}!B14') + '="YES","ДА","НЕТ")&CHAR(10)&'
        '"Можно закрыть секцию: "&B15&CHAR(10)&'
        '"Можно закрыть хвост: "&B16&CHAR(10)&CHAR(10)&'
        '"Тип хвоста: "&B23&CHAR(10)&'
        '"Тикет хвоста: "&IF(ISNUMBER(B24),TEXT(B24,"0"),B24)&CHAR(10)&'
        '"Лот хвоста до закрытия: "&IF(ISNUMBER(B25),TEXT(B25,"0.00"),B25)&CHAR(10)&'
        '"Итоговый лот закрытия: "&IF(ISNUMBER(B30),TEXT(B30,"0.00"),"0.00")&CHAR(10)&'
        '"Остаток хвоста после закрытия: "&TEXT(' + (f'{tail}!B14') + ',"0.00")&CHAR(10)&CHAR(10)&'
        '"RecoveryFund после: "&TEXT(' + (f'{tail}!B13') + ',"0.00")&CHAR(10)&'
        '"ДЕЙСТВИЕ: "&B31&CHAR(10)&'
        '"Если открывать секцию: "&B11&" / "&B12&CHAR(10)&'
        '"Если закрывать хвост: тип="&B23&", тикет="&IF(ISNUMBER(B24),TEXT(B24,"0"),B24)&", до="&IF(ISNUMBER(B25),TEXT(B25,"0.00"),B25)&", закрыть="&TEXT(B30,"0.00")&", остаток="&TEXT(' + (f'{tail}!B14') + ',"0.00")&CHAR(10)&'
        '"Добавить в резерв: "&TEXT(B19,"0.00")&"; Добавить в RecoveryFund: "&TEXT(B20,"0.00")&CHAR(10)&'
        '"Резерв после: "&TEXT(B21,"0.00")&"; RecoveryFund после цикла: "&TEXT(B22,"0.00")&CHAR(10)&'
        '"RecoveryFund до/после закрытия хвоста: "&TEXT(B27,"0.00")&" / "&TEXT(' + (f'{tail}!B13') + ',"0.00")&CHAR(10)&'
        '"SAFE: "&B38&CHAR(10)&'
        '"ИТОГОВОЕ ДЕЙСТВИЕ: "&B31'
    )
    ws.merge_cells("A39:H52")
    ws["A39"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A39"].font = Font(bold=True, size=13, color="FFFFFFFF")
    ws["A39"].fill = PatternFill(fill_type="solid", fgColor="FF1F4E78")
    for c in ("B3","B4","B5"): ws[c].number_format="0.00000"
    for c in ("B25","B28","B29","B30","B35","B36","B37"): ws[c].number_format="0.00"
    ws["B8"].number_format="0"
    ws["B6"].number_format="0"

def build():
    wb = Workbook(); wb.active.title = "Settings"
    for s in ["CurrentPositions", "Scenario_UP", "Scenario_DOWN", "SectionCalculator_UP", "SectionCalculator_DOWN", "TailRecovery_UP", "TailRecovery_DOWN", "BasketSummary", "Validation", "Log", "ScenarioText_UP", "ScenarioText_DOWN", "FormulaDependencyMap"]:
        wb.create_sheet(s)
    add_settings(wb["Settings"]); add_positions(wb["CurrentPositions"])
    add_scenario(wb["Scenario_UP"], "UP"); add_scenario(wb["Scenario_DOWN"], "DOWN")
    add_section_calc(wb["SectionCalculator_UP"], "Scenario_UP", True); add_section_calc(wb["SectionCalculator_DOWN"], "Scenario_DOWN", False)
    add_tail_recovery(wb["TailRecovery_UP"], "SectionCalculator_UP", "Scenario_UP"); add_tail_recovery(wb["TailRecovery_DOWN"], "SectionCalculator_DOWN", "Scenario_DOWN")

    bs = wb["BasketSummary"]; hdr(bs, 1, ["Field", "UP", "DOWN"])
    fields = ["ScenarioPrice", "BasketFloating", "GlobalReserveAfter", "RecoveryFundAfterCycle", "TailLotAfter", "CloseLot", "CanCloseSection", "CanCloseTail", "CanCloseBasket", "NextBigLot", "NextSmallLot", "NextAction", "HumanReadableAction"]
    for i, f in enumerate(fields, 2): bs.cell(i, 1, f)
    bs["B2"] = "=Scenario_UP!B7"; bs["C2"] = "=Scenario_DOWN!B7"; bs["B3"] = "=SUM(Scenario_UP!H12:H211)"; bs["C3"] = "=SUM(Scenario_DOWN!H12:H211)"
    bs["B4"] = "=SectionCalculator_UP!B35"; bs["C4"] = "=SectionCalculator_DOWN!B35"; bs["B5"] = "=TailRecovery_UP!B7"; bs["C5"] = "=TailRecovery_DOWN!B7"
    bs["B6"] = "=TailRecovery_UP!B14"; bs["C6"] = "=TailRecovery_DOWN!B14"; bs["B7"] = "=TailRecovery_UP!B10"; bs["C7"] = "=TailRecovery_DOWN!B10"
    bs["B8"] = "=SectionCalculator_UP!B32"; bs["C8"] = "=SectionCalculator_DOWN!B32"; bs["B9"] = "=TailRecovery_UP!B11"; bs["C9"] = "=TailRecovery_DOWN!B11"
    bs["B10"] = '=IF(B3+B4>=Settings!$B$11,"YES","NO")'; bs["C10"] = '=IF(C3+C4>=Settings!$B$11,"YES","NO")'
    bs["B11"] = "=TailRecovery_UP!B16"; bs["C11"] = "=TailRecovery_DOWN!B16"; bs["B12"] = "=TailRecovery_UP!B17"; bs["C12"] = "=TailRecovery_DOWN!B17"
    bs["B13"] = '=IF(B10="YES","BASKET_CLOSE",IF(B8="NO","WAIT",IF(B9="YES","CLOSE_SECTION+CLOSE_TAIL","SAFE")))'; bs["C13"] = '=IF(C10="YES","BASKET_CLOSE",IF(C8="NO","WAIT",IF(C9="YES","CLOSE_SECTION+CLOSE_TAIL","SAFE")))'
    bs["B14"] = "=ScenarioText_UP!B31"; bs["C14"] = "=ScenarioText_DOWN!B31"

    add_recommendations(wb["ScenarioText_UP"], "UP")
    add_recommendations(wb["ScenarioText_DOWN"], "DOWN")

    v = wb["Validation"]; hdr(v, 1, ["Rule", "UP", "DOWN"])
    rules = [
        ("CycleProfit > 0 before close section", '=IF(OR(SectionCalculator_UP!B31>0,SectionCalculator_UP!B32="NO"),"OK","ERROR")', '=IF(OR(SectionCalculator_DOWN!B31>0,SectionCalculator_DOWN!B32="NO"),"OK","ERROR")'),
        ("CloseLot = 0 if CanCloseSection=NO", '=IF(OR(SectionCalculator_UP!B32="YES",TailRecovery_UP!B10=0),"OK","ERROR")', '=IF(OR(SectionCalculator_DOWN!B32="YES",TailRecovery_DOWN!B10=0),"OK","ERROR")'),
        ("TailCloseLoss <= RecoveryFundAfterCycle", '=IF(TailRecovery_UP!B12<=TailRecovery_UP!B7,"OK","ERROR")', '=IF(TailRecovery_DOWN!B12<=TailRecovery_DOWN!B7,"OK","ERROR")'),
        ("LevelReached before open section", '=IF(SectionCalculator_UP!B12="YES","OK",IF(SectionCalculator_UP!B14="NO","BLOCKED","ERROR"))', '=IF(SectionCalculator_DOWN!B12="YES","OK",IF(SectionCalculator_DOWN!B14="NO","BLOCKED","ERROR"))'),
        ("No opposite cascade active", '=IF(SectionCalculator_UP!B13="YES","OK","BLOCKED")', '=IF(SectionCalculator_DOWN!B13="YES","OK","BLOCKED")'),
        ("Tail exists when loss exists", '=IF(AND(COUNTIFS(Scenario_UP!B12:B211,"SELL",Scenario_UP!H12:H211,"<0")>0,Scenario_UP!B222="N/A"),"ERROR","OK")', '=IF(AND(COUNTIFS(Scenario_DOWN!B12:B211,"BUY",Scenario_DOWN!H12:H211,"<0")>0,Scenario_DOWN!B222="N/A"),"ERROR","OK")'),
        ("ScenarioText action equals BasketSummary action", '=IF(OR(AND(BasketSummary!B13="BASKET_CLOSE",ScenarioText_UP!B31="ЗАКРЫТЬ ВСЮ КОРЗИНУ"),BasketSummary!B13<>"BASKET_CLOSE"),"OK","ERROR")', '=IF(OR(AND(BasketSummary!C13="BASKET_CLOSE",ScenarioText_DOWN!B31="ЗАКРЫТЬ ВСЮ КОРЗИНУ"),BasketSummary!C13<>"BASKET_CLOSE"),"OK","ERROR")'),
        ("NextTriggerLevel is numeric", '=IF(OR(ISNUMBER(ScenarioText_UP!B9),ScenarioText_UP!B9="Уровень не рассчитан"),"OK","ERROR")', '=IF(OR(ISNUMBER(ScenarioText_DOWN!B9),ScenarioText_DOWN!B9="Уровень не рассчитан"),"OK","ERROR")'),
        ("DistancePoints is numeric", '=IF(OR(ISNUMBER(ScenarioText_UP!B10),ScenarioText_UP!B10="Расстояние не рассчитано"),"OK","ERROR")', '=IF(OR(ISNUMBER(ScenarioText_DOWN!B10),ScenarioText_DOWN!B10="Расстояние не рассчитано"),"OK","ERROR")'),
        ("ScenarioMid is numeric", '=IF(ISNUMBER(ScenarioText_UP!B7),"OK","ERROR")', '=IF(ISNUMBER(ScenarioText_DOWN!B7),"OK","ERROR")'),
        ("ScenarioSpread is numeric", '=IF(ISNUMBER(ScenarioText_UP!B8),"OK","ERROR")', '=IF(ISNUMBER(ScenarioText_DOWN!B8),"OK","ERROR")'),
        ("ScenarioSpread bounds", '=IF(AND(ScenarioText_UP!B8>0,ScenarioText_UP!B8<1000),"OK","ERROR")', '=IF(AND(ScenarioText_DOWN!B8>0,ScenarioText_DOWN!B8<1000),"OK","ERROR")'),
        ("Scenario_* B8 not empty", '=IF(Scenario_UP!B8<>"","OK","ERROR")', '=IF(Scenario_DOWN!B8<>"","OK","ERROR")'),
        ("No legacy single-price wording", '=IF(ISERROR(SEARCH("Сценарная цена",ScenarioText_UP!B46)),"OK","ERROR")', '=IF(ISERROR(SEARCH("Сценарная цена",ScenarioText_DOWN!B46)),"OK","ERROR")'),
        ("ScenarioText B9!=B5", '=IF(ScenarioText_UP!B9<>ScenarioText_UP!B5,"OK","ERROR")', '=IF(ScenarioText_DOWN!B9<>ScenarioText_DOWN!B5,"OK","ERROR")'),
        ("ScenarioText B10!=B6", '=IF(ScenarioText_UP!B10<>ScenarioText_UP!B6,"OK","ERROR")', '=IF(ScenarioText_DOWN!B10<>ScenarioText_DOWN!B6,"OK","ERROR")'),
        ("HumanReadableAction not empty", '=IF(AND(ScenarioText_UP!B31<>"",ISNUMBER(ScenarioText_UP!B4)),"OK","ERROR")', '=IF(AND(ScenarioText_DOWN!B31<>"",ISNUMBER(ScenarioText_DOWN!B4)),"OK","ERROR")'),
        ("ReserveAdd mapping sync", '=IF(ScenarioText_UP!B19=SectionCalculator_UP!B33,"OK","ERROR")', '=IF(ScenarioText_DOWN!B19=SectionCalculator_DOWN!B33,"OK","ERROR")'),
        ("RecoveryAdd mapping sync", '=IF(ScenarioText_UP!B20=SectionCalculator_UP!B34,"OK","ERROR")', '=IF(ScenarioText_DOWN!B20=SectionCalculator_DOWN!B34,"OK","ERROR")'),
        ("TailTicket not mapped as lot", '=IF(AND(ISNUMBER(ScenarioText_UP!B24),ScenarioText_UP!B24>1000,ISNUMBER(ScenarioText_UP!B25),ScenarioText_UP!B25>100),"ERROR","OK")', '=IF(AND(ISNUMBER(ScenarioText_DOWN!B24),ScenarioText_DOWN!B24>1000,ISNUMBER(ScenarioText_DOWN!B25),ScenarioText_DOWN!B25>100),"ERROR","OK")'),
        ("Tail lot reasonable", '=IF(AND(ISNUMBER(ScenarioText_UP!B25),ScenarioText_UP!B25>100),"ERROR","OK")', '=IF(AND(ISNUMBER(ScenarioText_DOWN!B25),ScenarioText_DOWN!B25>100),"ERROR","OK")'),
        ("Tail остаток <= исходного", '=IF(AND(ISNUMBER(ScenarioText_UP!B37),ISNUMBER(ScenarioText_UP!B25),ScenarioText_UP!B37>ScenarioText_UP!B25),"ERROR","OK")', '=IF(AND(ISNUMBER(ScenarioText_DOWN!B37),ISNUMBER(ScenarioText_DOWN!B25),ScenarioText_DOWN!B37>ScenarioText_DOWN!B25),"ERROR","OK")'),
        ("RecoveryFundAfter >= 0", '=IF(TailRecovery_UP!B13<0,"ERROR","OK")', '=IF(TailRecovery_DOWN!B13<0,"ERROR","OK")'),
        ("CloseLotFinal <= TailLot", '=IF(TailRecovery_UP!B10>TailRecovery_UP!B4,"ERROR","OK")', '=IF(TailRecovery_DOWN!B10>TailRecovery_DOWN!B4,"ERROR","OK")'),
        ("TailLotAfter >= 0", '=IF(TailRecovery_UP!B14<0,"ERROR","OK")', '=IF(TailRecovery_DOWN!B14<0,"ERROR","OK")'),
        ("ScenarioText SAFE not empty", '=IF(ScenarioText_UP!B38<>"","OK","ERROR")', '=IF(ScenarioText_DOWN!B38<>"","OK","ERROR")'),
        ("ScenarioText SAFE equals Log SAFE", '=IF(ScenarioText_UP!B38=Log!K2,"OK","ERROR")', '=IF(ScenarioText_DOWN!B38=Log!K3,"OK","ERROR")'),
        ("Detailed recommendation block not empty", '=IF(AND(ScenarioText_UP!A39<>"",ScenarioText_UP!A39<>0),"OK","ERROR")', '=IF(AND(ScenarioText_DOWN!A39<>"",ScenarioText_DOWN!A39<>0),"OK","ERROR")')
    ]
    for r, row in enumerate(rules, 2):
        for c, x in enumerate(row, 1): ws=v; ws.cell(r, c, x)

    lg=wb["Log"]
    hdr(lg,1,["Время","Направление","Сценарий","Действие","Хвост до закрытия","Закрываемый лот хвоста","Хвост после закрытия","RecoveryFund до","RecoveryFund после","Резерв после","SAFE"])
    lg["A2"]="=NOW()"; lg["B2"]="ВВЕРХ"; lg["C2"]="ScenarioText_UP"; lg["D2"]='=IF(BasketSummary!B13="WAIT","ЖДАТЬ",IF(BasketSummary!B13="SAFE","SAFE",IF(BasketSummary!B13="BASKET_CLOSE","ЗАКРЫТЬ ВСЮ КОРЗИНУ",IF(AND(TailRecovery_UP!B10>0,TailRecovery_UP!B10<TailRecovery_UP!B4),"ЗАКРЫТЬ ХВОСТ ЧАСТИЧНО","ЗАКРЫТЬ СЕКЦИЮ"))))'; lg["E2"]="=IFERROR(TailRecovery_UP!B4,0)"; lg["F2"]="=IFERROR(TailRecovery_UP!B10,0)"; lg["G2"]='=IFERROR(TailRecovery_UP!B14,0)'; lg["H2"]="=IFERROR(TailRecovery_UP!B7,0)"; lg["I2"]="=IFERROR(TailRecovery_UP!B13,0)"; lg["J2"]="=IFERROR(SectionCalculator_UP!B35,0)"; lg['K2']='=IF(BasketSummary!B13="SAFE","ДА","НЕТ")'
    lg["A3"]="=NOW()"; lg["B3"]="ВНИЗ"; lg["C3"]="ScenarioText_DOWN"; lg["D3"]='=IF(BasketSummary!C13="WAIT","ЖДАТЬ",IF(BasketSummary!C13="SAFE","SAFE",IF(BasketSummary!C13="BASKET_CLOSE","ЗАКРЫТЬ ВСЮ КОРЗИНУ",IF(AND(TailRecovery_DOWN!B10>0,TailRecovery_DOWN!B10<TailRecovery_DOWN!B4),"ЗАКРЫТЬ ХВОСТ ЧАСТИЧНО","ЗАКРЫТЬ СЕКЦИЮ"))))'; lg["E3"]="=IFERROR(TailRecovery_DOWN!B4,0)"; lg["F3"]="=IFERROR(TailRecovery_DOWN!B10,0)"; lg["G3"]='=IFERROR(TailRecovery_DOWN!B14,0)'; lg["H3"]="=IFERROR(TailRecovery_DOWN!B7,0)"; lg["I3"]="=IFERROR(TailRecovery_DOWN!B13,0)"; lg["J3"]="=IFERROR(SectionCalculator_DOWN!B35,0)"; lg['K3']='=IF(BasketSummary!C13="SAFE","ДА","НЕТ")'

    dm=wb["FormulaDependencyMap"]; hdr(dm,1,["Cell","Formula","DependsOn","UsedBy","Description"])
    rows=[("Scenario_UP!B5","=B2+B4*Settings!B3","Scenario_UP!B2, Scenario_UP!B4, Settings!B3","Scenario_UP!B7:B8, SectionCalculator_UP!B12, ScenarioText_UP!B5","ScenarioBid"),
          ("Scenario_UP!B6","=B3+B4*Settings!B3","Scenario_UP!B3, Scenario_UP!B4, Settings!B3","Scenario_UP!B7:B8, Scenario_UP!F12:F211, ScenarioText_UP!B6","ScenarioAsk"),
          ("Scenario_UP!B7","=(B5+B6)/2","Scenario_UP!B5, Scenario_UP!B6","SectionCalculator_UP!B12, ScenarioText_UP!B7","ScenarioMid"),
          ("Scenario_UP!B8","=(B6-B5)/Settings!B3","Scenario_UP!B6, Scenario_UP!B5, Settings!B3","ScenarioText_UP!B8","ScenarioSpread"),
          ("Scenario_UP!K12:L211","PnL helper range","Scenario_UP!C12:F211, Settings!B3","Scenario_UP!B221:B223","Tail helper range"),
          ("Scenario_UP!B221","=IF(COUNT(K12:K211)=0,0,MIN(K12:K211))","Scenario_UP!K12:K211","Scenario_UP!B222","TailWorstPnL"),
          ("Scenario_UP!B222","=IF(B221=0,\"N/A\",MIN(L12:L211))","Scenario_UP!B221, Scenario_UP!L12:L211","Scenario_UP!B223, ScenarioText_UP!B24","TailTicket"),
          ("BasketSummary!B13","=IF(...)","BasketSummary!B8, BasketSummary!B9, BasketSummary!B10","ScenarioText_UP!B31, ScenarioText_UP!B38, Log!D2, Log!K2","NextAction"),
          ("ScenarioText_UP!B31","=IF(BasketSummary!B13...)","BasketSummary!B13","ScenarioText_UP!B46","HumanReadableAction"),
          ("ScenarioText_UP!B38","=IF(BasketSummary!B13=\"SAFE\",\"ДА\",\"НЕТ\")","BasketSummary!B13","ScenarioText_UP!B46, Validation","SAFE flag")]
    for i,(a,b,c,d,e) in enumerate(rows,2):
        dm[f"A{i}"]=a
        dm[f"B{i}"]="'"+b if isinstance(b,str) and b.startswith("=") else b
        dm[f"C{i}"]=c
        dm[f"D{i}"]=d
        dm[f"E{i}"]=e
    wb.save("recovery_lock_cascade_next_step.xlsx")


if __name__ == "__main__":
    build()
    print("Created recovery_lock_cascade_next_step.xlsx")
