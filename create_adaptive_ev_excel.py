from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws_in = wb.active
ws_in.title = "Ввод"
ws_pos = wb.create_sheet("Позиции")
ws_calc = wb.create_sheet("Расчеты")
ws_next = wb.create_sheet("Следующий шаг")
ws_open = wb.create_sheet("Новые позиции")

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
sub_fill = PatternFill("solid", fgColor="D9E1F2")
warn_fill = PatternFill("solid", fgColor="FCE4D6")

ws_in["A1"] = "Калькулятор Adaptive EV"
ws_in["A1"].font = Font(bold=True, size=14)

inputs = [
    ("Брокер", "DemoBroker"), ("Символ", "EURUSD"), ("Знаков после запятой", 5),
    ("Размер контракта", 100000), ("Кредитное плечо", 100), ("Валюта счета", "USD"),
    ("Баланс", 100000), ("Эквити", 99998.70), ("Свободная маржа", 99875.50),
    ("Текущая цена (mid)", 1.23184), ("Bid", "=B12-B15/2"), ("Ask", "=B12+B15/2"),
    ("Спред (цена)", 0.00008), ("ATR", 0.00080), ("Alpha (α)", 0.30),
    ("Шаг Δ", 0.00028), ("Gamma (γ)", 0.08), ("Минимум SELL Ls,min", 0.25),
    ("Предел просадки Dmax, %", 2.0), ("Лимит нагрузки маржи, %", 60),
    ("Коэффициент режима (0.5..2.0)", 1.00), ("k_min", 1.20), ("k_max", 2.20),
    ("Lambda k(ATR)", 1.00), ("Emax (лот)", 0.50), ("Смещение bias, %", 5.0),
]
for i, (k, v) in enumerate(inputs, start=3):
    ws_in[f"A{i}"] = k
    ws_in[f"B{i}"] = v
    ws_in[f"A{i}"].fill = sub_fill
ws_in["A25"], ws_in["B25"] = "Уровень маржи, %", "=IF(Расчеты!B13=0,0,B10/Расчеты!B13*100)"
ws_in["A26"], ws_in["B26"] = "Статус риска", "=IF(OR(Расчеты!B9=\"FAIL\",Расчеты!B14>B22,Расчеты!B16<-B21),\"СТРЕСС/СОКРАЩЕНИЕ\",\"НОРМА\")"
ws_in["A27"], ws_in["B27"] = "Состояние FSM", "=Расчеты!B22"
ws_in["A25"].fill = sub_fill
ws_in["A26"].fill = warn_fill
ws_in["A27"].fill = sub_fill

headers = ["ID","Активна (Y/N)","Направление (BUY/SELL)","Лот","Цена открытия","SL","TP","Комиссия","Своп","Текущая цена","Разница цены","P/L, USD","Маржа, USD","Лот частичного закрытия","Действие","Расчет ΔL_s/ΔL_b","Ключ актив+направление"]
for c, h in enumerate(headers, 1):
    cell = ws_pos.cell(1, c, h)
    cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center")

samples = [(1,"Y","SELL",0.10,1.23125,0,0,0,0),(2,"Y","BUY",0.10,1.23267,0,0,0,0)]
for r, data in enumerate(samples, 2):
    for c, v in enumerate(data, 1): ws_pos.cell(r, c, v)

for r in range(2, 202):
    ws_pos[f"J{r}"] = f"=IF(C{r}=\"BUY\",Ввод!B13,Ввод!B14)"
    ws_pos[f"K{r}"] = f"=IF(C{r}=\"BUY\",J{r}-E{r},E{r}-J{r})"
    ws_pos[f"L{r}"] = f"=IF(B{r}=\"Y\",K{r}*D{r}*Ввод!B6+H{r}+I{r},0)"
    ws_pos[f"M{r}"] = f"=IF(B{r}=\"Y\",D{r}*Ввод!B6*J{r}/Ввод!B7,0)"
    ws_pos[f"N{r}"] = f"=IF(B{r}=\"Y\",ROUND(MIN(D{r}*Ввод!B19,D{r}*0.5),2),0)"
    ws_pos[f"O{r}"] = f"=IF(B{r}<>\"Y\",\"НЕТ\",IF(AND(Расчеты!B11=\"ДА\",ABS(K{r})>=Ввод!B18),\"ЧАСТИЧНО ЗАКРЫТЬ\",\"ДЕРЖАТЬ\"))"
    ws_pos[f"P{r}"] = f"=IF(B{r}<>\"Y\",0,IFERROR((Ввод!B12-Расчеты!B3-Ввод!B15/2+Расчеты!B6)/(Расчеты!B4-Ввод!B12-Ввод!B15/2-Расчеты!B6),0))"
    ws_pos[f"Q{r}"] = f"=IF(B{r}=\"Y\",B{r}&C{r},\"\")"
    ws_pos[f"D{r}"].number_format = ws_pos[f"N{r}"].number_format = "0.00"

calc_rows = [
    ("Сумма BUY лотов", "=SUMIFS(Позиции!D2:D201,Позиции!C2:C201,\"BUY\",Позиции!B2:B201,\"Y\")"),
    ("Сумма SELL лотов", "=SUMIFS(Позиции!D2:D201,Позиции!C2:C201,\"SELL\",Позиции!B2:B201,\"Y\")"),
    ("Средняя цена BUY", "=IF(B1=0,0,SUMPRODUCT((Позиции!C2:C201=\"BUY\")*(Позиции!B2:B201=\"Y\")*Позиции!D2:D201*Позиции!E2:E201)/B1)"),
    ("Средняя цена SELL", "=IF(B2=0,0,SUMPRODUCT((Позиции!C2:C201=\"SELL\")*(Позиции!B2:B201=\"Y\")*Позиции!D2:D201*Позиции!E2:E201)/B2)"),
    ("P_avg", "=IF(B1+B2=0,0,(B1*B3+B2*B4)/(B1+B2))"),
    ("δ = S + α*ATR", "=Ввод!B15 + Ввод!B17*Ввод!B16"),
    ("Левая часть survival: δ*Ls", "=B6*B2"), ("Правая часть survival: Δ*Lb", "=Ввод!B18*B1"),
    ("Условие survival", "=IF(B7>B8,\"OK\",\"FAIL\")"), ("|P-P_avg|", "=ABS(Ввод!B12-B5)"),
    ("Триггер |P-P_avg|>Δ", "=IF(B10>Ввод!B18,\"ДА\",\"НЕТ\")"), ("Общий плавающий P/L", "=SUM(Позиции!L2:L201)"),
    ("Общая маржа", "=SUM(Позиции!M2:M201)"), ("Нагрузка на депозит, %", "=IF(Ввод!B9=0,0,B13/Ввод!B9*100)"),
    ("Чистая экспозиция (BUY-SELL)", "=B1-B2"), ("Текущая просадка от баланса, %", "=IF(Ввод!B9=0,0,(B12/Ввод!B9)*100)"),
    ("ID первой активной BUY", "=IFERROR(INDEX(Позиции!A$2:A$201,MATCH(\"YBUY\",Позиции!Q$2:Q$201,0)),\"НЕТ\")"),
    ("ID первой активной SELL", "=IFERROR(INDEX(Позиции!A$2:A$201,MATCH(\"YSELL\",Позиции!Q$2:Q$201,0)),\"НЕТ\")"),
    ("Лот выбранной BUY", "=IFERROR(INDEX(Позиции!D$2:D$201,MATCH(B17,Позиции!A$2:A$201,0)),0)"),
    ("Лот выбранной SELL", "=IFERROR(INDEX(Позиции!D$2:D$201,MATCH(B18,Позиции!A$2:A$201,0)),0)"),
    ("Экспозиция E=|Lb-Ls|", "=ABS(B1-B2)"),
    ("Норм. волатильность v=ATR/Δ", "=IF(Ввод!B18=0,0,Ввод!B16/Ввод!B18)"),
    ("k(ATR)", "=Ввод!B24+(Ввод!B25-Ввод!B24)*EXP(-Ввод!B26*B24)"),
    ("k адаптивный", "=MAX(Ввод!B24,MIN(Ввод!B25,B23*(1-0.5*B22/MAX(Ввод!B27,0.0001))*(1-0.5*MAX(0,-B16)/MAX(Ввод!B21,0.0001))*(1-0.4*B2/MAX(1,1))))"),
    ("Режим FSM", "=IF(OR(B14>Ввод!B22,B16<-Ввод!B21),\"ESCAPE\",IF(OR(B23>0.7*Ввод!B27,B14>0.7*Ввод!B22),\"STRESS\",\"FLOW\"))"),
]
for i,(k,v) in enumerate(calc_rows,1):
    ws_calc[f"A{i}"]=k; ws_calc[f"B{i}"]=v; ws_calc[f"A{i}"].fill=sub_fill

next_headers=["Сценарий","Целевая цена","ID Позиции","Для BUY/SELL","Направление шага","Рекоменд. лот","Действие","Частично закрыть лот","Полностью закрыть?","Причина","Открыть позицию?","Направление открытия","Лот открытия","Цена открытия"]
for c,h in enumerate(next_headers,1):
    cell=ws_next.cell(1,c,h); cell.fill=header_fill; cell.font=header_font

rows=[
 (2,"Цена вверх на Δ","=Ввод!B12+Ввод!B18","BUY","BUY (Рекомендации для выбранной BUY)","BUY trim / SELL hedge","=IF(Расчеты!B17=\"НЕТ\",0,MAX(0.01,ROUND(Расчеты!B19*Ввод!B19*Ввод!B23*0.5,2)))"),
 (3,"Цена вверх на Δ","=Ввод!B12+Ввод!B18","SELL","SELL (Рекомендации для выбранной SELL)","SELL partial","=IF(Расчеты!B18=\"НЕТ\",0,MAX(0.01,ROUND(Расчеты!B20*Ввод!B19*Ввод!B23*2,2)))"),
 (4,"Цена вниз на Δ","=Ввод!B12-Ввод!B18","BUY","BUY (Рекомендации для выбранной BUY)","BUY partial","=IF(Расчеты!B17=\"НЕТ\",0,MAX(0.01,ROUND(Расчеты!B19*Ввод!B19*Ввод!B23*2,2)))"),
 (5,"Цена вниз на Δ","=Ввод!B12-Ввод!B18","SELL","SELL (Рекомендации для выбранной SELL)","SELL trim / BUY hedge","=IF(Расчеты!B18=\"НЕТ\",0,MAX(0.01,ROUND(Расчеты!B20*Ввод!B19*Ввод!B23*0.5,2)))")]

for r,sc,target,side,label,dirn,rec in rows:
    ws_next[f"A{r}"]=sc; ws_next[f"B{r}"]=target
    ws_next[f"C{r}"]='=IF(D{0}="BUY (Рекомендации для выбранной BUY)",Расчеты!B17,Расчеты!B18)'.format(r)
    ws_next[f"D{r}"]=label; ws_next[f"E{r}"]=dirn; ws_next[f"F{r}"]=rec
    ws_next[f"G{r}"]='=IF(AND(Расчеты!B11="ДА",F{0}>0),"ЧАСТИЧНАЯ РЕБАЛАНСИРОВКА","ОЖИДАТЬ")'.format(r)
    ws_next[f"H{r}"]='=IF(F{0}=0,0,IF(OR(E{0}="BUY partial",E{0}="SELL partial"),ROUND(F{0},2),ROUND(F{0}*0.6,2)))'.format(r)
    ws_next[f"I{r}"]='=IF(OR(Расчеты!B9="FAIL",Расчеты!B14>Ввод!B22,Расчеты!B16<-Ввод!B21),"ДА","НЕТ")'
    ws_next[f"J{r}"]='=IF(I{0}="ДА","Риск превышен: survival/маржа/просадка","Нормальный цикл")'.format(r)

    ws_next[f"K{r}"] = '=IF(AND(F{0}>0,I{0}="НЕТ"),"ДА","НЕТ")'.format(r)
    ws_next[f"L{r}"] = '=IF(K{0}="НЕТ","-",IF(D{0}="BUY (Рекомендации для выбранной BUY)","BUY","SELL"))'.format(r)
    ws_next[f"M{r}"] = '=IF(K{0}="НЕТ",0,ROUND(MAX(0.01,F{0}),2))'.format(r)
    ws_next[f"N{r}"] = '=IF(K{0}="НЕТ",0,B{0})'.format(r)
    ws_next[f"F{r}"].number_format=ws_next[f"H{r}"].number_format="0.00"
    ws_next[f"M{r}"].number_format="0.00"
    ws_next[f"N{r}"].number_format="0.00000"

open_headers = [
    "Сценарий", "Триггер", "Рекомендация открытия", "Направление",
    "Рекомендуемый лот", "Рекомендуемая цена", "Оценка маржи, USD",
    "Свободная маржа после открытия, USD", "Комментарий"
]
for c, h in enumerate(open_headers, 1):
    cell = ws_open.cell(1, c, h)
    cell.fill = header_fill
    cell.font = header_font

open_rows = [
    (2, "Пробой вверх (+Δ)", 'Ввод!B12>=Расчеты!B5+Ввод!B18', "ОТКРЫТЬ"),
    (3, "Пробой вниз (-Δ)", 'Ввод!B12<=Расчеты!B5-Ввод!B18', "ОТКРЫТЬ"),
]
for r, scenario, trigger_formula, action in open_rows:
    ws_open[f"A{r}"] = scenario
    ws_open[f"B{r}"] = f'=IF({trigger_formula},"ДА","НЕТ")'
    ws_open[f"C{r}"] = f'=IF(B{r}="ДА","{action}","ЖДАТЬ")'
    ws_open[f"D{r}"] = '=IF(A{0}="Пробой вверх (+Δ)","BUY","SELL")'.format(r)
    ws_open[f"E{r}"] = (
        '=IF(C{0}<>"ОТКРЫТЬ",0,ROUND(MAX(0.01,MIN('
        'Ввод!B20*0.01*Ввод!B23,'
        '(Ввод!B10*Ввод!B22/100-Расчеты!B13)*Ввод!B7/(Ввод!B6*Ввод!B12)'
        ')),2))'
    ).format(r)
    ws_open[f"F{r}"] = '=IF(E{0}=0,0,IF(D{0}="BUY",Ввод!B14-Ввод!B18*(Ввод!B28/100),Ввод!B13+Ввод!B18*(Ввод!B28/100)))'.format(r)
    ws_open[f"G{r}"] = '=IF(E{0}=0,0,E{0}*Ввод!B6*F{0}/Ввод!B7)'.format(r)
    ws_open[f"H{r}"] = '=Ввод!B10-G{0}'.format(r)
    ws_open[f"I{r}"] = '=IF(E{0}=0,"Нет сигнала или лимит маржи","Открыть только при рыночном подтверждении")'.format(r)
    ws_open[f"E{r}"].number_format = "0.00"
    ws_open[f"F{r}"].number_format = "0.00000"
    ws_open[f"G{r}"].number_format = ws_open[f"H{r}"].number_format = "#,##0.00"

for ws in [ws_in,ws_pos,ws_calc,ws_next,ws_open]:
    for col in "ABCDEFGHIJKLMNOPQRST": ws.column_dimensions[col].width=28

wb.save("adaptive_ev_calculator.xlsx")
print("Создан файл adaptive_ev_calculator.xlsx")
