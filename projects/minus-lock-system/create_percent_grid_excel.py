from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import LineChart, Reference
from openpyxl.workbook.defined_name import DefinedName


def build_workbook(output_path: str) -> None:
    wb = Workbook()
    wb.active.title = "Settings"
    for name in ["DownTrend", "UpTrend", "Summary", "Checks", "Manual"]:
        wb.create_sheet(name)

    s = wb["Settings"]
    rows = [
        ("StartLot", 1.00), ("PointStep", 100), ("MaxLevels", 5), ("LotStep", 0.01),
        ("RoundMode", "Nearest"), ("TargetSkewMin%", 5), ("TargetSkewMax%", 25),
        ("UseRounding", True), ("Direction", "DOWN"),
        ("BigRoundMode_DOWN", "DOWN"), ("SmallRoundMode_DOWN", "UP"), ("CloseRoundMode_DOWN", "SAFE"),
        ("BigRoundMode_UP", "DOWN"), ("SmallRoundMode_UP", "UP"), ("CloseRoundMode_UP", "SAFE"),
        ("EnableRiskSafeRounding", True), ("EnableInputValidation", True),
    ]
    for r, (k, v) in enumerate(rows, 2):
        s[f"A{r}"] = k; s[f"A{r}"].font = Font(bold=True); s[f"B{r}"] = v

    for c, h in zip("ABCDE", ["Level", "Big%", "Small%", "TargetSkew%", "ManualClose%"]):
        s[f"{c}21"] = h; s[f"{c}21"].font = Font(bold=True)
    grid = [(1, 90, 30, 0, None), (2, 30, 15, 15, None), (3, 20, 15, 10, None), (4, 10, 10, 10, None), (5, 5, 5, 10, None)]
    for r, data in enumerate(grid, 22):
        for c, v in enumerate(data, 1): s.cell(r, c, v)

    names = [("LevelRange", "$A$22:$A$200"), ("BigPercentRange", "$B$22:$B$200"), ("SmallPercentRange", "$C$22:$C$200"), ("TargetSkewRange", "$D$22:$D$200"), ("ManualCloseRange", "$E$22:$E$200")]
    for n, ref in names: wb.defined_names.add(DefinedName(n, attr_text=f"Settings!{ref}"))

    def fill(ws, down=True):
        headers = ["Level","PriceStep","Big BUY %" if down else "Big SELL %","Big Lot","Small SELL %" if down else "Small BUY %","Small Lot","Start Before %","Before Close %","Opposite After Add %","Target Skew %","Auto Close %","Manual Close %","Final Close %","Start After %","Sum Big %","Sum Small %","Total Main %","Total Opp %","Skew %","Status","Comment","Big Raw Lot","Big Rounded","Small Raw Lot","Small Rounded","Close Raw Lot","Close Rounded","Safe Rounding Status","Rounding Comment"]
        for i,h in enumerate(headers,1): ws.cell(1,i,h).font=Font(bold=True)
        for r in range(2,42):
            ws[f"A{r}"]=0 if r==2 else r-2
            ws[f"B{r}"]=0 if r==2 else f"=A{r}*Settings!$B$3"
            ws[f"C{r}"]=0 if r==2 else f"=IF(A{r}<=Settings!$B$4,IFERROR(XLOOKUP(A{r},LevelRange,BigPercentRange),0),0)"
            ws[f"E{r}"]=0 if r==2 else f"=IF(A{r}<=Settings!$B$4,IFERROR(XLOOKUP(A{r},LevelRange,SmallPercentRange),0),0)"
            ws[f"V{r}"]=0 if r==2 else f"=Settings!$B$2*C{r}/100"
            ws[f"X{r}"]=0 if r==2 else f"=Settings!$B$2*E{r}/100"
            ws[f"W{r}"]=0 if r==2 else f"=IF(Settings!$B$17,IF(Settings!$B$9,FLOOR(V{r},Settings!$B$5),V{r}),IF(Settings!$B$9,MROUND(V{r},Settings!$B$5),V{r}))"
            ws[f"Y{r}"]=0 if r==2 else f"=IF(Settings!$B$17,IF(Settings!$B$9,CEILING(X{r},Settings!$B$5),X{r}),IF(Settings!$B$9,MROUND(X{r},Settings!$B$5),X{r}))"
            ws[f"D{r}"]=f"=W{r}"; ws[f"F{r}"]=f"=Y{r}"
            ws[f"G{r}"]=100 if r in (2,3) else f"=N{r-1}"
            ws[f"H{r}"]=100 if r==2 else f"=G{r}+SUM($C$3:C{r})"
            ws[f"I{r}"]=100 if r==2 else f"=100+SUM($E$3:E{r})"
            ws[f"J{r}"]=0 if r==2 else f"=IF(A{r}<=Settings!$B$4,IFERROR(XLOOKUP(A{r},LevelRange,TargetSkewRange),0),0)"
            ws[f"K{r}"]=0 if r==2 else f"=MIN(G{r},MAX(0,H{r}-I{r}+J{r}))"
            ws[f"L{r}"]="" if r==2 else f"=IF(A{r}<=Settings!$B$4,IFERROR(XLOOKUP(A{r},LevelRange,ManualCloseRange),\"\"),\"\")"
            ws[f"M{r}"]=0 if r==2 else f"=MIN(G{r},IF(LEN(L{r})=0,K{r},L{r}))"
            ws[f"Z{r}"]=0 if r==2 else f"=Settings!$B$2*M{r}/100"
            if down:
                ws[f"AA{r}"]=0 if r==2 else f"=IF(Settings!$B$17,MIN(Settings!$B$2*G{r}/100,IF(Settings!$B$9,CEILING(MAX(0,V{r}+Settings!$B$2*G{r}/100-X{r}-Settings!$B$2*J{r}/100),Settings!$B$5),MAX(0,V{r}+Settings!$B$2*G{r}/100-X{r}-Settings!$B$2*J{r}/100))),IF(Settings!$B$9,MROUND(Z{r},Settings!$B$5),Z{r}))"
            else:
                ws[f"AA{r}"]=0 if r==2 else f"=IF(Settings!$B$17,MIN(Settings!$B$2*G{r}/100,IF(Settings!$B$9,CEILING(MAX(0,V{r}+Settings!$B$2*G{r}/100-X{r}-Settings!$B$2*J{r}/100),Settings!$B$5),MAX(0,V{r}+Settings!$B$2*G{r}/100-X{r}-Settings!$B$2*J{r}/100))),IF(Settings!$B$9,MROUND(Z{r},Settings!$B$5),Z{r}))"
            ws[f"M{r}"]=0 if r==2 else f"=MIN(G{r},AA{r}/Settings!$B$2*100)"
            ws[f"N{r}"]=100 if r==2 else f"=G{r}-M{r}"
            ws[f"O{r}"]=0 if r==2 else f"=SUM($C$3:C{r})"
            ws[f"P{r}"]=0 if r==2 else f"=SUM($E$3:E{r})"
            ws[f"Q{r}"]=f"=N{r}+O{r}"; ws[f"R{r}"]=f"=100+P{r}"; ws[f"S{r}"]=f"=R{r}-Q{r}"
            berr = "Big BUY must be >= Small SELL" if down else "Big SELL must be >= Small BUY"
            balerr = "Rounding broke protection balance"
            bal_cond = f"Q{r}>R{r}"
            ws[f"T{r}"]=(
                f"=IF(Settings!$B$18=FALSE,\"OK\",IF(OR(NOT(ISNUMBER(Settings!$B$2)),Settings!$B$2<=0),\"ERROR: Invalid StartLot\","
                f"IF(OR(NOT(ISNUMBER(Settings!$B$5)),Settings!$B$5<=0),\"ERROR: Invalid LotStep\","
                f"IF(OR(Settings!$B$4<1,Settings!$B$4>20),\"ERROR: Invalid MaxLevels\","
                f"IF(AND(Settings!$B$10<>\"DOWN\",Settings!$B$10<>\"UP\"),\"ERROR: Invalid Direction\","
                f"IF(OR(C{r}<0,E{r}<0,J{r}<0,AND(LEN(L{r})>0,L{r}<0)),\"ERROR: Negative input\","
                f"IF(C{r}<E{r},\"ERROR: {berr}\",IF(AND(LEN(L{r})>0,L{r}>G{r}),\"ERROR: ManualClose exceeds remaining Start position\","
                f"IF(OR(N{r}<0,Q{r}<0,R{r}<0),\"ERROR: Negative totals\",IF({bal_cond},\"ERROR: {balerr}\",IF(S{r}>Settings!$B$8,\"WARNING: Skew exceeds recommended limit\",\"OK\")))))))))))"
            )
            ws[f"AB{r}"]=f"=IF(T{r}=\"ERROR: Rounding broke protection balance\",\"ERROR\",IF(AND(Settings!$B$17,ABS(AA{r}-Z{r})>1E-9),\"FIXED\",IF(ABS(S{r})<=Settings!$B$7,\"SAFE\",\"WARNING\")))"
            ws[f"AC{r}"]=f"=IF(AB{r}=\"FIXED\",\"Close adjusted by SAFE mode\",IF(AB{r}=\"SAFE\",\"Balance preserved\",IF(AB{r}=\"WARNING\",\"Near protection boundary\",\"Protection broken\")))"

        # colors
        ws.conditional_formatting.add("AB2:AB41", FormulaRule(formula=['AB2="SAFE"'], fill=PatternFill("solid", fgColor="C6EFCE")))
        ws.conditional_formatting.add("AB2:AB41", FormulaRule(formula=['AB2="FIXED"'], fill=PatternFill("solid", fgColor="9CC2E5")))
        ws.conditional_formatting.add("AB2:AB41", FormulaRule(formula=['AB2="WARNING"'], fill=PatternFill("solid", fgColor="FFEB9C")))
        ws.conditional_formatting.add("AB2:AB41", FormulaRule(formula=['AB2="ERROR"'], fill=PatternFill("solid", fgColor="FFC7CE")))

    fill(wb['DownTrend'], True)
    fill(wb['UpTrend'], False)

    sm = wb['Summary']
    sm['A1']='Parameter'; sm['B1']='Value'; sm['A1'].font=sm['B1'].font=Font(bold=True)
    vals=[('StartLot','=Settings!B2'),('Direction','=Settings!B10'),('MaxLevels','=Settings!B4'),('Final Total BUY %','=IF(Settings!B10="DOWN",INDEX(DownTrend!Q:Q,Settings!B4+3),INDEX(UpTrend!R:R,Settings!B4+3))'),('Final Total SELL %','=IF(Settings!B10="DOWN",INDEX(DownTrend!R:R,Settings!B4+3),INDEX(UpTrend!Q:Q,Settings!B4+3))'),('Final Skew %','=ABS(B5-B4)'),('Final Start Remaining %','=IF(Settings!B10="DOWN",INDEX(DownTrend!N:N,Settings!B4+3),INDEX(UpTrend!N:N,Settings!B4+3))'),('Final Status','=IF(Settings!B10="DOWN",INDEX(DownTrend!T:T,Settings!B4+3),INDEX(UpTrend!T:T,Settings!B4+3))')]
    for r,(k,f) in enumerate(vals,2): sm[f'A{r}']=k; sm[f'B{r}']=f

    charts=[('Raw vs Rounded Lots','DownTrend',22,25),('BUY/SELL after rounding','DownTrend',17,18),('SAFE adjustments count','DownTrend',28,28),('Skew before/after','DownTrend',10,19),('Skew after rounding','DownTrend',19,19)]
    for i,(t,sh,c1,c2) in enumerate(charts):
        ch=LineChart(); ch.title=t
        data=Reference(wb[sh],min_col=c1,min_row=2,max_col=c2,max_row=8)
        ch.add_data(data,titles_from_data=True)
        ch.set_categories(Reference(wb[sh],min_col=1,min_row=3,max_row=8))
        sm.add_chart(ch,f'D{1+i*12}')

    ck=wb['Checks']; ck['A1']='Check'; ck['B1']='Result'; ck['A1'].font=ck['B1'].font=Font(bold=True)
    checks=[
        ('Invalid StartLot','=IF(OR(NOT(ISNUMBER(Settings!B2)),Settings!B2<=0),"ERROR","OK")'),
        ('Invalid LotStep','=IF(OR(NOT(ISNUMBER(Settings!B5)),Settings!B5<=0),"ERROR","OK")'),
        ('Invalid Direction','=IF(AND(Settings!B10<>"DOWN",Settings!B10<>"UP"),"ERROR","OK")'),
        ('Big>=Small Down/Up','=IF(SUMPRODUCT(--(DownTrend!C3:C41<DownTrend!E3:E41))+SUMPRODUCT(--(UpTrend!C3:C41<UpTrend!E3:E41))>0,"ERROR","OK")'),
        ('TotalBUY<=TotalSELL Down','=IF(SUMPRODUCT(--(DownTrend!Q3:Q41>DownTrend!R3:R41))>0,"ERROR","OK")'),
        ('TotalSELL<=TotalBUY Up','=IF(SUMPRODUCT(--(UpTrend!Q3:Q41>UpTrend!R3:R41))>0,"ERROR","OK")'),
        ('SAFE rounding preserved','=IF(OR(COUNTIF(DownTrend!AB3:AB41,"ERROR")>0,COUNTIF(UpTrend!AB3:AB41,"ERROR")>0),"ERROR","OK")'),
        ('Negative values','=IF(OR(MIN(DownTrend!N3:N41)<0,MIN(UpTrend!N3:N41)<0,MIN(DownTrend!Q3:R41)<0,MIN(UpTrend!Q3:R41)<0),"ERROR","OK")'),
    ]
    for r,(a,b) in enumerate(checks,2): ck[f'A{r}']=a; ck[f'B{r}']=b

    wb['Manual']['A1']='v2: Risk-Safe Rounding + Full Input Validation enabled.'
    wb.save(output_path)


if __name__ == '__main__':
    build_workbook('projects/minus-lock-system/MinusLock_Percent_Grid_Calculator.xlsx')
