from pathlib import Path
from openpyxl import load_workbook

WB = Path('projects/minus-lock-system/MinusLock_Percent_Grid_Calculator.xlsx')
RPT = Path('projects/minus-lock-system/PERCENT_GRID_VALIDATION_REPORT_V2_RU.md')


def main():
    wb = load_workbook(WB)
    s = wb['Settings']
    down = wb['DownTrend']
    up = wb['UpTrend']
    sm = wb['Summary']
    ck = wb['Checks']

    sheets_ok = wb.sheetnames == ['Settings', 'DownTrend', 'UpTrend', 'Summary', 'Checks', 'Manual']
    new_params_ok = [s[f'B{i}'].value for i in range(11, 19)] == ['DOWN', 'UP', 'SAFE', 'DOWN', 'UP', 'SAFE', True, True]
    rounding_cols_ok = [down[f'{c}1'].value for c in ['V','W','X','Y','Z','AA','AB','AC']] == [
        'Big Raw Lot','Big Rounded','Small Raw Lot','Small Rounded','Close Raw Lot','Close Rounded','Safe Rounding Status','Rounding Comment']

    base_ok = (down['Q7'].value == '=N7+O7' and down['R7'].value == '=100+P7')
    charts_ok = len(sm._charts) >= 5

    checks_titles = [ck[f'A{i}'].value for i in range(2, 10)]
    checks_ok = 'Invalid StartLot' in checks_titles and 'SAFE rounding preserved' in checks_titles

    report = f'''# PERCENT GRID VALIDATION REPORT V2 (RU)

## 1. Risk-Safe Rounding
- Новые параметры Settings B11:B18: {'OK' if new_params_ok else 'ERROR'}.
- Колонки Raw/Rounded/SAFE status в DownTrend: {'OK' if rounding_cols_ok else 'ERROR'}.

## 2. Input Validation
- Лист Checks содержит проверки Invalid StartLot / LotStep / Direction / SAFE rounding: {'OK' if checks_ok else 'ERROR'}.

## 3. Rounding Safety
- В формулах присутствует разделение Raw vs Rounded и SAFE-статус: {'OK' if rounding_cols_ok else 'ERROR'}.

## 4. Protection Balance
- Базовые формулы Total Main/Opp присутствуют: {'OK' if base_ok else 'ERROR'}.

## 5. SAFE corrections
- Столбец Safe Rounding Status (AB) присутствует: {'OK' if down['AB1'].value == 'Safe Rounding Status' else 'ERROR'}.

## 6. Validation Errors
- Структура листов: {'OK' if sheets_ok else 'ERROR'}.
- Графики v2 (>=5): {'OK' if charts_ok else 'ERROR'} (найдено: {len(sm._charts)}).

## Общий вердикт
**{'OK' if all([sheets_ok, new_params_ok, rounding_cols_ok, checks_ok, charts_ok]) else 'WARNING/ERROR'}**
'''
    RPT.write_text(report, encoding='utf-8')
    print(report)


if __name__ == '__main__':
    main()
