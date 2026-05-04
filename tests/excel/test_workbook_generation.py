from pathlib import Path
import runpy

from openpyxl import load_workbook


def _column_values(ws, col: str):
    return [cell.value for cell in ws[col] if cell.value is not None]


def test_workbook_generation_and_required_fields():
    runpy.run_path('create_adaptive_lock_ev_excel.py', run_name='__main__')

    workbook_path = Path('adaptive_lock_ev_calculator.xlsx')
    assert workbook_path.exists(), 'Workbook file was not created'

    wb = load_workbook(workbook_path, data_only=False)

    assert 'Excel_Audit' in wb.sheetnames
    assert 'Scenario_Up' in wb.sheetnames
    assert 'Scenario_Down' in wb.sheetnames

    excel_audit_labels = _column_values(wb['Excel_Audit'], 'A')
    assert 'Z_Up' in excel_audit_labels
    assert 'Z_Down' in excel_audit_labels

    for sheet_name in ['Scenario_Up', 'Scenario_Down']:
        labels = _column_values(wb[sheet_name], 'A')
        assert 'Projected Exposure' in labels
        assert 'Block Reason' in labels
        assert 'selected_position_id' in labels
        assert 'locked_volume_before' in labels
        assert 'locked_volume_after' in labels
