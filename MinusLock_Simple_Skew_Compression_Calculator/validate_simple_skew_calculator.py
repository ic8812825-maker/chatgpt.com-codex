from __future__ import annotations

import re
import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "MinusLock_Simple_Skew_Compression_Calculator.xlsx"
REQUIRED_SHEETS = ["Калькулятор", "РИСК_АНАЛИЗ", "Тесты", "Руководство", "Описание"]
ERROR_TOKENS = {"#NAME?", "#ИМЯ?", "#VALUE!", "#ЗНАЧ!", "#REF!", "#ССЫЛКА!", "#DIV/0!", "#Н/Д", "#N/A"}
UNQUOTED_SHEET_RE = re.compile(r"=(Калькулятор|РИСК_АНАЛИЗ|Тесты|Руководство|Описание)!", re.IGNORECASE)


def die(message: str) -> None:
    print(f"VALIDATION FAILED: {message}")
    raise SystemExit(1)


def ok(condition: bool, message: str) -> None:
    if not condition:
        die(message)


def recalc_with_libreoffice(src: Path) -> Path:
    soffice = shutil.which("libreoffice") or shutil.which("soffice")
    ok(soffice is not None, "LibreOffice/soffice недоступен")
    tmp_dir = Path(tempfile.mkdtemp(prefix="minuslock_recalc_"))
    tmp_src = tmp_dir / src.name
    tmp_src.write_bytes(src.read_bytes())

    ods_file = tmp_dir / f"{tmp_src.stem}.ods"
    p1 = subprocess.run([soffice, "--headless", "--convert-to", "ods", "--outdir", str(tmp_dir), str(tmp_src)], capture_output=True, text=True)
    ok(p1.returncode == 0 and ods_file.exists(), f"ошибка LibreOffice xlsx->ods: {p1.stderr.strip() or p1.stdout.strip()}")
    p2 = subprocess.run([soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(tmp_dir), str(ods_file)], capture_output=True, text=True)
    ok(p2.returncode == 0, f"ошибка LibreOffice ods->xlsx: {p2.stderr.strip() or p2.stdout.strip()}")

    recalc_file = tmp_dir / f"{ods_file.stem}.xlsx"
    ok(recalc_file.exists(), "пересчитанный xlsx не создан")
    return recalc_file


def scan_workbook_formulas_and_errors(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    for s in REQUIRED_SHEETS:
        ok(s in wb.sheetnames, f"нет листа {s}")

    found_formula_cells = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    if value in ERROR_TOKENS:
                        die(f"ошибка Excel {value} в {ws.title}!{cell.coordinate}")
                    if value.startswith("="):
                        found_formula_cells += 1
                        if UNQUOTED_SHEET_RE.search(value):
                            die(f"некавыченная межлистовая ссылка в {ws.title}!{cell.coordinate}: {value}")
    ok(found_formula_cells > 0, "формулы не найдены")


def check_dynamic_expectations(recalc_path: Path) -> None:
    wb = load_workbook(recalc_path, data_only=True)
    c = wb["Калькулятор"]
    r = wb["РИСК_АНАЛИЗ"]
    t = wb["Тесты"]

    start_lot = float(c["B2"].value)

    exp = {
        "База Main 165": 165,
        "База Opp 175": 175,
        "База Skew 10": 10,
        "Close 60": 60,
        "Close 30": 30,
        "Close 0": 0,
        "Сумма Big": start_lot * 1.55,
        "Сумма Small": start_lot * 0.75,
        "Сумма Close": start_lot * 0.90,
        "Финальный Rounded Main": start_lot * 1.65,
        "Финальный Rounded Opp": start_lot * 1.75,
        "Финальный Rounded Skew": start_lot * 0.10,
        "Максимальная маржа": start_lot * 3740,
        "Максимальная нагрузка %": start_lot * 37.4,
        "Риск-анализ синхронизирован": "YES",
        "Human Summary есть": "ИТОГОВЫЙ ЧЕЛОВЕЧЕСКИЙ РАСЧЁТ ВСЕХ УРОВНЕЙ",
    }

    tests = {}
    for rr in range(2, t.max_row + 1):
        name = t[f"A{rr}"].value
        if not name:
            continue
        fact = t[f"B{rr}"].value
        result = t[f"D{rr}"].value
        tests[name] = (fact, result, rr)

    for key, want in exp.items():
        ok(key in tests, f"нет обязательного теста '{key}'")
        fact, result, rr = tests[key]
        if isinstance(want, (int, float)):
            ok(abs(float(fact) - float(want)) < 1e-6, f"{key}: факт {fact}, ожидание {want}")
        else:
            ok(str(fact) == str(want), f"{key}: факт {fact}, ожидание {want}")
        ok(result == "PASS", f"Тесты!D{rr} для '{key}' не PASS: {result}")

    # human comment existence
    ok("Human Comment есть" in tests, "нет обязательного теста 'Human Comment есть'")
    fact, result, rr = tests["Human Comment есть"]
    ok(str(fact) != "", "Human Comment пустой")
    ok(result == "PASS", f"Тесты!D{rr} для Human Comment не PASS")

    # anti-test: if expected is wrong, strict formula must return FAIL
    anti = Path(tempfile.mkdtemp(prefix="minuslock_anti_")) / "anti.xlsx"
    anti.write_bytes(recalc_path.read_bytes())
    anti_wb = load_workbook(anti, data_only=False)
    anti_t = anti_wb["Тесты"]
    target_row = None
    for rr in range(2, anti_t.max_row + 1):
        if anti_t[f"A{rr}"].value == "Сумма Big":
            target_row = rr
            break
    ok(target_row is not None, "anti-test: нет строки Сумма Big")
    anti_t[f"C{target_row}"] = 999999
    anti_wb.save(anti)
    anti_recalc = recalc_with_libreoffice(anti)
    anti_vals = load_workbook(anti_recalc, data_only=True)["Тесты"]
    ok(anti_vals[f"D{target_row}"].value == "FAIL", "anti-test: PASS появился при неправильном ожидании")

    # risk sync rows
    for lvl in range(1, 6):
        rr = lvl + 2
        cr = lvl + 17
        mapping = [("A", "J"), ("B", "R"), ("C", "Q"), ("D", "N"), ("E", "T"), ("F", "X"), ("G", "AA")]
        for rc, cc in mapping:
            rv = r[f"{rc}{rr}"].value
            cv = c[f"{cc}{cr}"].value
            if isinstance(cv, (int, float)):
                ok(abs(float(rv) - float(cv)) < 1e-6, f"рассинхрон РИСК_АНАЛИЗ {rc}{rr} vs {cc}{cr}")
            else:
                ok(str(rv) == str(cv), f"рассинхрон РИСК_АНАЛИЗ {rc}{rr} vs {cc}{cr}")


def main() -> None:
    ok(WORKBOOK_PATH.exists(), "нет workbook")
    scan_workbook_formulas_and_errors(WORKBOOK_PATH)
    recalc = recalc_with_libreoffice(WORKBOOK_PATH)
    check_dynamic_expectations(recalc)
    print("ALL TESTS PASSED")
    print("DYNAMIC TESTS FIXED")
    print("SELECTED STARTLOT SCENARIO VALIDATED")


if __name__ == "__main__":
    main()
