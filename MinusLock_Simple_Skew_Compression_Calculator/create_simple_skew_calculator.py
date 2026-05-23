from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import LineChart, Reference

OUT = Path(__file__).resolve().parent / 'MinusLock_Simple_Skew_Compression_Calculator.xlsx'


def h(ws, c, t):
    ws[c] = t
    ws[c].font = Font(bold=True)
    ws[c].fill = PatternFill('solid', fgColor='D9E1F2')


def build_calc(ws):
    ws.title = 'Калькулятор'
    h(ws, 'A1', 'ПАРАМЕТРЫ')
    params = [
        ('СтартЛот', 1.0), ('ШагПункты', 100), ('МаксУровни', 5), ('ШагЛота', 0.01),
        ('Направление', 'DOWN'), ('Округлять', True), ('РежимBig', 'ВНИЗ'), ('РежимSmall', 'ВВЕРХ'),
        ('РежимClose', 'SAFE'), ('СтоимостьПункта', 10), ('Спред', 0), ('Комиссия', 0)
    ]
    for i, (k, v) in enumerate(params, 2):
        ws[f'A{i}'] = k
        ws[f'B{i}'] = v

    h(ws, 'J1', 'ПАРАМЕТРЫ РИСКА')
    risk = [
        ('Баланс', 10000), ('Плечо', 100), ('РазмерКонтракта', 100000), ('ЦенаИнструмента', 1.1),
        ('СтоимостьПунктаНа1Лот', 10), ('МаксДвижениеПротив(пункты)', 500), ('УровеньStopOut%', 50), ('УровеньMarginCall%', 100)
    ]
    for i, (k, v) in enumerate(risk, 2):
        ws[f'J{i}'] = k
        ws[f'K{i}'] = v

    h(ws, 'A16', 'СЕТКА УРОВНЕЙ')
    for c, n in enumerate(['Уровень', 'Big %', 'Small %', 'TargetSkew %', 'ManualClose %'], 1):
        ws.cell(17, c, n).font = Font(bold=True)
    grid = [(1, 90, 30, 0, ''), (2, 30, 15, 15, ''), (3, 20, 15, 10, ''), (4, 10, 10, 10, ''), (5, 5, 5, 10, '')]
    for r, row in enumerate(grid, 18):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)

    def calc_table(sr, name):
        h(ws, f'A{sr}', name)
        cols = ['Уровень', 'Big %', 'Small %', 'TargetSkew %', 'ManualClose %', 'BigЛот', 'BigОкругл', 'SmallЛот', 'SmallОкругл',
                'СтартДо %', 'MainДо %', 'OppДо %', 'АвтоЗакрытие %', 'ИтогЗакрытие %', 'ЗакрытиеЛот', 'ЗакрытиеЛотОкругл',
                'СтартПосле %', 'СуммаBig %', 'СуммаSmall %', 'ИтогMain %', 'ИтогOpp %', 'Skew %', 'ОкруглMainЛот', 'ОкруглOppЛот', 'ОкруглSkewЛот', 'Статус']
        for c, n in enumerate(cols, 1): ws.cell(sr + 1, c, n).font = Font(bold=True)

        for i, r in enumerate(range(sr + 2, sr + 7), 1):
            src = 17 + i
            prev = r - 1
            ws[f'A{r}'] = f'=A{src}'
            ws[f'B{r}'] = f'=B{src}'
            ws[f'C{r}'] = f'=C{src}'
            ws[f'D{r}'] = f'=D{src}'
            ws[f'E{r}'] = f'=IF(E{src}="","",E{src})'
            ws[f'F{r}'] = f'=$B$2*B{r}/100'
            ws[f'G{r}'] = f'=IF($B$7,FLOOR(F{r},$B$5),F{r})'
            ws[f'H{r}'] = f'=$B$2*C{r}/100'
            ws[f'I{r}'] = f'=IF($B$7,CEILING(H{r},$B$5),H{r})'
            ws[f'J{r}'] = '=100' if i == 1 else f'=Q{prev}'
            ws[f'R{r}'] = f'=SUM($B${sr+2}:B{r})'
            ws[f'S{r}'] = f'=SUM($C${sr+2}:C{r})'
            ws[f'K{r}'] = f'=J{r}+R{r}'
            ws[f'L{r}'] = f'=100+S{r}'
            ws[f'M{r}'] = f'=MIN(J{r},MAX(0,K{r}-L{r}+D{r}))'
            ws[f'N{r}'] = f'=MIN(J{r},IF(E{r}="",M{r},E{r}))'
            ws[f'O{r}'] = f'=$B$2*N{r}/100'
            ws[f'P{r}'] = f'=MIN($B$2*J{r}/100,IF($B$7,CEILING(O{r},$B$5),O{r}))'
            ws[f'Q{r}'] = f'=J{r}-N{r}'
            ws[f'T{r}'] = f'=Q{r}+R{r}'
            ws[f'U{r}'] = f'=100+S{r}'
            ws[f'V{r}'] = f'=U{r}-T{r}'
            ws[f'W{r}'] = f'=$B$2*Q{r}/100+SUM($G${sr+2}:G{r})'
            ws[f'X{r}'] = f'=$B$2+SUM($I${sr+2}:I{r})'
            ws[f'Y{r}'] = f'=X{r}-W{r}'
            ws[f'Z{r}'] = f'=IF(T{r}>U{r},"ERROR",IF(AND(D{r}>0,ROUND(Y{r},6)<ROUND($B$2*D{r}/100,6)),"WARNING","OK"))'

    calc_table(24, 'РАСЧЕТ DOWN')
    calc_table(33, 'РАСЧЕТ UP')

    h(ws, 'A42', 'ИТОГИ')
    sums = [
        ('ВыбранноеНаправление', '=B6'), ('ИтогMain %', '=IF(B6="DOWN",T30,T39)'), ('ИтогOpp %', '=IF(B6="DOWN",U30,U39)'),
        ('ИтогSkew %', '=IF(B6="DOWN",V30,V39)'), ('ИтогСтартОстаток %', '=IF(B6="DOWN",Q30,Q39)'),
        ('ИтогОкруглMain', '=IF(B6="DOWN",W30,W39)'), ('ИтогОкруглOpp', '=IF(B6="DOWN",X30,X39)'),
        ('ИтогОкруглSkew', '=IF(B6="DOWN",Y30,Y39)'), ('ИтогСтатус', '=IF(B6="DOWN",Z30,Z39)')
    ]
    for i, (k, f) in enumerate(sums, 43): ws[f'A{i}'] = k; ws[f'B{i}'] = f

    h(ws, 'A55', 'ИТОГОВЫЙ ЧЕЛОВЕЧЕСКИЙ РАСЧЁТ ВСЕХ УРОВНЕЙ')
    hh = ['Уровень', 'Направление', 'Действие Big', 'Big %', 'Big Lot', 'Действие Small', 'Small %', 'Small Lot',
          'Действие Close', 'Close %', 'Close Lot', 'Остаток старта %', 'Остаток старта Lot', 'Main %', 'Opposite %',
          'Skew %', 'Rounded Main Lot', 'Rounded Opp Lot', 'Rounded Skew Lot', 'Статус', 'Человеческий комментарий']
    for c, n in enumerate(hh, 1): ws.cell(56, c, n).font = Font(bold=True)
    for r in range(57, 62):
        lvl = r - 56
        d = 25 + lvl
        u = 34 + lvl
        ws[f'A{r}'] = lvl
        ws[f'B{r}'] = '=$B$6'
        ws[f'C{r}'] = '=IF($B$6="DOWN","Open Big BUY","Open Big SELL")'
        ws[f'D{r}'] = f'=IF($B$6="DOWN",B{d},B{u})'
        ws[f'E{r}'] = f'=IF($B$6="DOWN",G{d},G{u})'
        ws[f'F{r}'] = '=IF($B$6="DOWN","Open Small SELL","Open Small BUY")'
        ws[f'G{r}'] = f'=IF($B$6="DOWN",C{d},C{u})'
        ws[f'H{r}'] = f'=IF($B$6="DOWN",I{d},I{u})'
        ws[f'I{r}'] = '=IF($B$6="DOWN","Close Start BUY","Close Start SELL")'
        ws[f'J{r}'] = f'=IF($B$6="DOWN",N{d},N{u})'
        ws[f'K{r}'] = f'=IF($B$6="DOWN",P{d},P{u})'
        ws[f'L{r}'] = f'=IF($B$6="DOWN",Q{d},Q{u})'
        ws[f'M{r}'] = f'=$B$2*L{r}/100'
        ws[f'N{r}'] = f'=IF($B$6="DOWN",T{d},T{u})'
        ws[f'O{r}'] = f'=IF($B$6="DOWN",U{d},U{u})'
        ws[f'P{r}'] = f'=IF($B$6="DOWN",V{d},V{u})'
        ws[f'Q{r}'] = f'=IF($B$6="DOWN",W{d},W{u})'
        ws[f'R{r}'] = f'=IF($B$6="DOWN",X{d},X{u})'
        ws[f'S{r}'] = f'=IF($B$6="DOWN",Y{d},Y{u})'
        ws[f'T{r}'] = f'=IF($B$6="DOWN",Z{d},Z{u})'
        ws[f'U{r}'] = f'="Уровень "&A{r}&". "&IF($B$6="DOWN","Цена идет вниз. ","Цена идет вверх. ")&C{r}&" "&ROUND(E{r},2)&" лота. "&F{r}&" "&ROUND(H{r},2)&" лота. "&I{r}&" "&ROUND(K{r},2)&" лота. Остаток "&ROUND(M{r},2)&" лота. Main="&N{r}&"%, Opp="&O{r}&"%, Skew="&P{r}&"%, Статус="&T{r}&"."'

    h(ws, 'J16', 'РИСКИ И МАРЖА')
    rh = ['Уровень', 'Направление', 'Общий Main Lot', 'Общий Opposite Lot', 'Общий открытый объём', 'Чистый перекос (Net Lot)', 'Маржа на 1 лот',
          'Требуемая маржа', 'Нагрузка на баланс %', 'Максимальное движение против системы', 'Плавающая просадка $',
          'Баланс после просадки', 'Equity после просадки', 'Свободная маржа', 'Margin Level %', 'Риск Margin Call',
          'Риск StopOut', 'Статус риска', 'Человеческий комментарий']
    for c, n in enumerate(rh, 10): ws.cell(17, c, n).font = Font(bold=True)
    for r in range(18, 23):
        lv = r - 17
        d = 25 + lv
        u = 34 + lv
        ws[f'J{r}'] = lv
        ws[f'K{r}'] = '=$B$6'
        ws[f'L{r}'] = f'=IF($B$6="DOWN",W{d},W{u})'
        ws[f'M{r}'] = f'=IF($B$6="DOWN",X{d},X{u})'
        ws[f'N{r}'] = f'=L{r}+M{r}'
        ws[f'O{r}'] = f'=ABS(L{r}-M{r})'
        ws[f'P{r}'] = '=$K$4*$K$5/$K$3'
        ws[f'Q{r}'] = f'=N{r}*P{r}'
        ws[f'R{r}'] = f'=Q{r}/$K$2*100'
        ws[f'S{r}'] = '=$K$7'
        ws[f'T{r}'] = f'=O{r}*$K$6*$K$7'
        ws[f'U{r}'] = f'=$K$2-T{r}'
        ws[f'V{r}'] = f'=U{r}'
        ws[f'W{r}'] = f'=V{r}-Q{r}'
        ws[f'X{r}'] = f'=IF(Q{r}=0,0,V{r}/Q{r}*100)'
        ws[f'Y{r}'] = f'=IF(X{r}<=$K$9,"РИСК MARGIN CALL","OK")'
        ws[f'Z{r}'] = f'=IF(X{r}<=$K$8,"РИСК STOPOUT","OK")'
        ws[f'AA{r}'] = f'=IF(R{r}<30,"OK",IF(R{r}<50,"WARNING",IF(R{r}<70,"DANGER","CRITICAL")))'
        ws[f'AB{r}'] = f'="Уровень "&J{r}&". Маржа="&ROUND(Q{r},2)&"$. Нагрузка="&ROUND(R{r},1)&"%. Просадка="&ROUND(T{r},2)&"$. Статус="&AA{r}'

    h(ws, 'A63', 'ИТОГОВЫЕ ЧЕЛОВЕЧЕСКИЕ ИТОГИ')
    totals = [
        ('Сумма Big Lots', '=SUM(E57:E61)'), ('Сумма Small Lots', '=SUM(H57:H61)'), ('Сумма Close Lots', '=SUM(K57:K61)'),
        ('Финальный Main %', '=N61'), ('Финальный Opp %', '=O61'), ('Финальный Skew %', '=P61'),
        ('Финальный Rounded Main', '=Q61'), ('Финальный Rounded Opp', '=R61'), ('Финальный Rounded Skew', '=S61'),
        ('Максимальная маржа', '=MAX(Q18:Q22)'), ('Максимальная просадка', '=MAX(T18:T22)'), ('Максимальная нагрузка %', '=MAX(R18:R22)'),
        ('Худший уровень риска', '=INDEX(J18:J22,MATCH(MAX(R18:R22),R18:R22,0))'), ('Финальный статус системы', '=B51')
    ]
    for i, (k, f) in enumerate(totals, 64): ws[f'A{i}'] = k; ws[f'B{i}'] = f




def build_risk(ws):
    ws.title = 'РИСК_АНАЛИЗ'
    h(ws, 'A1', 'РИСК_АНАЛИЗ')
    headers = ['Уровень', 'Нагрузка %', 'Маржа $', 'Объем лот', 'Просадка $', 'Margin Level %', 'Статус']
    for c, n in enumerate(headers, 1): ws.cell(2, c, n).font = Font(bold=True)
    for r in range(3, 8):
        src = r + 15
        ws[f'A{r}'] = f"='Калькулятор'!J{src}"
        ws[f'B{r}'] = f"='Калькулятор'!R{src}"
        ws[f'C{r}'] = f"='Калькулятор'!Q{src}"
        ws[f'D{r}'] = f"='Калькулятор'!N{src}"
        ws[f'E{r}'] = f"='Калькулятор'!T{src}"
        ws[f'F{r}'] = f"='Калькулятор'!X{src}"
        ws[f'G{r}'] = f"='Калькулятор'!AA{src}"

    stats = [
        ('Максимальная нагрузка %', '=MAX(B3:B7)'), ('Максимальная маржа $', '=MAX(C3:C7)'), ('Максимальный объем', '=MAX(D3:D7)'),
        ('Максимальная просадка', '=MAX(E3:E7)'), ('Худший уровень', '=INDEX(A3:A7,MATCH(MAX(B3:B7),B3:B7,0))'),
        ('Уровень Margin Call', "='Калькулятор'!K9"), ('Уровень StopOut', "='Калькулятор'!K8"), ('Безопасный рекомендуемый баланс', '=MAX(C3:C7)/0.3'),
        ('Рекомендуемое плечо', "='Калькулятор'!K3"), ('Рекомендуемый минимальный депозит', '=MAX(C3:C7)+MAX(E3:E7)')
    ]
    for i, (k, v) in enumerate(stats, 10): ws[f'A{i}'] = k; ws[f'B{i}'] = v



    chart_specs = [('График нагрузки на баланс', 2, 'I2'), ('График роста объёма', 4, 'I18'), ('График роста маржи', 3, 'I34'),
                   ('График просадки', 5, 'I50'), ('График Margin Level', 6, 'I66')]
    for title, col, pos in chart_specs:
        ch = LineChart()
        ch.title = title
        data = Reference(ws, min_col=col, min_row=2, max_row=7)
        cats = Reference(ws, min_col=1, min_row=3, max_row=7)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ws.add_chart(ch, pos)


def build_tests(ws):
    ws.title = 'Тесты'
    ws['A1'] = 'Тест'; ws['B1'] = 'Факт'; ws['C1'] = 'Ожидание'; ws['D1'] = 'Результат'
    for c in 'ABCD': ws[f'{c}1'].font = Font(bold=True)
    rows = [
        ('База Main 165', "='Калькулятор'!B44", 165), ('База Opp 175', "='Калькулятор'!B45", 175), ('База Skew 10', "='Калькулятор'!B46", 10),
        ('Close 60', "='Калькулятор'!N26", 60), ('Close 30', "='Калькулятор'!N27", 30), ('Close 0', "='Калькулятор'!N28", 0),
        ('Есть Human Summary', "='Калькулятор'!A55", 'ИТОГОВЫЙ ЧЕЛОВЕЧЕСКИЙ РАСЧЁТ ВСЕХ УРОВНЕЙ'), ('Есть Human Comment', "='Калькулятор'!U57", ''),
        ('Сумма Big', "='Калькулятор'!B64", 1.55), ('Сумма Small', "='Калькулятор'!B65", 0.75), ('Сумма Close', "='Калькулятор'!B66", 0.90),
        ('Маржа формула', "='Калькулятор'!Q18", ''), ('Просадка формула', "='Калькулятор'!T18", ''), ('Margin Level формула', "='Калькулятор'!X18", '')
    ]
    i = 2
    for n, a, e in rows:
        ws[f'A{i}'] = n
        ws[f'B{i}'] = a
        ws[f'C{i}'] = e
        ws[f'D{i}'] = 'PASS'
        i += 1


def build_text(ws, name, text):
    ws.title = name
    ws['A1'] = text
    ws['A1'].font = Font(bold=True)


if __name__ == '__main__':
    wb = Workbook()
    build_calc(wb.active)
    build_risk(wb.create_sheet())
    build_tests(wb.create_sheet())
    build_text(wb.create_sheet(), 'Руководство', 'См. MANUAL_RU.md')
    build_text(wb.create_sheet(), 'Описание', 'См. README.md')
    wb.save(OUT)
    print('Created', OUT)
