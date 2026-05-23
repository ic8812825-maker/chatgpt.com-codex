from __future__ import annotations

import re
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
XLSX = BASE_DIR / "MinusLock_Simple_Skew_Compression_Calculator.xlsx"
ERRORS = {"#NAME?", "#ИМЯ?", "#REF!", "#ССЫЛКА!", "#VALUE!", "#ЗНАЧ!", "#DIV/0!", "#ДЕЛ/0!"}
RUS_SHEETS = ["Калькулятор", "РИСК_АНАЛИЗ", "Тесты", "Руководство", "Описание"]
UNQUOTED_RE = re.compile(r"=(Калькулятор|РИСК_АНАЛИЗ|Тесты|Руководство|Описание)!", re.IGNORECASE)


def fail(msg: str) -> None:
    print(f"SMOKE CHECK FAILED: {msg}")
    raise SystemExit(1)


def ok(cond: bool, msg: str) -> None:
    if not cond:
        fail(msg)


def recalc_via_libreoffice(src: Path) -> Path:
    out_dir = Path(tempfile.mkdtemp(prefix="smoke_recalc_"))
    ods = out_dir / "recalc.ods"
    xlsx = out_dir / "recalc.xlsx"

    p1 = subprocess.run(["libreoffice", "--headless", "--convert-to", "ods", "--outdir", str(out_dir), str(src)], capture_output=True, text=True)
    ok(p1.returncode == 0, f"libreoffice xlsx->ods failed: {p1.stderr or p1.stdout}")

    generated_ods = out_dir / f"{src.stem}.ods"
    ok(generated_ods.exists(), "ods not generated")
    generated_ods.rename(ods)

    p2 = subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", str(out_dir), str(ods)], capture_output=True, text=True)
    ok(p2.returncode == 0, f"libreoffice ods->xlsx failed: {p2.stderr or p2.stdout}")
    ok(xlsx.exists(), "xlsx not generated after recalc")
    return xlsx


def scan_formulas(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    for sheet in RUS_SHEETS:
        ok(sheet in wb.sheetnames, f"missing sheet {sheet}")

    found = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    if v in ERRORS:
                        fail(f"error token {v} in {ws.title}!{cell.coordinate}")
                    if v.startswith("="):
                        found += 1
                        if UNQUOTED_RE.search(v):
                            fail(f"unquoted russian sheet ref {ws.title}!{cell.coordinate}: {v}")
    ok(found > 0, "no formulas found")


def check_values(recalc_path: Path) -> None:
    wb = load_workbook(recalc_path, data_only=True)
    c = wb["Калькулятор"]
    r = wb["РИСК_АНАЛИЗ"]

    expected = {"B70": 1.65, "B71": 1.75, "B72": 0.10, "B73": 3740, "B75": 37.4}
    for cell, val in expected.items():
        cur = c[cell].value
        ok(cur is not None and cur != "", f"empty summary {cell}")
        ok(abs(float(cur) - float(val)) < 1e-6, f"bad summary {cell}: {cur} != {val}")

    ok(isinstance(c["B70"].value, (int, float)), "Rounded Main not numeric")
    ok(isinstance(c["B71"].value, (int, float)), "Rounded Opp not numeric")
    ok(isinstance(c["B72"].value, (int, float)), "Rounded Skew not numeric")

    for lvl in range(1, 6):
        rr = lvl + 2
        cr = lvl + 17
        mapping = [("A", "J"), ("B", "R"), ("C", "Q"), ("D", "N"), ("E", "T"), ("F", "X"), ("G", "AA")]
        for rc, cc in mapping:
            rv = r[f"{rc}{rr}"].value
            cv = c[f"{cc}{cr}"].value
            ok(rv is not None and cv is not None, f"empty sync at level {lvl} {rc}/{cc}")
            if isinstance(cv, (int, float)):
                ok(abs(float(rv) - float(cv)) < 1e-6, f"desync level {lvl}: RISK {rc}{rr}={rv} vs CALC {cc}{cr}={cv}")
            else:
                ok(str(rv) == str(cv), f"desync level {lvl}: RISK {rc}{rr}={rv} vs CALC {cc}{cr}={cv}")


if __name__ == "__main__":
    ok(XLSX.exists(), f"missing workbook {XLSX.name}")
    scan_formulas(XLSX)
    recalc = recalc_via_libreoffice(XLSX)
    check_values(recalc)
    print("SMOKE CHECK PASSED")
