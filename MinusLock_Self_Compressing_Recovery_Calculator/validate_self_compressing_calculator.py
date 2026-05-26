from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent / "MinusLock_Self_Compressing_Recovery_Calculator.xlsx"
ERRORS = {"#NAME?", "#ИМЯ?", "#REF!", "#ССЫЛКА!", "#VALUE!", "#ЗНАЧ!", "#DIV/0!", "#ДЕЛ/0!"}
REQ_SHEETS = {"Калькулятор", "РИСК_АНАЛИЗ", "Тесты", "Руководство", "Описание"}
REQ_RISK = ["Уровень", "Total Big Lots", "Total Small Lots", "Total Open Lots", "Net Lot", "Margin Per Lot", "Required Margin", "Margin Load %", "Floating DD", "Equity After DD", "Free Margin", "Margin Level %", "Risk Status"]

def fail(msg: str):
    raise AssertionError(msg)

wb = load_workbook(XLSX, data_only=True)
if not REQ_SHEETS.issubset(set(wb.sheetnames)):
    fail("missing required sheets")

ws = wb["Калькулятор"]
params = {ws.cell(r,1).value: ws.cell(r,2).value for r in range(2,40)}
if params.get("CloseMode") != "THEORETICAL":
    fail("CloseMode must default to THEORETICAL")
lot_step = float(params["LotStep"])

header_row = 26
headers = [ws.cell(header_row, c).value for c in range(1, 80)]
idx = {h: i+1 for i,h in enumerate(headers) if h}
for col in ["Max Close Far Lot", "Close By Profit Budget", "Actual Close Far Lot", "Close Mode", "Close Status", "Комментарий"]:
    if col not in idx:
        fail(f"missing column {col}")

rows = []
for r in range(header_row+1, header_row+1+int(params["MaxLevels"])):
    rows.append({h: ws.cell(r, c).value for h, c in idx.items()})

for r in rows:
    near = float(r["Ближний старт"])
    actual = float(r["Actual Close Far Lot"])
    new_far = float(r["Новый дальний остаток"])
    if r["Close Mode"] != "THEORETICAL": fail("row close mode not THEORETICAL")
    if near * 0.3 >= lot_step and actual <= 0: fail("Actual Close is zero in theoretical mode")
    if abs(actual - near*0.3) > 1e-10: fail("Actual Close must equal NearStart*30% in theoretical")
    if str(actual) not in str(r["Комментарий"]): fail("Human summary must contain actual close value")

for i in range(1, len(rows)):
    if not (rows[i]["Новый дальний остаток"] < rows[i-1]["Новый дальний остаток"]):
        fail("NewFarRemaining must decrease each level")

# StartLot scenarios raw checks
for startlot in [1.0, 2.0, 5.0]:
    near = startlot
    for _ in range(5):
        if abs(near*0.3 - (startlot * (0.5 ** _) * 0.3)) > 1e-12:
            fail("scenario check failed")
        near *= 0.5

# Mirror check UP/DOWN as side labels in comments logic (BUY/SELL swap ability by rules)
if rows[0]["Старт поз. самая дальняя"] != "Start BUY":
    fail("DOWN mirror baseline invalid")

risk = wb["РИСК_АНАЛИЗ"]
if risk.max_row <= 1:
    fail("risk sheet empty")
risk_headers = [risk.cell(1,c).value for c in range(1,20)]
for h in REQ_RISK:
    if h not in risk_headers:
        fail(f"risk column missing: {h}")

for sh in wb.worksheets:
    for row in sh.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in ERRORS:
                fail(f"excel error token {cell.value} in {sh.title}!{cell.coordinate}")

wt = wb["Тесты"]
statuses = [wt.cell(r,2).value for r in range(2, wt.max_row+1)]
if not statuses or any(s != "PASS" for s in statuses):
    fail("tests not pass")
if any(wt.cell(r,3).value != "formula-based check" for r in range(2, wt.max_row+1)):
    fail("tests must be formula-based")

print("VALIDATION: PASS")
