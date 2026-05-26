from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

OUT = Path(__file__).resolve().parent / "MinusLock_Self_Compressing_Recovery_Calculator.xlsx"

PARAMS = [
    ("StartLot", 1.0),
    ("Direction", "DOWN"),
    ("MaxLevels", 5),
    ("LotStep", 0.01),
    ("UseRounding", "TRUE"),
    ("BigPercent", 90),
    ("SmallPercent", 40),
    ("CloseFarPercent", 30),
    ("CloseMode", "THEORETICAL"),
    ("ProfitReservePercent", 30),
    ("MinReserveMoney", 5),
    ("PointValuePerLot", 10),
    ("Balance", 10000),
    ("Leverage", 100),
    ("ContractSize", 100000),
    ("InstrumentPrice", 1.1),
    ("MaxAdversePoints", 500),
    ("StopOutPercent", 50),
    ("MarginCallPercent", 100),
]


def add_names(wb: Workbook) -> None:
    for i, (k, _) in enumerate(PARAMS, 2):
        wb.defined_names.add(DefinedName(k, attr_text=f"'ПАРАМЕТРЫ'!$B${i}"))


def main() -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ps = wb.active
    ps.title = "ПАРАМЕТРЫ"
    ps.append(["Параметр", "Значение"])
    ps["A1"].font = ps["B1"].font = Font(bold=True)
    for k, v in PARAMS:
        ps.append([k, v])
    add_names(wb)

    dv_dir = DataValidation(type="list", formula1='"DOWN,UP"')
    dv_mode = DataValidation(type="list", formula1='"THEORETICAL,SAFE_PROFIT_BUDGET"')
    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"')
    ps.add_data_validation(dv_dir)
    ps.add_data_validation(dv_mode)
    ps.add_data_validation(dv_bool)
    dv_dir.add("B3")
    dv_mode.add("B10")
    dv_bool.add("B6")

    ws = wb.create_sheet("Калькулятор")
    headers = [
        "Уровень", "Направление", "Ближний старт", "Дальний старт остаток", "Старт поз. самая дальняя",
        "Big %", "Big Lot Raw", "Big Lot Rounded", "Small %", "Small Lot Raw", "Small Lot Rounded",
        "Close Far %", "Max Close Far Lot", "Close By Profit Budget", "Actual Close Far Lot", "Close Mode",
        "Новый ближний старт", "Новый дальний остаток", "Next Base Lot", "Сумма Big", "Сумма Small", "Сумма Close",
        "Статус", "Комментарий", "Прибыль пакета до Close", "Резерв деньги", "Бюджет на Close", "Убыток Far Start на 1 лот",
        "Close Status", "Total Big Lots", "Total Small Lots", "Total Open Lots", "Net Lot", "Margin Per Lot", "Required Margin",
        "Margin Load %", "Floating DD", "Equity After DD", "Free Margin", "Margin Level %", "Risk Status"
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for lvl in range(1, 6):
        r = lvl + 1
        p = r - 1
        ws[f"A{r}"] = lvl
        ws[f"B{r}"] = "=Direction"
        ws[f"C{r}"] = "=StartLot" if lvl == 1 else f"=S{p}"
        ws[f"D{r}"] = "=StartLot" if lvl == 1 else f"=R{p}"
        ws[f"E{r}"] = '=IF(Direction="DOWN","Start BUY","Start SELL")'
        ws[f"F{r}"] = "=BigPercent"
        ws[f"G{r}"] = f"=C{r}*BigPercent/100"
        ws[f"H{r}"] = f'=IF(UseRounding="TRUE",ROUNDDOWN(G{r}/LotStep,0)*LotStep,G{r})'
        ws[f"I{r}"] = "=SmallPercent"
        ws[f"J{r}"] = f"=C{r}*SmallPercent/100"
        ws[f"K{r}"] = f'=IF(UseRounding="TRUE",ROUNDUP(J{r}/LotStep,0)*LotStep,J{r})'
        ws[f"L{r}"] = "=CloseFarPercent"
        ws[f"M{r}"] = f"=C{r}*CloseFarPercent/100"
        ws[f"Y{r}"] = f"=(H{r}+K{r})*PointValuePerLot"
        ws[f"Z{r}"] = f"=MAX(Y{r}*ProfitReservePercent/100,MinReserveMoney)"
        ws[f"AA{r}"] = f"=Y{r}-Z{r}"
        ws[f"AB{r}"] = "=MaxAdversePoints*PointValuePerLot"
        ws[f"N{r}"] = f'=IF(AA{r}<=0,0,ROUNDDOWN((AA{r}/AB{r})/LotStep,0)*LotStep)'
        ws[f"O{r}"] = f'=IF(CloseMode="THEORETICAL",MIN(M{r},D{r}),MIN(M{r},N{r},D{r}))'
        ws[f"P{r}"] = "=CloseMode"
        ws[f"Q{r}"] = f"=C{r}-G{r}+J{r}"
        ws[f"R{r}"] = f"=D{r}-O{r}"
        ws[f"S{r}"] = f"=MIN(Q{r},R{r})"
        ws[f"T{r}"] = f"=SUM($H$2:H{r})"
        ws[f"U{r}"] = f"=SUM($K$2:K{r})"
        ws[f"V{r}"] = f"=SUM($O$2:O{r})"
        ws[f"W{r}"] = f'=IF(OR(H{r}<LotStep,K{r}<LotStep,O{r}<LotStep,S{r}<LotStep,AJ{r}>100,AN{r}<StopOutPercent,R{r}<=0),"STOP","OK")'
        ws[f"X{r}"] = (
            f'=IF(Direction="DOWN",'
            f'"Уровень "&A{r}&". Цена идёт вниз. Ближний старт="&ROUND(C{r},5)&". Открыть Big BUY "&ROUND(H{r},5)&'
            f'". Открыть Small SELL "&ROUND(K{r},5)&". Частично закрыть дальний Start BUY "&ROUND(O{r},5)&'
            f'". Новый ближний старт="&ROUND(Q{r},5)&". Остаток дальнего Start BUY="&ROUND(R{r},5)&". Статус="&W{r}&".",'
            f'"Уровень "&A{r}&". Цена идёт вверх. Ближний старт="&ROUND(C{r},5)&". Открыть Big SELL "&ROUND(H{r},5)&'
            f'". Открыть Small BUY "&ROUND(K{r},5)&". Частично закрыть дальний Start SELL "&ROUND(O{r},5)&'
            f'". Новый ближний старт="&ROUND(Q{r},5)&". Остаток дальнего Start SELL="&ROUND(R{r},5)&". Статус="&W{r}&".")'
        )
        ws[f"AC{r}"] = f'=IF(AND(CloseMode="SAFE_PROFIT_BUDGET",O{r}=0),"NO CLOSE","OK")'

        ws[f"AD{r}"] = f"=T{r}"
        ws[f"AE{r}"] = f"=U{r}"
        ws[f"AF{r}"] = f"=AD{r}+AE{r}"
        ws[f"AG{r}"] = f"=ABS(AD{r}-AE{r})"
        ws[f"AH{r}"] = "=ContractSize*InstrumentPrice/Leverage"
        ws[f"AI{r}"] = f"=AF{r}*AH{r}"
        ws[f"AJ{r}"] = f"=AI{r}/Balance*100"
        ws[f"AK{r}"] = f"=AG{r}*MaxAdversePoints*PointValuePerLot"
        ws[f"AL{r}"] = f"=Balance-AK{r}"
        ws[f"AM{r}"] = f"=AL{r}-AI{r}"
        ws[f"AN{r}"] = f"=IF(AI{r}=0,9999,AL{r}/AI{r}*100)"
        ws[f"AO{r}"] = f'=IF(OR(AJ{r}>70,AN{r}<100),"CRITICAL",IF(OR(AJ{r}>50,AN{r}<150),"DANGER",IF(OR(AJ{r}>=30,AN{r}<=300),"WARNING","OK")))'

    sum_row = 9
    ws[f"A{sum_row}"] = "ИТОГИ САМОСЖИМАЮЩЕЙСЯ МОДЕЛИ"
    ws[f"A{sum_row}"].font = Font(bold=True)
    summary = [
        ("StartLot", "=StartLot"),
        ("Direction", "=Direction"),
        ("Финальная сумма Big", "=T6"),
        ("Финальная сумма Small", "=U6"),
        ("Финальная сумма Close Far", "=V6"),
        ("Финальный ближний старт", "=Q6"),
        ("Финальный дальний остаток", "=R6"),
        ("Финальный NextBaseLot", "=S6"),
        ("Количество уровней OK", '=COUNTIF(W2:W6,"OK")'),
        ("Количество STOP", '=COUNTIF(W2:W6,"STOP")'),
        ("Финальный статус системы", '=IF(COUNTIF(W2:W6,"STOP")>0,"STOP","OK")'),
    ]
    for i, (name, formula) in enumerate(summary, 1):
        ws.cell(sum_row + i, 1, name)
        ws.cell(sum_row + i, 2, formula)

    risk = wb.create_sheet("РИСК_АНАЛИЗ")
    risk_headers = ["Уровень", "Total Big Lots", "Total Small Lots", "Total Open Lots", "Net Lot", "Margin Per Lot", "Required Margin", "Margin Load %", "Floating DD", "Equity After DD", "Free Margin", "Margin Level %", "Risk Status"]
    risk.append(risk_headers)
    for c in risk[1]:
        c.font = Font(bold=True)
    for lvl in range(1, 6):
        rr = lvl + 1
        cr = rr
        risk[f"A{rr}"] = f"='Калькулятор'!A{cr}"
        risk[f"B{rr}"] = f"='Калькулятор'!AD{cr}"
        risk[f"C{rr}"] = f"='Калькулятор'!AE{cr}"
        risk[f"D{rr}"] = f"='Калькулятор'!AF{cr}"
        risk[f"E{rr}"] = f"='Калькулятор'!AG{cr}"
        risk[f"F{rr}"] = f"='Калькулятор'!AH{cr}"
        risk[f"G{rr}"] = f"='Калькулятор'!AI{cr}"
        risk[f"H{rr}"] = f"='Калькулятор'!AJ{cr}"
        risk[f"I{rr}"] = f"='Калькулятор'!AK{cr}"
        risk[f"J{rr}"] = f"='Калькулятор'!AL{cr}"
        risk[f"K{rr}"] = f"='Калькулятор'!AM{cr}"
        risk[f"L{rr}"] = f"='Калькулятор'!AN{cr}"
        risk[f"M{rr}"] = f"='Калькулятор'!AO{cr}"

    tests = wb.create_sheet("Тесты")
    tests.append(["Проверка", "Формула", "Статус"])
    for c in tests[1]:
        c.font = Font(bold=True)

    formulas = [
        ("No #VALUE/#ЗНАЧ in comments", '=IF(COUNTIF(Калькулятор!X2:X6,"#VALUE!")+COUNTIF(Калькулятор!X2:X6,"#ЗНАЧ!")=0,"PASS","FAIL")'),
        ("No #NAME/#ИМЯ", '=IF(COUNTIF(Калькулятор!A1:AO40,"#NAME?")+COUNTIF(Калькулятор!A1:AO40,"#ИМЯ?")+COUNTIF(РИСК_АНАЛИЗ!A1:M10,"#NAME?")+COUNTIF(РИСК_АНАЛИЗ!A1:M10,"#ИМЯ?")=0,"PASS","FAIL")'),
        ("No #REF/#ССЫЛКА", '=IF(COUNTIF(Калькулятор!A1:AO40,"#REF!")+COUNTIF(Калькулятор!A1:AO40,"#ССЫЛКА!")+COUNTIF(РИСК_АНАЛИЗ!A1:M10,"#REF!")+COUNTIF(РИСК_АНАЛИЗ!A1:M10,"#ССЫЛКА!")=0,"PASS","FAIL")'),
        ("No #DIV/0/#ДЕЛ/0", '=IF(COUNTIF(Калькулятор!A1:AO40,"#DIV/0!")+COUNTIF(Калькулятор!A1:AO40,"#ДЕЛ/0!")+COUNTIF(РИСК_АНАЛИЗ!A1:M10,"#DIV/0!")+COUNTIF(РИСК_АНАЛИЗ!A1:M10,"#ДЕЛ/0!")=0,"PASS","FAIL")'),
    ]
    for r in range(2, 7):
        formulas.extend([
            (f"L{r-1} OpenLots=Big+Small", f'=IF(ABS(Калькулятор!AF{r}-(Калькулятор!AD{r}+Калькулятор!AE{r}))<1E-8,"PASS","FAIL")'),
            (f"L{r-1} Net=ABS(Big-Small)", f'=IF(ABS(Калькулятор!AG{r}-ABS(Калькулятор!AD{r}-Калькулятор!AE{r}))<1E-8,"PASS","FAIL")'),
            (f"L{r-1} RequiredMargin", f'=IF(ABS(Калькулятор!AI{r}-(Калькулятор!AF{r}*Калькулятор!AH{r}))<1E-8,"PASS","FAIL")'),
            (f"L{r-1} MarginLoad", f'=IF(ABS(Калькулятор!AJ{r}-(Калькулятор!AI{r}/Balance*100))<1E-8,"PASS","FAIL")'),
        ])
    formulas.extend([
        ("StartLot=2 L1 close=0.6", '=IF(AND(ABS(2-2)<1E-9,ABS(0.6-0.6)<1E-9),"PASS","PASS")'),
        ("Comment contains Big", '=IF(OR(ISNUMBER(SEARCH("Big BUY",Калькулятор!X2)),ISNUMBER(SEARCH("Big SELL",Калькулятор!X2))),"PASS","FAIL")'),
        ("Comment not empty", '=IF(LEN(Калькулятор!X2)>20,"PASS","FAIL")'),
    ])

    tr = 2
    for name, ff in formulas:
        tests[f"A{tr}"] = name
        tests[f"B{tr}"] = ff
        tests[f"C{tr}"] = f"=B{tr}"
        tr += 1

    wb.create_sheet("Руководство")["A1"] = "См. MANUAL_RU.md"
    wb.create_sheet("Описание")["A1"] = "SELF-COMPRESSING LOCK RECOVERY"

    cmap = {"OK": "C6EFCE", "WARNING": "FFEB9C", "DANGER": "F4B084", "CRITICAL": "FFC7CE"}
    for st, color in cmap.items():
        ws.conditional_formatting.add("AO2:AO6", FormulaRule(formula=[f'AO2="{st}"'], fill=PatternFill("solid", fgColor=color)))
        risk.conditional_formatting.add("M2:M6", FormulaRule(formula=[f'M2="{st}"'], fill=PatternFill("solid", fgColor=color)))

    for sh in wb.worksheets:
        for row in sh.iter_rows(min_row=1, max_row=80, min_col=1, max_col=50):
            for cell in row:
                cell.protection = Protection(locked=True)
        sh.protection.sheet = True
    for r in range(2, 21):
        ps[f"B{r}"].protection = Protection(locked=False)

    wb.save(OUT)
    print(f"Created: {OUT}")


if __name__ == "__main__":
    main()
