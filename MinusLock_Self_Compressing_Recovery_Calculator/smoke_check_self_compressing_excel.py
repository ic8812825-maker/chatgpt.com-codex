from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent / "MinusLock_Self_Compressing_Recovery_Calculator.xlsx"
ERRORS = {"#NAME?", "#ИМЯ?", "#REF!", "#ССЫЛКА!", "#VALUE!", "#ЗНАЧ!", "#DIV/0!", "#ДЕЛ/0!"}

wb = load_workbook(XLSX, data_only=True)

risk = wb["РИСК_АНАЛИЗ"]
if risk.max_row <= 1:
    raise SystemExit("РИСК_АНАЛИЗ is empty")

calc = wb["Калькулятор"]
headers = [calc.cell(26,c).value for c in range(1,80)]
idx = {h:i+1 for i,h in enumerate(headers) if h}
for key in ["Уровень", "Ближний старт", "Actual Close Far Lot", "Close Mode"]:
    if key not in idx:
        raise SystemExit(f"missing key column: {key}")
for r in range(27, 32):
    near = calc.cell(r, idx["Ближний старт"]).value
    actual = calc.cell(r, idx["Actual Close Far Lot"]).value
    mode = calc.cell(r, idx["Close Mode"]).value
    if not isinstance(near, (int,float)) or not isinstance(actual, (int,float)):
        raise SystemExit("text instead of numbers")
    if mode == "THEORETICAL" and near * 0.3 >= 0.01 and actual == 0:
        raise SystemExit("zero actual close in theoretical mode")

for sh in wb.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if isinstance(c.value, str) and c.value in ERRORS:
                raise SystemExit(f"excel token {c.value} at {sh.title}!{c.coordinate}")

wt = wb["Тесты"]
for r in range(2, wt.max_row+1):
    if wt.cell(r,2).value == "PASS" and wt.cell(r,3).value != "formula-based check":
        raise SystemExit("fake PASS detected")

print("SMOKE: PASS")
