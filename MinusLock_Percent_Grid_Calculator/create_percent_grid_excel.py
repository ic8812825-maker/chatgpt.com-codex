from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName


BASE_SHEETS = ["Settings", "DownTrend", "UpTrend", "Summary", "Checks", "Manual"]
V3_SHEETS = [
    "MarketModel", "AdaptiveEngine", "MarginControl", "MonteCarlo",
    "EquityModel", "RecoveryMap", "StressTest", "RiskDashboard",
]


def add_settings(ws):
    params = [
        ("StartLot", 1.00), ("PointStep", 100), ("MaxLevels", 5), ("LotStep", 0.01),
        ("RoundMode", "Nearest"), ("TargetSkewMin%", 5), ("TargetSkewMax%", 25),
        ("UseRounding", True), ("Direction", "DOWN"),
        ("BigRoundMode_DOWN", "DOWN"), ("SmallRoundMode_DOWN", "UP"), ("CloseRoundMode_DOWN", "SAFE"),
        ("BigRoundMode_UP", "DOWN"), ("SmallRoundMode_UP", "UP"), ("CloseRoundMode_UP", "SAFE"),
        ("EnableRiskSafeRounding", True), ("EnableInputValidation", True),
    ]
    for r, (k, v) in enumerate(params, 2):
        ws[f"A{r}"] = k
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = v

    headers = ["Level", "Big%", "Small%", "TargetSkew%", "ManualClose%"]
    for i, h in enumerate(headers, 1):
        ws.cell(21, i, h).font = Font(bold=True)
    grid = [(1, 90, 30, 0, None), (2, 30, 15, 15, None), (3, 20, 15, 10, None), (4, 10, 10, 10, None), (5, 5, 5, 10, None)]
    for r, row in enumerate(grid, 22):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)


def add_named_ranges(wb):
    ranges = {
        "LevelRange": "$A$22:$A$200",
        "BigPercentRange": "$B$22:$B$200",
        "SmallPercentRange": "$C$22:$C$200",
        "TargetSkewRange": "$D$22:$D$200",
        "ManualCloseRange": "$E$22:$E$200",
    }
    for name, ref in ranges.items():
        wb.defined_names.add(DefinedName(name, attr_text=f"Settings!{ref}"))


def add_trend_sheet(ws, down=True):
    headers = [
        "Level", "PriceStep", "Big %", "Big Lot", "Small %", "Small Lot", "Start Before %", "Before Close %", "Opposite After Add %",
        "Target Skew %", "Auto Close %", "Manual Close %", "Final Close %", "Start After %", "Sum Big %", "Sum Small %",
        "Total Main %", "Total Opposite %", "Skew %", "Status", "Comment",
        "Big Raw Lot", "Big Rounded", "Small Raw Lot", "Small Rounded", "Close Raw Lot", "Close Rounded", "Safe Rounding Status", "Rounding Comment",
        "Rounded Start After Lot", "Rounded Sum Big Lot", "Rounded Sum Small Lot", "Rounded Total Main Lot", "Rounded Total Opposite Lot", "Rounded Skew Lot", "Rounded Status",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h).font = Font(bold=True)

    for r in range(2, 42):
        ws[f"A{r}"] = 0 if r == 2 else r - 2
        ws[f"B{r}"] = 0 if r == 2 else f"=A{r}*Settings!$B$3"
        ws[f"C{r}"] = 0 if r == 2 else f"=IF(A{r}<=Settings!$B$4,IFERROR(XLOOKUP(A{r},LevelRange,BigPercentRange),0),0)"
        ws[f"E{r}"] = 0 if r == 2 else f"=IF(A{r}<=Settings!$B$4,IFERROR(XLOOKUP(A{r},LevelRange,SmallPercentRange),0),0)"

        ws[f"V{r}"] = 0 if r == 2 else f"=Settings!$B$2*C{r}/100"
        ws[f"X{r}"] = 0 if r == 2 else f"=Settings!$B$2*E{r}/100"
        ws[f"W{r}"] = 0 if r == 2 else f"=IF(Settings!$B$9,FLOOR(V{r},Settings!$B$5),V{r})"
        ws[f"Y{r}"] = 0 if r == 2 else f"=IF(Settings!$B$9,CEILING(X{r},Settings!$B$5),X{r})"
        ws[f"D{r}"] = f"=W{r}"
        ws[f"F{r}"] = f"=Y{r}"

        ws[f"G{r}"] = 100 if r in (2, 3) else f"=N{r-1}"
        ws[f"H{r}"] = 100 if r == 2 else f"=G{r}+SUM($C$3:C{r})"
        ws[f"I{r}"] = 100 if r == 2 else f"=100+SUM($E$3:E{r})"
        ws[f"J{r}"] = 0 if r == 2 else f"=IF(A{r}<=Settings!$B$4,IFERROR(XLOOKUP(A{r},LevelRange,TargetSkewRange),0),0)"
        ws[f"K{r}"] = 0 if r == 2 else f"=MIN(G{r},MAX(0,H{r}-I{r}+J{r}))"
        ws[f"L{r}"] = "" if r == 2 else f"=IF(A{r}<=Settings!$B$4,IFERROR(XLOOKUP(A{r},LevelRange,ManualCloseRange),\"\"),\"\")"
        ws[f"M{r}"] = 0 if r == 2 else f"=MIN(G{r},IF(LEN(L{r})=0,K{r},L{r}))"
        ws[f"Z{r}"] = 0 if r == 2 else f"=Settings!$B$2*M{r}/100"

        # SAFE close in lot space to preserve protection balance
        ws[f"AA{r}"] = 0 if r == 2 else (
            f"=IF(Settings!$B$17,"
            f"MIN(Settings!$B$2*G{r}/100,IF(Settings!$B$9,CEILING(MAX(0,Settings!$B$2*(H{r}-I{r}+J{r})/100),Settings!$B$5),MAX(0,Settings!$B$2*(H{r}-I{r}+J{r})/100))),"
            f"IF(Settings!$B$9,MROUND(Z{r},Settings!$B$5),Z{r}))"
        )
        ws[f"M{r}"] = 0 if r == 2 else f"=MIN(G{r},AA{r}/Settings!$B$2*100)"
        ws[f"N{r}"] = 100 if r == 2 else f"=G{r}-M{r}"
        ws[f"O{r}"] = 0 if r == 2 else f"=SUM($C$3:C{r})"
        ws[f"P{r}"] = 0 if r == 2 else f"=SUM($E$3:E{r})"
        ws[f"Q{r}"] = f"=N{r}+O{r}"
        ws[f"R{r}"] = f"=100+P{r}"
        ws[f"S{r}"] = f"=R{r}-Q{r}"

        big_msg = "Big BUY must be >= Small SELL" if down else "Big SELL must be >= Small BUY"
        bal_expr = f"Q{r}>R{r}"  # mirrored table is already aligned as main/opposite

        ws[f"T{r}"] = (
            f"=IF(Settings!$B$18=FALSE,\"OK\","
            f"IF(OR(NOT(ISNUMBER(Settings!$B$2)),Settings!$B$2<=0),\"ERROR: Invalid StartLot\","
            f"IF(OR(NOT(ISNUMBER(Settings!$B$5)),Settings!$B$5<=0),\"ERROR: Invalid LotStep\","
            f"IF(OR(Settings!$B$4<1,Settings!$B$4>20),\"ERROR: Invalid MaxLevels\","
            f"IF(AND(Settings!$B$10<>\"DOWN\",Settings!$B$10<>\"UP\"),\"ERROR: Invalid Direction\","
            f"IF(OR(C{r}<0,E{r}<0,J{r}<0,AND(LEN(L{r})>0,L{r}<0)),\"ERROR: Negative input\","
            f"IF(C{r}<E{r},\"ERROR: {big_msg}\","
            f"IF(AND(LEN(L{r})>0,L{r}>G{r}),\"ERROR: ManualClose exceeds remaining Start position\","
            f"IF(OR(N{r}<0,Q{r}<0,R{r}<0),\"ERROR: Negative totals\","
            f"IF({bal_expr},\"ERROR: Rounding broke protection balance\","
            f"IF(S{r}>Settings!$B$8,\"WARNING: Skew exceeds recommended limit\",\"OK\")))))))))))"
        )
        ws[f"U{r}"] = f"=IF(LEFT(T{r},5)=\"ERROR\",\"Check input/risk\",\"\")"
        ws[f"AB{r}"] = (
            f"=IF(LEFT(T{r},5)=\"ERROR\",\"ERROR\","
            f"IF(AND(Settings!$B$17,ABS(AA{r}-Z{r})>1E-9),\"FIXED\","
            f"IF(ABS(S{r})<=Settings!$B$7,\"SAFE\",\"WARNING\")))"
        )
        ws[f"AC{r}"] = (
            f"=IF(AB{r}=\"FIXED\",\"Close adjusted by SAFE mode\","
            f"IF(AB{r}=\"SAFE\",\"BUY/SELL balance preserved\","
            f"IF(AB{r}=\"WARNING\",\"Near protection boundary\",\"Protection broken\")))"
        )
        ws[f"AD{r}"] = 0 if r == 2 else f"=MAX(0,Settings!$B$2*G{r}/100-AA{r})"
        ws[f"AE{r}"] = 0 if r == 2 else f"=SUM($W$3:W{r})"
        ws[f"AF{r}"] = 0 if r == 2 else f"=SUM($Y$3:Y{r})"
        ws[f"AG{r}"] = 0 if r == 2 else f"=AD{r}+AE{r}"
        ws[f"AH{r}"] = 0 if r == 2 else f"=Settings!$B$2+AF{r}"
        ws[f"AI{r}"] = 0 if r == 2 else f"=AH{r}-AG{r}"
        ws[f"AJ{r}"] = 0 if r == 2 else f"=IF(OR(AND(C{r}>0,W{r}=0),AND(E{r}>0,Y{r}=0),Settings!$B$2<Settings!$B$5),\"ERROR: LotStep too coarse\",IF(AG{r}>AH{r},\"ERROR: Rounded balance broken\",IF(AI{r}<Settings!$B$2*Settings!$B$7/100,\"WARNING\",\"OK\")))"

    ws.conditional_formatting.add("AB2:AB41", FormulaRule(formula=['AB2="SAFE"'], fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add("AB2:AB41", FormulaRule(formula=['AB2="FIXED"'], fill=PatternFill("solid", fgColor="9CC2E5")))
    ws.conditional_formatting.add("AB2:AB41", FormulaRule(formula=['AB2="WARNING"'], fill=PatternFill("solid", fgColor="FFEB9C")))
    ws.conditional_formatting.add("AB2:AB41", FormulaRule(formula=['AB2="ERROR"'], fill=PatternFill("solid", fgColor="FFC7CE")))


def add_market_model(ws):
    ws["A1"] = "Market Model Inputs"
    ws["A1"].font = Font(bold=True)
    params = [
        ("CurrentPrice", 1.1000), ("ATR(14)", 80), ("ATR Multiplier", 2.0), ("Spread", 12), ("Swap", -3),
        ("Commission", 7), ("Daily Volatility", 1.2), ("Weekly Volatility", 3.1), ("Trend Strength", 60),
        ("Volatility Regime", ""), ("Symbol Digits", 5), ("Point Value", 10), ("Contract Size", 100000),
    ]
    for i, (k, v) in enumerate(params, 2):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
    ws["B11"] = '=IF(B3<60,"LOW_VOL",IF(B3<120,"NORMAL",IF(B3<200,"HIGH_VOL","EXTREME")))'
    ws["A16"] = "Adaptive Step"
    ws["B16"] = "=B3*B4"


def add_adaptive_engine(ws):
    ws["A1"] = "Adaptive Engine"
    ws["A1"].font = Font(bold=True)
    fields = [
        "CurrentSkew", "LockWidth", "DistanceFromAverage", "MarginLoad", "RecoveryPressure",
        "TrendRisk", "RollbackProbability", "DD%", "DecayK", "BaseBig", "BaseSmall", "RiskFactor",
    ]
    defaults = [10, 300, 120, 40, 35, 55, 45, 18, 0.25, 90, 30, 0.85]
    for i, (f, d) in enumerate(zip(fields, defaults), 2):
        ws[f"A{i}"] = f
        ws[f"B{i}"] = d
    ws["A16"] = "Level"
    ws["B16"] = "DynamicBig%"
    ws["C16"] = "DynamicSmall%"
    ws["D16"] = "DynamicTargetSkew%"
    for r in range(17, 37):
        ws[f"A{r}"] = r - 16
        ws[f"B{r}"] = f"=MAX(1,$B$11*EXP(-$B$10*A{r})*$B$12)"
        ws[f"C{r}"] = f"=MAX(1,$B$12*($B$6/50)*MAX(0.5,1+$B$2/100))"
        ws[f"D{r}"] = f"=MIN(30,MAX(5,5+$B$6/5+$B$4/10))"


def add_margin_control(ws):
    ws["A1"] = "Margin Control"
    ws["A1"].font = Font(bold=True)
    rows = [
        ("Balance", 10000), ("Equity", 9400), ("FreeMargin", 7000), ("Leverage", 100), ("MarginPerLot", 1000),
        ("MarginCallLevel", 80), ("StopOutLevel", 50), ("UsedMargin", 2400), ("NextLevelLot", 0.20),
    ]
    for i, (k, v) in enumerate(rows, 2):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
    ws["A12"] = "RequiredMargin"
    ws["B12"] = "=B9*B5"
    ws["A13"] = "MarginLoad%"
    ws["B13"] = "=B8/B3*100"
    ws["A14"] = "FreeMarginAfterNextLevel"
    ws["B14"] = "=B4-B12"
    ws["A15"] = "WorstCaseMargin"
    ws["B15"] = "=B8+B12"
    ws["A16"] = "MarginStress"
    ws["B16"] = '=IF(B13<30,"SAFE",IF(B13<50,"NORMAL",IF(B13<70,"WARNING",IF(B13<90,"DANGER","CRITICAL"))))'
    ws["A17"] = "MaxSafeLevels"
    ws["B17"] = "=MAX(1,INT(B4/B12))"


def add_simple_table(ws, title, headers, rows):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h).font = Font(bold=True)
    for r, row in enumerate(rows, 4):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)


def add_summary_and_dashboard(wb):
    sm = wb["Summary"]
    sm["A1"] = "V3 Dashboard Summary"
    sm["A1"].font = Font(bold=True)
    pairs = [
        ("StartLot", "=Settings!B2"), ("Direction", "=Settings!B10"),
        ("Final Total BUY %", "=INDEX(DownTrend!Q:Q,Settings!B4+3)"), ("Final Total SELL %", "=INDEX(DownTrend!R:R,Settings!B4+3)"),
        ("Final Skew %", "=INDEX(DownTrend!S:S,Settings!B4+3)"), ("Final id1 Remaining %", "=INDEX(DownTrend!N:N,Settings!B4+3)"),
        ("Adaptive Step", "=MarketModel!B16"), ("ATR Regime", "=MarketModel!B11"), ("Margin Load", "=MarginControl!B13"),
        ("Risk Score", "=RiskDashboard!B4"), ("Risk Status", "=RiskDashboard!B5"), ("Survival Probability", "=RiskDashboard!B7"),
        ("Final System Status", '=IF(COUNTIF(DownTrend!T3:INDEX(DownTrend!T:T,Settings!B4+3),"ERROR*")+COUNTIF(UpTrend!T3:INDEX(UpTrend!T:T,Settings!B4+3),"ERROR*")>0,"ERROR",IF(COUNTIF(DownTrend!T3:INDEX(DownTrend!T:T,Settings!B4+3),"WARNING*")+COUNTIF(UpTrend!T3:INDEX(UpTrend!T:T,Settings!B4+3),"WARNING*")>0,"WARNING","OK"))'),
        ("Rounded System Status", '=IF(COUNTIF(DownTrend!AJ3:INDEX(DownTrend!AJ:AJ,Settings!B4+3),"ERROR*")+COUNTIF(UpTrend!AJ3:INDEX(UpTrend!AJ:AJ,Settings!B4+3),"ERROR*")>0,"ERROR",IF(COUNTIF(DownTrend!AJ3:INDEX(DownTrend!AJ:AJ,Settings!B4+3),"WARNING*")+COUNTIF(UpTrend!AJ3:INDEX(UpTrend!AJ:AJ,Settings!B4+3),"WARNING*")>0,"WARNING","OK"))'),
    ]
    for i, (k, f) in enumerate(pairs, 3):
        sm[f"A{i}"] = k
        sm[f"B{i}"] = f

    chart_specs = [
        ("Equity Curve", "EquityModel", 2, 3), ("Floating DD", "EquityModel", 4, 4), ("Margin Load", "MarginControl", 13, 13),
        ("Skew Evolution", "AdaptiveEngine", 4, 4), ("Compression Speed", "RecoveryMap", 4, 4),
        ("Recovery Probability", "RiskDashboard", 5, 5), ("ATR Regime Step", "MarketModel", 16, 16),
        ("Risk Score", "RiskDashboard", 2, 2), ("Monte Carlo", "MonteCarlo", 3, 3), ("BreakEven Map", "RecoveryMap", 2, 2),
    ]
    for i, (title, sh, c1, c2) in enumerate(chart_specs):
        ch = LineChart()
        ch.title = title
        data = Reference(wb[sh], min_col=c1, min_row=3, max_col=c2, max_row=10)
        ch.add_data(data, titles_from_data=False)
        ch.set_categories(Reference(wb[sh], min_col=1, min_row=4, max_row=10))
        sm.add_chart(ch, f"D{1 + i * 12}")


def add_checks(ws):
    ws["A1"] = "Checks"
    ws["A1"].font = Font(bold=True)
    checks = [
        ("Invalid StartLot", '=IF(OR(NOT(ISNUMBER(Settings!B2)),Settings!B2<=0),"ERROR","OK")'),
        ("Invalid LotStep", '=IF(OR(NOT(ISNUMBER(Settings!B5)),Settings!B5<=0),"ERROR","OK")'),
        ("Invalid Direction", '=IF(AND(Settings!B10<>"DOWN",Settings!B10<>"UP"),"ERROR","OK")'),
        ("Negative values", '=IF(OR(MIN(Settings!B22:E200)<0,MIN(DownTrend!N3:N41)<0,MIN(UpTrend!N3:N41)<0),"ERROR","OK")'),
        ("Big < Small", '=IF(SUMPRODUCT(--(DownTrend!C3:C41<DownTrend!E3:E41))+SUMPRODUCT(--(UpTrend!C3:C41<UpTrend!E3:E41))>0,"ERROR","OK")'),
        ("ManualClose > Remaining", '=IF(OR(SUMPRODUCT(--(DownTrend!L3:L41>DownTrend!G3:G41))>0,SUMPRODUCT(--(UpTrend!L3:L41>UpTrend!G3:G41))>0),"ERROR","OK")'),
        ("Remaining < 0", '=IF(OR(MIN(DownTrend!N3:N41)<0,MIN(UpTrend!N3:N41)<0),"ERROR","OK")'),
        ("Protection balance", '=IF(OR(SUMPRODUCT(--(DownTrend!Q3:Q41>DownTrend!R3:R41))>0,SUMPRODUCT(--(UpTrend!Q3:Q41>UpTrend!R3:R41))>0),"ERROR","OK")'),
        ("Rounded safety preserved", '=IF(OR(COUNTIF(DownTrend!AJ3:AJ41,"ERROR*")>0,COUNTIF(UpTrend!AJ3:AJ41,"ERROR*")>0),"ERROR","OK")'),
        ("Rounded Big zero", '=IF(OR(SUMPRODUCT(--(DownTrend!C3:C41>0),--(DownTrend!W3:W41=0))>0,SUMPRODUCT(--(UpTrend!C3:C41>0),--(UpTrend!W3:W41=0))>0),"ERROR","OK")'),
        ("Rounded Small zero", '=IF(OR(SUMPRODUCT(--(DownTrend!E3:E41>0),--(DownTrend!Y3:Y41=0))>0,SUMPRODUCT(--(UpTrend!E3:E41>0),--(UpTrend!Y3:Y41=0))>0),"ERROR","OK")'),
        ("LotStep too coarse", '=IF(Settings!B2<Settings!B5,"ERROR","OK")'),
        ("FinalStatus from T/AJ", '=IF(OR(COUNTIF(DownTrend!T3:T41,"ERROR*")>0,COUNTIF(UpTrend!T3:T41,"ERROR*")>0,COUNTIF(DownTrend!AJ3:AJ41,"ERROR*")>0,COUNTIF(UpTrend!AJ3:AJ41,"ERROR*")>0),"ERROR",IF(OR(COUNTIF(DownTrend!T3:T41,"WARNING*")>0,COUNTIF(UpTrend!T3:T41,"WARNING*")>0,COUNTIF(DownTrend!AJ3:AJ41,"WARNING*")>0,COUNTIF(UpTrend!AJ3:AJ41,"WARNING*")>0),"WARNING","OK"))'),
    ]
    for i, (k, f) in enumerate(checks, 2):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = f


def build_workbook(output_path: str) -> None:
    wb = Workbook()
    wb.active.title = BASE_SHEETS[0]
    for name in BASE_SHEETS[1:] + V3_SHEETS:
        wb.create_sheet(name)

    add_settings(wb["Settings"])
    add_named_ranges(wb)
    add_trend_sheet(wb["DownTrend"], True)
    add_trend_sheet(wb["UpTrend"], False)
    add_market_model(wb["MarketModel"])
    add_adaptive_engine(wb["AdaptiveEngine"])
    add_margin_control(wb["MarginControl"])
    add_simple_table(wb["MonteCarlo"], "Monte Carlo Scenarios", ["Scenario", "Max DD", "Max Levels", "Margin Stress", "Recovery Distance", "Required Rollback", "Survival"], [
        ("Strong Downtrend", "=25+AdaptiveEngine!B17/4", "=INT(5+AdaptiveEngine!B17/20)", "WARNING", "=500+AdaptiveEngine!B17*6", "=200+AdaptiveEngine!D17*8", "=IF(B4>60,\"LOW\",\"MEDIUM\")"), ("Strong Uptrend", 33, 8, "WARNING", 820, 400, "MEDIUM"),
        ("Volatile Trend", 45, 10, "DANGER", 1200, 620, "LOW"), ("Flash Crash", 60, 12, "CRITICAL", 1800, 900, "LOW"),
        ("Long Compression", 28, 7, "NORMAL", 640, 300, "HIGH"), ("Endless Trend", 72, 14, "CRITICAL", 2600, 1200, "VERY LOW"),
        ("Trend + Spike", 55, 11, "DANGER", 1500, 760, "LOW"),
    ])
    add_simple_table(wb["EquityModel"], "Equity Model", ["Step", "Balance", "Equity", "Floating DD", "Recovery", "Margin Load"], [
        (0, 10000, 10000, 0, 0, 24), (1, 10000, 9800, 2, 10, 28), (2, 10000, 9500, 5, 20, 35),
        (3, 10000, 9300, 7, 35, 42), (4, 10000, 9400, 6, 50, 39), (5, 10000, 9700, 3, 72, 32), (6, 10000, 9950, 1, 90, 26),
    ])
    add_simple_table(wb["RecoveryMap"], "Recovery Map", ["Level", "RecoveryDistance", "TimeBars", "CompressionCycles", "Basket>=0"], [
        (1, 120, 8, 1, "YES"), (2, 260, 16, 2, "YES"), (3, 430, 28, 3, "YES"), (4, 660, 45, 4, "NO"),
        (5, 980, 70, 6, "NO"), (6, 1300, 95, 8, "NO"),
    ])
    add_simple_table(wb["StressTest"], "Stress Test", ["Case", "Result", "Risk"], [
        ("ATR x2", "PASS", "NORMAL"), ("ATR x5", "WARN", "DANGER"), ("Spread x3", "WARN", "WARNING"),
        ("Gap 500", "WARN", "DANGER"), ("Flash move", "FAIL", "CRITICAL"), ("Infinite trend", "FAIL", "CRITICAL"),
        ("No rollback", "FAIL", "CRITICAL"),
    ])
    add_simple_table(wb["RiskDashboard"], "Risk Dashboard", ["Metric", "Value"], [
        ("RiskScore", "=MIN(100,MAX(0,0.35*EquityModel!D6+0.35*MarginControl!B13+0.2*AdaptiveEngine!B2+0.1*MarketModel!B3/2))"),
        ("RiskStatus", '=IF(B4<20,"SAFE",IF(B4<40,"NORMAL",IF(B4<60,"WARNING",IF(B4<80,"DANGER","CRITICAL"))))'),
        ("CurrentSkew", "=AdaptiveEngine!B2"),
        ("SurvivalProbability", "=MAX(0,100-B4)"),
        ("AdaptiveLevelStop", '=IF(OR(MarginControl!B13>70,EquityModel!D6>25,MarketModel!B3>200,AdaptiveEngine!D17>25),"STOP NEW LEVELS","ALLOW")'),
    ])

    add_summary_and_dashboard(wb)
    add_checks(wb["Checks"])
    wb["Manual"]["A1"] = "V3: Adaptive probabilistic recovery engine prototype with risk dashboard."
    wb.save(output_path)


if __name__ == "__main__":
    build_workbook("MinusLock_Percent_Grid_Calculator.xlsx")
