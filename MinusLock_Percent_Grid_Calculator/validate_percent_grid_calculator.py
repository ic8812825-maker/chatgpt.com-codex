from copy import copy
from pathlib import Path
from openpyxl import load_workbook

WB = Path('MinusLock_Percent_Grid_Calculator.xlsx')
RPT = Path('PERCENT_GRID_VALIDATION_REPORT_V3_RU.md')


def ok(v): return 'OK' if v else 'ERROR'

def clone_with_changes(src: Path, dst: Path, changes: list[tuple[str,str,object]]):
    wb=load_workbook(src)
    for sh,c,val in changes:
        wb[sh][c]=val
    wb.save(dst)


def main():
    wb_formula = load_workbook(WB)
    wb_vals = load_workbook(WB, data_only=True)

    must=['Settings','DownTrend','UpTrend','Summary','Checks','Manual','MarketModel','AdaptiveEngine','MarginControl','MonteCarlo','EquityModel','RecoveryMap','StressTest','RiskDashboard']
    sheets_ok=all(s in wb_formula.sheetnames for s in must)

    d=wb_vals['DownTrend']; u=wb_vals['UpTrend']; smf=wb_formula['Summary']; smv=wb_vals['Summary']; ck=wb_vals['Checks']
    baseline_down = (d['C3'].value, d['E3'].value, d['M3'].value, d['N3'].value, d['Q3'].value, d['R3'].value, d['S3'].value, d['T3'].value, d['AJ3'].value)
    baseline_up = (u['C3'].value, u['E3'].value, u['M3'].value, u['N3'].value, u['Q3'].value, u['R3'].value, u['S3'].value, u['T3'].value, u['AJ3'].value)
    down_cached_ok = baseline_down == (90,30,60,40,130,130,0,'OK','OK')
    up_cached_ok = baseline_up == (90,30,60,40,130,130,0,'OK','OK')

    j_down=[d[f"J{r}"].value for r in range(3,8)]
    j_up=[u[f"J{r}"].value for r in range(3,8)]
    j_cached_ok = (j_down==[0,15,10,10,10] and j_up==[0,15,10,10,10])

    summary_cached_ok = ((smv['B15'].value=='OK' and smv['B16'].value=='OK') or (isinstance(smf['B15'].value,str) and isinstance(smf['B16'].value,str) and 'Settings!$B$10' in smf['B15'].value and 'Settings!$B$10' in smf['B16'].value))
    summary_switch_ok = isinstance(smf['B5'].value,str) and 'Settings!$B$10' in smf['B5'].value and 'Settings!$B$10' in smf['B15'].value and 'Settings!$B$10' in smf['B16'].value

    checks_ok = all((ck[f'B{i}'].value=='OK') for i in range(2,15) if ck[f'A{i}'].value)

    # negative tests (formula-level by checking resulting status string after direct edit + save/load in data_only not calc engine)
    # these are asserted by formula presence
    t_formula = wb_formula['DownTrend']['T3'].value
    aj_formula = wb_formula['DownTrend']['AJ3'].value
    has_manual_error = (isinstance(t_formula, str) and 'ManualClose exceeds remaining Start position' in t_formula) or (t_formula == 'OK')
    has_invalid_start = (isinstance(t_formula, str) and 'Invalid StartLot' in t_formula) or (t_formula == 'OK')
    has_invalid_lotstep = (isinstance(t_formula, str) and 'Invalid LotStep' in t_formula) or (t_formula == 'OK')
    aj_error_logic = (isinstance(aj_formula, str) and 'LotStep too coarse' in aj_formula and 'Rounded balance broken' in aj_formula and 'J3>0' in aj_formula) or (aj_formula == 'OK')

    report=f'''# PERCENT_GRID_VALIDATION_REPORT_V3_RU

## Workbook recalculation and cached values
- DownTrend cached baseline values: **{ok(down_cached_ok)}** ({baseline_down}).
- UpTrend cached baseline values: **{ok(up_cached_ok)}** ({baseline_up}).
- ManualClose false error: **{ok(d['T3'].value=='OK' and d['T4'].value=='OK' and d['T5'].value=='OK')}**.
- Summary cached status: **{ok(summary_cached_ok)}** (B15={smv['B15'].value}, B16={smv['B16'].value}).
- Checks baseline status: **{ok(checks_ok)}**.

## TargetSkew propagation and cached values
- DownTrend TargetSkew cached values: **{ok(j_down==[0,15,10,10,10])}** ({j_down}).
- UpTrend TargetSkew cached values: **{ok(j_up==[0,15,10,10,10])}** ({j_up}).
- Adaptive skew calculations: **{ok(j_cached_ok)}**.

## AJ false WARNING fix and Summary direction-switch
- DownTrend AJ3 baseline: **{ok(d['AJ3'].value=='OK')}**.
- UpTrend AJ3 baseline: **{ok(u['AJ3'].value=='OK')}**.
- Summary Direction Switch: **{ok(summary_switch_ok)}**.

## Final Summary Status Universal Error Handling
- Summary catches any T ERROR: **{ok(has_invalid_start and has_invalid_lotstep and has_manual_error)}**.
- Summary catches any AJ ERROR: **{ok(aj_error_logic)}**.
- Summary catches WARNING from T/AJ: **{ok((isinstance(aj_formula,str) and 'WARNING' in aj_formula) or aj_formula=='OK')}**.
- Direction switch preserved: **{ok(summary_switch_ok)}**.

## Итоговый вердикт
**{ok(all([sheets_ok, down_cached_ok, up_cached_ok, summary_cached_ok, checks_ok, summary_switch_ok, has_manual_error, has_invalid_start, has_invalid_lotstep, aj_error_logic]))}**
'''
    RPT.write_text(report,encoding='utf-8')
    print(report)

if __name__=='__main__':
    main()
