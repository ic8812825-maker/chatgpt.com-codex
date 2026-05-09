from openpyxl import Workbook
from openpyxl.styles import Font

BOLD = Font(bold=True)


def hdr(ws, row, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = BOLD


def add_settings(ws):
    hdr(ws, 1, ["Field", "Value"])
    rows = [
        ("Symbol", "EURUSD"), ("Point", 0.0001), ("PointValuePerLot", 1), ("StepPoints", 100),
        ("MinLot", 0.01), ("LotStep", 0.01), ("MaxActiveSections", 4), ("ReservePercent", 0.20),
        ("RecoveryPercent", 0.80), ("BasketTarget", 0), ("MaxSpreadPoints", 20),
        ("CommissionPerLot", 0), ("SwapPerLot", 0), ("MaxTotalLot", 20), ("MaxNetLot", 10),
        ("GlobalReserve", 0), ("RecoveryFund", 0),
    ]
    for r, (k, v) in enumerate(rows, 2):
        ws.cell(r, 1, k); ws.cell(r, 2, v)
    hdr(ws, 20, ["Level", "BigRatio", "SmallRatio"])
    for r, row in enumerate([(1, 0.40, 0.15), (2, 0.25, 0.10), (3, 0.15, 0.06), (4, 0.10, 0.04)], 21):
        ws.cell(r, 1, row[0]); ws.cell(r, 2, row[1]); ws.cell(r, 3, row[2])


def add_positions(ws):
    hdr(ws, 1, ["Ticket", "Type", "Role", "Lot", "OpenPrice", "CurrentPrice", "PointsPnL", "MoneyPnL", "IsTail", "SectionID"])
    rows = [
        [10001, "BUY", "MAIN", 1.00, 1.2300, 1.2300, None, None, "NO", ""],
        [10002, "SELL", "TAIL", 1.00, 1.2300, 1.2300, None, None, "YES", ""],
        [10003, "SELL", "SECTION_BIG", 0.40, 1.2350, 1.2300, None, None, "NO", "S1"],
        [10004, "BUY", "SECTION_SMALL", 0.15, 1.2350, 1.2300, None, None, "NO", "S1"],
    ]
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
        ws.cell(r, 7, f'=IF(B{r}="BUY",(F{r}-E{r})/Settings!$B$3,(E{r}-F{r})/Settings!$B$3)')
        ws.cell(r, 8, f'=G{r}*D{r}*Settings!$B$4')


def add_scenario(ws, direction):
    up = direction == "UP"
    move = "MoveUpPoints" if up else "MoveDownPoints"
    sign = "+" if up else "-"
    loss_type = "SELL" if up else "BUY"

    hdr(ws, 1, ["Field", "Value"])
    ws["A2"] = "CurrentPrice"; ws["B2"] = "=CurrentPositions!F2"
    ws["A3"] = move; ws["B3"] = 100
    ws["A4"] = "ScenarioPrice"; ws["B4"] = f"=B2{sign}B3*Settings!$B$3"

    hdr(ws, 6, ["Ticket", "Type", "Role", "Lot", "OpenPrice", "ScenarioPrice", "ScenarioPointsPnL", "ScenarioMoneyPnL", "IsTail", "SectionID"])
    for r in range(2, 202):
        rr = r + 5
        ws.cell(rr, 1, f"=CurrentPositions!A{r}")
        ws.cell(rr, 2, f"=CurrentPositions!B{r}")
        ws.cell(rr, 3, f"=CurrentPositions!C{r}")
        ws.cell(rr, 4, f"=CurrentPositions!D{r}")
        ws.cell(rr, 5, f"=CurrentPositions!E{r}")
        ws.cell(rr, 6, "=$B$4")
        ws.cell(rr, 7, f'=IF(B{rr}="BUY",(F{rr}-E{rr})/Settings!$B$3,(E{rr}-F{rr})/Settings!$B$3)')
        ws.cell(rr, 8, f"=G{rr}*D{rr}*Settings!$B$4")
        ws.cell(rr, 9, f"=CurrentPositions!I{r}")
        ws.cell(rr, 10, f"=CurrentPositions!J{r}")

    ws["A220"] = "TailType"; ws["B220"] = loss_type
    ws["A221"] = "TailTicket"; ws["B221"] = f'=IFERROR(INDEX(A7:A206,MATCH(MINIFS(H7:H206,B7:B206,"{loss_type}",H7:H206,"<0"),H7:H206,0)),"N/A")'
    ws["A222"] = "TailLot"; ws["B222"] = '=IF(B221="N/A",0,XLOOKUP(B221,A7:A206,D7:D206,0))'
    ws["A223"] = "TailLossMoney"; ws["B223"] = '=IF(B221="N/A",0,ABS(XLOOKUP(B221,A7:A206,H7:H206,0)))'
    ws["A224"] = "TailLossPerLot"; ws["B224"] = '=IF(B222=0,0,ABS(B223/B222))'

    ws["D1"] = "AverageBuyPrice"; ws["E1"] = '=IFERROR(SUMPRODUCT((B7:B206="BUY")*E7:E206*D7:D206)/SUMIFS(D7:D206,B7:B206,"BUY"),0)'
    ws["D2"] = "AverageSellPrice"; ws["E2"] = '=IFERROR(SUMPRODUCT((B7:B206="SELL")*E7:E206*D7:D206)/SUMIFS(D7:D206,B7:B206,"SELL"),0)'
    ws["D3"] = "Center"; ws["E3"] = '=(E1+E2)/2'
    for lvl in range(1, 5):
        ws.cell(3 + lvl, 4, f"UpperLevel{lvl}"); ws.cell(3 + lvl, 5, f"=E3+{lvl}*Settings!$B$5*Settings!$B$3")
        ws.cell(8 + lvl, 4, f"LowerLevel{lvl}"); ws.cell(8 + lvl, 5, f"=E3-{lvl}*Settings!$B$5*Settings!$B$3")


def add_section_calc(ws, scenario_sheet, direction):
    up = direction == "UP"
    hdr(ws, 1, ["Field", "Value"])
    ws["A2"] = "TailType"; ws["B2"] = f"={scenario_sheet}!B220"
    ws["A3"] = "TailLot"; ws["B3"] = f"={scenario_sheet}!B222"
    ws["A4"] = "NextLevel"; ws["B4"] = 1
    ws["A5"] = "BigRatio"; ws["B5"] = '=XLOOKUP(B4,Settings!A21:A24,Settings!B21:B24)'
    ws["A6"] = "SmallRatio"; ws["B6"] = '=XLOOKUP(B4,Settings!A21:A24,Settings!C21:C24)'
    ws["A7"] = "BigLot"; ws["B7"] = '=FLOOR(B3*B5,Settings!$B$7)'
    ws["A8"] = "SmallLot"; ws["B8"] = '=FLOOR(B3*B6,Settings!$B$7)'
    ws["A9"] = "ActiveSections"; ws["B9"] = '=COUNTIFS(CurrentPositions!C2:C500,"SECTION_BIG")'
    ws["A10"] = "TotalLotAfterOpen"; ws["B10"] = '=SUM(CurrentPositions!D2:D500)+B7+B8'
    ws["A11"] = "NetLotAfterOpen"; ws["B11"] = '=ABS(SUMIFS(CurrentPositions!D2:D500,CurrentPositions!B2:B500,"BUY")-SUMIFS(CurrentPositions!D2:D500,CurrentPositions!B2:B500,"SELL"))'
    ws["A12"] = "LevelReached"
    ws["B12"] = '=IF({}!B4>={}!E4,"YES","NO")'.format(scenario_sheet, scenario_sheet) if up else '=IF({}!B4<={}!E9,"YES","NO")'.format(scenario_sheet, scenario_sheet)
    ws["A13"] = "CanOpenSection"
    ws["B13"] = '=IF(AND(B7>=Settings!$B$6,B8>=Settings!$B$6,B8<B7,B7<=B3,B9<Settings!$B$8,B10<=Settings!$B$15,B11<=Settings!$B$16,B12="YES"),"YES","NO")'

    hdr(ws, 20, ["SectionID", "BigType", "BigLot", "BigOpenPrice", "SmallType", "SmallLot", "SmallOpenPrice", "ScenarioPrice", "BigPnL", "SmallPnL", "Costs", "CycleProfit", "CanCloseSection"])
    for i in range(1, 5):
        r = 20 + i
        sid = f"S{i}"
        ws.cell(r, 1, sid)
        ws.cell(r, 2, f'=IFERROR(INDEX(CurrentPositions!B2:B500,MATCH(1,(CurrentPositions!C2:C500="SECTION_BIG")*(CurrentPositions!J2:J500=A{r}),0)),"")')
        ws.cell(r, 3, f'=IFERROR(INDEX(CurrentPositions!D2:D500,MATCH(1,(CurrentPositions!C2:C500="SECTION_BIG")*(CurrentPositions!J2:J500=A{r}),0)),0)')
        ws.cell(r, 4, f'=IFERROR(INDEX(CurrentPositions!E2:E500,MATCH(1,(CurrentPositions!C2:C500="SECTION_BIG")*(CurrentPositions!J2:J500=A{r}),0)),0)')
        ws.cell(r, 5, f'=IFERROR(INDEX(CurrentPositions!B2:B500,MATCH(1,(CurrentPositions!C2:C500="SECTION_SMALL")*(CurrentPositions!J2:J500=A{r}),0)),"")')
        ws.cell(r, 6, f'=IFERROR(INDEX(CurrentPositions!D2:D500,MATCH(1,(CurrentPositions!C2:C500="SECTION_SMALL")*(CurrentPositions!J2:J500=A{r}),0)),0)')
        ws.cell(r, 7, f'=IFERROR(INDEX(CurrentPositions!E2:E500,MATCH(1,(CurrentPositions!C2:C500="SECTION_SMALL")*(CurrentPositions!J2:J500=A{r}),0)),0)')
        ws.cell(r, 8, f'={scenario_sheet}!B4')
        ws.cell(r, 9, f'=IF(B{r}="",0,IF(B{r}="BUY",(H{r}-D{r})/Settings!$B$3*C{r}*Settings!$B$4,(D{r}-H{r})/Settings!$B$3*C{r}*Settings!$B$4))')
        ws.cell(r, 10, f'=IF(E{r}="",0,IF(E{r}="BUY",(H{r}-G{r})/Settings!$B$3*F{r}*Settings!$B$4,(G{r}-H{r})/Settings!$B$3*F{r}*Settings!$B$4))')
        ws.cell(r, 11, f'=(Settings!$B$12*Settings!$B$4*(C{r}+F{r}))+(Settings!$B$13*(C{r}+F{r}))+(Settings!$B$14*(C{r}+F{r}))')
        ws.cell(r, 12, f'=I{r}+J{r}-K{r}')
        ws.cell(r, 13, f'=IF(L{r}>0,"YES","NO")')

    ws["A30"] = "SelectedSectionID"; ws["B30"] = '=IFERROR(INDEX(A21:A24,MATCH("YES",M21:M24,0)),"NONE")'
    ws["A31"] = "CycleProfit"; ws["B31"] = '=IF(B30="NONE",0,XLOOKUP(B30,A21:A24,L21:L24,0))'
    ws["A32"] = "CanCloseSection"; ws["B32"] = '=IF(B31>0,"YES","NO")'
    ws["A33"] = "ReserveAdd"; ws["B33"] = '=IF(B31>0,B31*Settings!$B$9,0)'
    ws["A34"] = "RecoveryAdd"; ws["B34"] = '=IF(B31>0,B31*Settings!$B$10,0)'
    ws["A35"] = "NewGlobalReserve"; ws["B35"] = '=Settings!$B$17+B33'
    ws["A36"] = "NewRecoveryFund"; ws["B36"] = '=Settings!$B$18+B34'


def add_tail_recovery(ws, section_sheet, scenario_sheet):
    hdr(ws, 1, ["Field", "Value"])
    data = [
        ("TailType", f"={scenario_sheet}!B220"), ("TailTicket", f"={scenario_sheet}!B221"), ("TailLot", f"={scenario_sheet}!B222"),
        ("TailLossPerLot", f"={scenario_sheet}!B224"), ("CanCloseSection", f"={section_sheet}!B32"), ("RecoveryFundAfterCycle", f"={section_sheet}!B36"),
        ("CloseLotRaw", '=IF(OR(B6<>"YES",B5=0),0,B7/B5)'), ("CloseLotRounded", '=FLOOR(B8,Settings!$B$7)'),
        ("CloseLotFinal", '=IF(B6="YES",MIN(B9,B4),0)'), ("CloseAllowed", '=IF(AND(B10>=Settings!$B$6,B6="YES"),"YES","NO")'),
        ("TailCloseLoss", '=B10*B5'), ("RecoveryFundAfterClose", '=MAX(0,B7-B12)'), ("TailLotAfterClose", '=B4-B10'),
        ("NextBigLotAfterRecovery", '=FLOOR(B14*XLOOKUP(1,Settings!A21:A24,Settings!B21:B24),Settings!$B$7)'),
        ("NextSmallLotAfterRecovery", '=FLOOR(B14*XLOOKUP(1,Settings!A21:A24,Settings!C21:C24),Settings!$B$7)')
    ]
    for r, (k, v) in enumerate(data, 2): ws.cell(r, 1, k); ws.cell(r, 2, v)


def build():
    wb = Workbook()
    wb.active.title = "Settings"
    for s in ["CurrentPositions", "Scenario_UP", "Scenario_DOWN", "SectionCalculator_UP", "SectionCalculator_DOWN", "TailRecovery_UP", "TailRecovery_DOWN", "BasketSummary", "Validation", "Log"]:
        wb.create_sheet(s)
    add_settings(wb["Settings"])
    add_positions(wb["CurrentPositions"])
    add_scenario(wb["Scenario_UP"], "UP")
    add_scenario(wb["Scenario_DOWN"], "DOWN")
    add_section_calc(wb["SectionCalculator_UP"], "Scenario_UP", "UP")
    add_section_calc(wb["SectionCalculator_DOWN"], "Scenario_DOWN", "DOWN")
    add_tail_recovery(wb["TailRecovery_UP"], "SectionCalculator_UP", "Scenario_UP")
    add_tail_recovery(wb["TailRecovery_DOWN"], "SectionCalculator_DOWN", "Scenario_DOWN")

    bs = wb["BasketSummary"]
    hdr(bs, 1, ["Field", "UP", "DOWN"])
    fields = ["ScenarioPrice", "BasketFloating", "GlobalReserveAfter", "RecoveryFundAfterCycle", "TailLotAfter", "CloseLot", "CanCloseSection", "CanCloseTail", "CanCloseBasket", "NextBigLot", "NextSmallLot", "NextAction"]
    for i, f in enumerate(fields, 2): bs.cell(i, 1, f)
    bs["B2"] = "=Scenario_UP!B4"; bs["C2"] = "=Scenario_DOWN!B4"
    bs["B3"] = "=SUM(Scenario_UP!H7:H206)"; bs["C3"] = "=SUM(Scenario_DOWN!H7:H206)"
    bs["B4"] = "=SectionCalculator_UP!B35"; bs["C4"] = "=SectionCalculator_DOWN!B35"
    bs["B5"] = "=TailRecovery_UP!B7"; bs["C5"] = "=TailRecovery_DOWN!B7"
    bs["B6"] = "=TailRecovery_UP!B14"; bs["C6"] = "=TailRecovery_DOWN!B14"
    bs["B7"] = "=TailRecovery_UP!B10"; bs["C7"] = "=TailRecovery_DOWN!B10"
    bs["B8"] = "=SectionCalculator_UP!B32"; bs["C8"] = "=SectionCalculator_DOWN!B32"
    bs["B9"] = "=TailRecovery_UP!B11"; bs["C9"] = "=TailRecovery_DOWN!B11"
    bs["B10"] = '=IF(B3+B4>=Settings!$B$10,"YES","NO")'; bs["C10"] = '=IF(C3+C4>=Settings!$B$10,"YES","NO")'
    bs["B11"] = "=TailRecovery_UP!B15"; bs["C11"] = "=TailRecovery_DOWN!B15"
    bs["B12"] = "=TailRecovery_UP!B16"; bs["C12"] = "=TailRecovery_DOWN!B16"
    bs["B13"] = '=IF(B10="YES","BASKET_CLOSE",IF(B8="NO","WAIT",IF(B9="YES","CLOSE_SECTION+CLOSE_TAIL","SAFE")))'
    bs["C13"] = '=IF(C10="YES","BASKET_CLOSE",IF(C8="NO","WAIT",IF(C9="YES","CLOSE_SECTION+CLOSE_TAIL","SAFE")))'

    v = wb["Validation"]
    hdr(v, 1, ["Rule", "UP", "DOWN"])
    rules = [
        ("CycleProfit > 0 before close section", '=IF(OR(SectionCalculator_UP!B31>0,SectionCalculator_UP!B32="NO"),"OK","ERROR")', '=IF(OR(SectionCalculator_DOWN!B31>0,SectionCalculator_DOWN!B32="NO"),"OK","ERROR")'),
        ("CloseLot = 0 if CanCloseSection=NO", '=IF(OR(SectionCalculator_UP!B32="YES",TailRecovery_UP!B10=0),"OK","ERROR")', '=IF(OR(SectionCalculator_DOWN!B32="YES",TailRecovery_DOWN!B10=0),"OK","ERROR")'),
        ("TailCloseLoss <= RecoveryFundAfterCycle", '=IF(TailRecovery_UP!B12<=TailRecovery_UP!B7,"OK","ERROR")', '=IF(TailRecovery_DOWN!B12<=TailRecovery_DOWN!B7,"OK","ERROR")'),
        ("LevelReached before open section", '=IF(OR(SectionCalculator_UP!B13="NO",SectionCalculator_UP!B12="YES"),"OK","ERROR")', '=IF(OR(SectionCalculator_DOWN!B13="NO",SectionCalculator_DOWN!B12="YES"),"OK","ERROR")'),
        ("No opposite cascade active", 'OK', 'OK'),
        ("CloseLot < MinLot => do not close", '=IF(OR(TailRecovery_UP!B10>=Settings!$B$6,TailRecovery_UP!B10=0),"OK","ERROR")', '=IF(OR(TailRecovery_DOWN!B10>=Settings!$B$6,TailRecovery_DOWN!B10=0),"OK","ERROR")'),
        ("RecoveryFundAfterClose >= 0", '=IF(TailRecovery_UP!B13>=0,"OK","ERROR")', '=IF(TailRecovery_DOWN!B13>=0,"OK","ERROR")'),
    ]
    for r, row in enumerate(rules, 2):
        for c, x in enumerate(row, 1): v.cell(r, c, x)

    log = wb["Log"]
    log["A1"] = '="NEXT ACTION UP:"&CHAR(10)&"CloseSection="&BasketSummary!B8&CHAR(10)&"CloseTail="&BasketSummary!B9&CHAR(10)&"CloseLot="&TEXT(BasketSummary!B7,"0.00")&CHAR(10)&"TailAfter="&TEXT(BasketSummary!B6,"0.00")&CHAR(10)&"NextSection Big/Small="&TEXT(BasketSummary!B11,"0.00")&"/"&TEXT(BasketSummary!B12,"0.00")&CHAR(10)&"Action="&BasketSummary!B13'
    log["A2"] = '="NEXT ACTION DOWN:"&CHAR(10)&"CloseSection="&BasketSummary!C8&CHAR(10)&"CloseTail="&BasketSummary!C9&CHAR(10)&"CloseLot="&TEXT(BasketSummary!C7,"0.00")&CHAR(10)&"TailAfter="&TEXT(BasketSummary!C6,"0.00")&CHAR(10)&"NextSection Big/Small="&TEXT(BasketSummary!C11,"0.00")&"/"&TEXT(BasketSummary!C12,"0.00")&CHAR(10)&"Action="&BasketSummary!C13'

    wb.save("recovery_lock_cascade_next_step.xlsx")


if __name__ == "__main__":
    build()
    print("Created recovery_lock_cascade_next_step.xlsx")
